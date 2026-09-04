"""
FastAPI Backend Server.
Defines endpoints for analyzing PDFs, extracting text, matching priced items,
and running compliance validation audits.
"""
from fastapi import FastAPI, File, UploadFile, HTTPException, Form, BackgroundTasks
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Dict, Any, Optional
import os
import re
import base64
import shutil
import tempfile
import fitz
import json

from pdf_engine import PDFEngine
from core.document_extractor import extract_document_elements
from services.matcher import (
    load_master_price_list,
    save_master_price_list,
    match_item_to_price_list,
    clear_user_mappings,
    update_price_item_in_excel,
    clear_price_item_in_excel,
    add_price_item_to_excel,
    generate_populated_boq_excel,
    write_cell_value_to_excel,
    clear_column_values_in_excel,
    clear_price_items_in_excel_batch,
    clear_all_price_items_in_excel,
    get_sheet_column_widths,
    get_excel_layout_metadata,
    get_price_list_path,
    sync_db_to_active_excel,
    price_list_locks,
    get_id_from_path
)
from services.ai_service import run_gemini_boq_deduplicator, run_gemini_boq_mapper_and_deduplicator, run_gemini_recheck_generator

ACTIVE_PRICE_LIST_PATH = os.path.join(os.path.dirname(__file__), "uploads", "active_price_list.xlsx")
# Startup setup
os.makedirs(os.path.dirname(ACTIVE_PRICE_LIST_PATH), exist_ok=True)

app = FastAPI(
    title="BOQ Automation PDF Engine API",
    description="Backend API powered by PyMuPDF and Gemini Vision to parse drawing BOQs",
    version="1.0.0"
)

# Enable CORS for frontend interaction
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from routers import pdf_router, project_router
app.include_router(pdf_router)
app.include_router(project_router)

from services.db import init_db, get_default_price_list_id

@app.on_event("startup")
def on_startup():
    init_db()

@app.get("/")
def read_root() -> Dict[str, str]:
    """Root health check endpoint."""
    return {"status": "running", "engine": "PyMuPDF (fitz) + Gemini Vision Scanner"}

@app.post("/api/analyze-pdf")
async def analyze_pdf(file: UploadFile = File(...)) -> Dict[str, Any]:
    """Uploads a PDF, runs page classification, extracts cloud notes, and extracts tables."""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Uploaded file must be a PDF.")

    temp_dir = tempfile.mkdtemp()
    temp_file_path = os.path.join(temp_dir, file.filename)
    
    try:
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        engine = PDFEngine(temp_file_path)
        page_count = engine.get_page_count()
        pages_analysis = []
        
        for p in range(page_count):
            pages_analysis.append({
                "page_index": p,
                "metadata": engine.get_page_metadata(p),
                "category": engine.classify_page_category(p),
                "cloud_notes": engine.extract_cloud_notes(p),
                "tables": engine.extract_tables(p)
            })
            
        engine.close()
        return {
            "filename": file.filename,
            "page_count": page_count,
            "pages": pages_analysis
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to analyze PDF: {str(e)}")
    finally:
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        if os.path.exists(temp_dir):
            os.rmdir(temp_dir)

@app.post("/api/extract-page-text")
async def extract_page_text(file: UploadFile = File(...), page_num: int = Form(...)) -> Dict[str, Any]:
    """Extracts sorted, clean text from a specific page."""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Uploaded file must be a PDF.")

    temp_dir = tempfile.mkdtemp()
    temp_file_path = os.path.join(temp_dir, file.filename)
    
    try:
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        engine = PDFEngine(temp_file_path)
        if page_num < 0 or page_num >= engine.get_page_count():
            raise HTTPException(status_code=400, detail=f"Invalid page index {page_num}.")
            
        text = engine.extract_page_text(page_num)
        engine.close()
        return {"page_num": page_num + 1, "text": text}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        if os.path.exists(temp_dir):
            os.rmdir(temp_dir)

@app.post("/api/analyze-pdf-path")
async def analyze_pdf_path(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Accepts local PDF path, parses using generic vision scanning page-by-page, and maps items to SOR codes."""
    pdf_path = payload.get("path")
    price_list_id = payload.get("price_list_id")
    if not pdf_path:
        raise HTTPException(status_code=400, detail="Missing 'path' in payload.")
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail=f"PDF not found at: {pdf_path}")
        
    file_path = get_price_list_path(price_list_id)
    try:
        from core.document_extractor import extract_document_elements
        res = extract_document_elements(pdf_path)
        elements = res.get("elements", [])
        raw_items = res.get("raw_items", [])
        formatted_tables = res.get("extracted_tables", [])
        
        consolidated = [i for i in raw_items if i.get("action", "").upper() not in ["EXISTING", "REUSE"]]
        price_list = []
        if os.path.exists(file_path):
            price_list = load_master_price_list(file_path, price_list_id)
            
        mapped_boq_items = []
        for idx, item in enumerate(consolidated):
            if item.get("action", "").upper() in ["EXISTING", "REUSE"]:
                continue
            source_sheet = item.get("source_sheet", f"Sheet {item['page']}")
            mapped = match_item_to_price_list(item, price_list)
            
            mapped_boq_items.append({
                "item_id": f"boq_{idx:03d}",
                "equipment_type": item.get("equipment_type", "OTHER"),
                "model": item.get("model", ""),
                "action": item.get("action", "INSTALL"),
                "quantity": item.get("quantity", 1.0),
                "source_sheet": source_sheet,
                "raw_text": item.get("raw_text", ""),
                "sor_code": mapped.get("code", "UNMAPPED") if mapped else "UNMAPPED",
                "item_name": mapped.get("name", item.get("clean_text", item.get("raw_text", ""))) if mapped else item.get("clean_text", item.get("raw_text", "")),
                "unit": mapped.get("unit", "each") if mapped else "each",
                "rate": mapped.get("rate", 0.0) if mapped else 0.0,
                "total_cost": (mapped.get("rate", 0.0) * item.get("quantity", 1.0)) if mapped else 0.0,
                "similarity": mapped.get("similarity", 0.0) if mapped else 0.0,
                "auto_matched": mapped.get("auto_matched", False) if mapped else False,
                "row_idx": mapped.get("row_idx") if mapped else None
            })
            
        pdf_quantities = {}
        for b_item in mapped_boq_items:
            if b_item.get("row_idx"):
                idx_str = str(b_item["row_idx"])
                pdf_quantities[idx_str] = pdf_quantities.get(idx_str, 0) + b_item["quantity"]
                
        if pdf_quantities:
            # Update database quantities for this price list
            from services.db import get_db_connection
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE price_items SET quantity = 0 WHERE price_list_id = ?", (price_list_id or 1,))
            for row_idx_str, qty in pdf_quantities.items():
                cursor.execute("UPDATE price_items SET quantity = ? WHERE id = ?", (qty, int(row_idx_str)))
            conn.commit()
            conn.close()
            
            # Generate dynamically on disk
            generate_populated_boq_excel(file_path, {}, file_path)
            
        pdf_corpus_lines = []
        try:
            c_doc = fitz.open(pdf_path)
            for page in c_doc:
                pdf_corpus_lines.append(page.get_text())
            c_doc.close()
        except Exception:
            pass
        
        pdf_corpus = "\n".join(pdf_corpus_lines)
        validation_results = run_checklist_validation(consolidated, mapped_boq_items, pdf_corpus)
        
        return {
            "status": "success",
            "drawing_name": os.path.basename(pdf_path),
            "raw_count": len(raw_items),
            "consolidated_count": len(consolidated),
            "elements": elements,
            "extracted_tables": formatted_tables,
            "mapped_items": mapped_boq_items,
            "checklist": validation_results
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to analyze PDF path: {str(e)}")

@app.post("/api/extract-drawing-data")
async def extract_drawing_data(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Step 1: Extract all drawing info page-by-page.
    Saves raw tables to cache folder backend/extracted_tables.json.
    """
    pdf_path = payload.get("path")
    selected_pages = payload.get("selected_pages")
    if not pdf_path:
        raise HTTPException(status_code=400, detail="Missing 'path' in payload.")
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail=f"PDF not found at: {pdf_path}")
        
    try:
        from core.document_extractor import extract_document_elements
        res = extract_document_elements(pdf_path, selected_pages=selected_pages)
        elements = res.get("elements", [])
        
        # Populate raw items and formatted tables from elements
        from processors.parser import extract_raw_items_from_elements
        raw_items = extract_raw_items_from_elements(elements)
        
        formatted_tables = []
        for el in elements:
            if el.get("type") == "structured":
                content = el.get("content", {})
                formatted_tables.append({
                    "page": el.get("page", 1),
                    "table_type": "STRUCTURED_TABLE",
                    "table_title": el.get("title", "EXTRACTED TABLE"),
                    "headers": content.get("headers", []),
                    "rows": content.get("rows", []),
                    "bbox": el.get("bbox"),
                    "cells": [],
                    "ai_highlights": [],
                    "token_analytics": {}
                })
        
        # Cache results to json
        cache_path = os.path.join(os.path.dirname(__file__), "extracted_tables.json")
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump({
                "pdf_path": pdf_path,
                "raw_items": raw_items,
                "elements": elements,
                "extracted_tables": formatted_tables
            }, f, indent=2)

        return {
            "status": "success",
            "drawing_name": os.path.basename(pdf_path),
            "raw_count": len(raw_items),
            "elements": elements,
            "extracted_tables": formatted_tables
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to extract drawing data: {str(e)}")


@app.post("/api/reextract-page-data")
async def reextract_page_data(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Re-extracts details for a single specific page with corrective instructions,
    and updates the cached elements for that page.
    """
    pdf_path = payload.get("path")
    page_num = payload.get("page")
    if not pdf_path or page_num is None:
        raise HTTPException(status_code=400, detail="Missing 'path' or 'page' in payload.")
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail=f"PDF not found at: {pdf_path}")
        
    try:
        # Load currently cached elements to find existing data for this page
        cache_path = os.path.join(os.path.dirname(__file__), "extracted_tables.json")
        cached_data = {}
        if os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    cached_data = json.load(f)
            except Exception:
                pass
                
        all_elements = cached_data.get("elements", [])
        page_existing = [el for el in all_elements if el.get("page") == page_num]

        from core.document_extractor import reextract_single_page_elements
        new_page_elements = reextract_single_page_elements(pdf_path, int(page_num), page_existing)
        
        # Replace only the elements for the requested page in the list
        updated_elements = [el for el in all_elements if el.get("page") != page_num]
        updated_elements.extend(new_page_elements)
        updated_elements.sort(key=lambda x: (x.get("page", 1), x.get("type", "")))
        
        # Re-populate raw items and formatted tables from updated elements
        from processors.parser import extract_raw_items_from_elements
        updated_raw_items = extract_raw_items_from_elements(updated_elements)
        
        updated_formatted_tables = []
        for el in updated_elements:
            if el.get("type") == "structured":
                content = el.get("content", {})
                updated_formatted_tables.append({
                    "page": el.get("page", 1),
                    "table_type": "STRUCTURED_TABLE",
                    "table_title": el.get("title", "EXTRACTED TABLE"),
                    "headers": content.get("headers", []),
                    "rows": content.get("rows", []),
                    "bbox": el.get("bbox"),
                    "cells": [],
                    "ai_highlights": [],
                    "token_analytics": {}
                })
        
        # Update the cache file
        cached_data["elements"] = updated_elements
        cached_data["raw_items"] = updated_raw_items
        cached_data["extracted_tables"] = updated_formatted_tables
        cached_data["pdf_path"] = pdf_path
        
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(cached_data, f, indent=2)
            
        return {
            "status": "success",
            "drawing_name": os.path.basename(pdf_path),
            "elements": updated_elements,
            "extracted_tables": updated_formatted_tables
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to re-extract page details: {str(e)}")


@app.post("/api/generate-boq")
@app.post("/api/generate-boq-deduplicated")
async def generate_boq(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generates commercial Bill of Quantities (BOQ) takeoff directly from extracted drawing tables
    and annotations using intelligent AI and price catalog matching.
    """
    pdf_path = payload.get("path")
    name = payload.get("name")
    
    # Resolve valid PDF path from uploads if needed
    from core.config import UPLOADS_DIR
    if (not pdf_path or not os.path.exists(pdf_path)) and name:
        candidate = UPLOADS_DIR / name
        if candidate.exists():
            pdf_path = str(candidate)
        else:
            matches = list(UPLOADS_DIR.glob(f"*{name}*"))
            if matches:
                pdf_path = str(matches[0])
                
    if not pdf_path or not os.path.exists(pdf_path):
        existing_pdfs = list(UPLOADS_DIR.glob("*.pdf"))
        if existing_pdfs:
            pdf_path = str(sorted(existing_pdfs, key=os.path.getmtime, reverse=True)[0])
        else:
            pdf_path = str(UPLOADS_DIR / (name or "drawing.pdf"))
        
    cache_path = os.path.join(os.path.dirname(__file__), "extracted_tables.json")

    import time
    _t_start = time.time()
    print(f"[Timer 0] Start generate_boq_deduplicated at {_t_start:.2f}")

    try:
        req_filename = os.path.basename(pdf_path).replace("_cleaned", "").replace(".pdf", "").strip().lower()

        # Multi-tiered extraction lookup:
        # Tier 1: Frontend supplied elements directly in the payload
        passed_extracted = payload.get("extracted_data") or payload.get("extractedData")
        has_passed_data = False
        elements = []
        raw_items = []
        extracted_tables = []

        if passed_extracted and isinstance(passed_extracted, dict):
            p_elements = passed_extracted.get("elements", [])
            if p_elements:
                elements = p_elements
                raw_items = passed_extracted.get("raw_items", [])
                extracted_tables = passed_extracted.get("extracted_tables", [])
                has_passed_data = True
                print(f"[Generate BOQ] Using {len(elements)} elements directly from active UI state for {pdf_path}")

        # Tier 2: Check per-document cache on disk (uploads/<clean_stem>_extracted.json)
        doc_cache_path = UPLOADS_DIR / f"{req_filename}_extracted.json"
        if not has_passed_data and doc_cache_path.exists():
            try:
                with open(doc_cache_path, "r", encoding="utf-8") as f:
                    doc_cached = json.load(f)
                if doc_cached.get("elements"):
                    elements = doc_cached.get("elements", [])
                    raw_items = doc_cached.get("raw_items", [])
                    extracted_tables = doc_cached.get("extracted_tables", [])
                    has_passed_data = True
                    print(f"[Generate BOQ] Using {len(elements)} elements from document cache: {doc_cache_path}")
            except Exception as e:
                print(f"[Generate BOQ] Document cache read error: {e}")

        # Tier 3: Check global extracted_tables.json
        if not has_passed_data and os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    cached = json.load(f)
                cached_pdf = cached.get("pdf_path", "") or cached.get("filename", "")
                clean_cached = os.path.basename(cached_pdf).replace("_cleaned", "").replace(".pdf", "").strip().lower() if cached_pdf else ""
                if (clean_cached == req_filename or not clean_cached) and cached.get("elements"):
                    elements = cached.get("elements", [])
                    raw_items = cached.get("raw_items", [])
                    extracted_tables = cached.get("extracted_tables", [])
                    has_passed_data = True
                    print(f"[Generate BOQ] Using {len(elements)} elements from global cache for {pdf_path}")
            except Exception as e:
                print(f"[Generate BOQ] Global cache read error: {e}")

        # Tier 4: Only extract on-the-fly if absolutely no data exists anywhere
        if not has_passed_data:
            print(f"[Generate BOQ] No cached or UI elements found for {req_filename}. Extracting elements on-the-fly for {pdf_path}...")
            from core.document_extractor import extract_document_elements
            extracted = extract_document_elements(pdf_path)
            raw_items = extracted.get("raw_items", [])
            elements = extracted.get("elements", [])
            extracted_tables = extracted.get("extracted_tables", [])

        # Auto-extract structured tables from elements if extracted_tables is empty
        if not extracted_tables and elements:
            for el in elements:
                if isinstance(el, dict) and el.get("type") == "structured":
                    c = el.get("content", {})
                    if isinstance(c, dict) and "rows" in c:
                        extracted_tables.append({
                            "page": el.get("page", 1),
                            "table_title": el.get("title", "Configuration Table"),
                            "headers": c.get("headers", []),
                            "rows": c.get("rows", []),
                            "sheet_name": f"Page {el.get('page', 1)}"
                        })

        # Persist to disk cache so subsequent clicks are instantaneous
        try:
            cache_payload = {
                "pdf_path": pdf_path,
                "filename": os.path.basename(pdf_path),
                "raw_items": raw_items,
                "elements": elements,
                "extracted_tables": extracted_tables
            }
            with open(doc_cache_path, "w", encoding="utf-8") as f:
                json.dump(cache_payload, f, indent=2)
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump(cache_payload, f, indent=2)
        except Exception:
            pass
        
        # Load API key
        from services.ai_service import load_env_file
        load_env_file()
        api_key = os.environ.get("GEMINI_API_KEY")
        
        if not api_key:
            raise HTTPException(status_code=500, detail="GEMINI_API_KEY not configured in environment.")

        # Load Price list
        price_list_id = payload.get("price_list_id")
        file_path = get_price_list_path(price_list_id)
        
        price_list = []
        if os.path.exists(file_path):
            price_list = load_master_price_list(file_path, price_list_id)

        mapped_boq_items = []
        
        # Step 1: Execute Direct AI Semantic Takeoff Mapping if API key is present
        try:
            if api_key and (extracted_tables or elements):
                print(f"[Generate BOQ] Running Gemini AI takeoff mapper...")
                ai_items = run_gemini_boq_mapper_and_deduplicator(
                    extracted_tables, elements, price_list, api_key
                )
                if ai_items and isinstance(ai_items, list):
                    for idx, m_item in enumerate(ai_items):
                        rate = float(m_item.get("rate", 0.0))
                        qty = float(m_item.get("quantity", 1.0))
                        mapped_boq_items.append({
                            "item_id": f"boq_{idx:03d}",
                            "equipment_type": m_item.get("equipment_type") or m_item.get("category", "EQUIPMENT"),
                            "model": m_item.get("model") or m_item.get("item_name", ""),
                            "action": m_item.get("action", "INSTALL"),
                            "quantity": qty,
                            "source_sheet": m_item.get("source_sheet", "Drawing Schedule"),
                            "raw_text": m_item.get("raw_text") or m_item.get("model", ""),
                            "sor_code": m_item.get("sor_code", "UNQUOTED"),
                            "item_name": m_item.get("item_name") or m_item.get("model", ""),
                            "unit": m_item.get("unit", "each"),
                            "rate": rate,
                            "total_cost": rate * qty,
                            "similarity": float(m_item.get("similarity", 95.0)),
                            "auto_matched": True,
                            "row_idx": m_item.get("row_idx"),
                            "comment": m_item.get("comment") or m_item.get("notes", "Mapped via AI Semantic Takeoff"),
                            "matched_by_rule": "AI Semantic Takeoff",
                            "confidence_score": float(m_item.get("confidence_score", 95.0)),
                            "confidence_level": m_item.get("confidence_level", "HIGH"),
                            "evidence": m_item.get("evidence") or {
                                "source_sheet": m_item.get("source_sheet", "Drawing Schedule"),
                                "source_table": "AI Takeoff",
                                "source_row": idx,
                                "page": m_item.get("page", 1),
                                "ant_id": m_item.get("ant_id", "-"),
                                "model": m_item.get("model", ""),
                                "action": m_item.get("action", "INSTALL"),
                                "quantity": qty,
                                "entity_class": m_item.get("equipment_type", "EQUIPMENT"),
                                "target_sor": m_item.get("sor_code", "UNQUOTED"),
                                "target_name": m_item.get("item_name", ""),
                                "rate": rate,
                                "validation_status": "VERIFIED_IN_LAYOUT",
                                "confidence_score": 95.0,
                                "confidence_level": "HIGH",
                                "raw_text": m_item.get("raw_text", "")
                            },
                            "sources": m_item.get("sources", []),
                            "additional_sources": []
                        })
        except Exception as e:
            print(f"[Generate BOQ] AI Takeoff Mapper failed: {e}")

        # Step 2: Direct schema & fuzzy catalog matching fallback
        if not mapped_boq_items:
            print(f"[Generate BOQ] Running direct schema & fuzzy matching on extracted items...")
            candidate_items = []
            for t in extracted_tables:
                rows = t.get('rows', [])
                sheet = t.get('sheet_name') or f"Page {t.get('page', 1)}"
                for r_idx, row in enumerate(rows):
                    if not row or not isinstance(row, list) or len(row) < 2:
                        continue
                    model_str = str(row[1] if len(row) > 1 else row[0]).strip()
                    if not model_str or model_str in ['-', 'N/A', 'NONE']:
                        continue
                    act = "INSTALL"
                    qty = 1.0
                    for cell in row:
                        c_str = str(cell).upper().strip()
                        if "REMOVE" in c_str or "RECOVER" in c_str:
                            act = "REMOVE"
                        elif "RELOCATE" in c_str:
                            act = "RELOCATE"
                        elif "EXIST" in c_str:
                            act = "EXISTING"
                        try:
                            f_val = float(c_str)
                            if 0 < f_val < 100:
                                qty = f_val
                        except ValueError:
                            pass
                    if act == "EXISTING":
                        continue
                    candidate_items.append({
                        "model": model_str,
                        "equipment_type": "EQUIPMENT",
                        "action": act,
                        "quantity": qty,
                        "source_sheet": sheet,
                        "raw_text": " | ".join(str(c) for c in row if c)
                    })

            for el in elements:
                txt = str(el.get("content") or el.get("text") or "").strip()
                if len(txt) > 10 and not any(ign in txt.upper() for ign in ['DRAWING', 'TITLE', 'SCALE', 'DO NOT SCALE']):
                    act = "REMOVE" if any(k in txt.upper() for k in ["REMOVE", "RECOVER"]) else "INSTALL"
                    if "EXISTING" in txt.upper() and not any(k in txt.upper() for k in ["RECOVER", "REMOVE", "REPLACE"]):
                        continue
                    candidate_items.append({
                        "model": txt[:80],
                        "equipment_type": "EQUIPMENT",
                        "action": act,
                        "quantity": 1.0,
                        "source_sheet": el.get("sheet_name") or f"Page {el.get('page', 1)}",
                        "raw_text": txt
                    })

            for idx, c_item in enumerate(candidate_items):
                matched = match_item_to_price_list(c_item, price_list)
                sor = matched.get("code", "UNQUOTED") if matched else "UNQUOTED"
                name = matched.get("name", c_item["model"]) if matched else c_item["model"]
                rate = float(matched.get("rate", 0.0)) if matched else 0.0
                unit = matched.get("unit", "each") if matched else "each"
                r_idx = matched.get("row_idx") if matched else None
                sim = float(matched.get("similarity", 75.0)) if matched else 50.0

                mapped_boq_items.append({
                    "item_id": f"boq_{idx:03d}",
                    "equipment_type": c_item.get("equipment_type", "EQUIPMENT"),
                    "model": c_item["model"],
                    "action": c_item["action"],
                    "quantity": c_item["quantity"],
                    "source_sheet": c_item["source_sheet"],
                    "raw_text": c_item["raw_text"],
                    "sor_code": sor,
                    "item_name": name,
                    "unit": unit,
                    "rate": rate,
                    "total_cost": rate * c_item["quantity"],
                    "similarity": sim,
                    "auto_matched": bool(matched),
                    "row_idx": r_idx,
                    "comment": "Matched via Catalog Fuzzy Matcher",
                    "matched_by_rule": "Catalog Matcher",
                    "confidence_score": sim,
                    "confidence_level": "HIGH" if sim >= 80 else "MEDIUM",
                    "evidence": {
                        "source_sheet": c_item["source_sheet"],
                        "source_table": "Takeoff Table",
                        "source_row": idx,
                        "page": 1,
                        "ant_id": "-",
                        "model": c_item["model"],
                        "action": c_item["action"],
                        "quantity": c_item["quantity"],
                        "entity_class": c_item.get("equipment_type", "EQUIPMENT"),
                        "target_sor": sor,
                        "target_name": name,
                        "rate": rate,
                        "validation_status": "VERIFIED_IN_LAYOUT",
                        "confidence_score": sim,
                        "confidence_level": "HIGH" if sim >= 80 else "MEDIUM",
                        "raw_text": c_item["raw_text"]
                    },
                    "sources": [],
                    "additional_sources": []
                })

        consolidated = mapped_boq_items

        # Fast local structural checks
        validation_results = []
        try:
            pdf_corpus_lines = []
            c_doc = fitz.open(pdf_path)
            for page in c_doc:
                pdf_corpus_lines.append(page.get_text())
            c_doc.close()
            pdf_corpus = "\n".join(pdf_corpus_lines)
            validation_results = []
        except Exception:
            validation_results = []

        # Reset SQLite database boq_items for this run
        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM boq_items WHERE price_list_id = ?", (price_list_id or 1,))

        # Insert fresh unquoted items into boq_items so they appear in BOQ view and Excel
        for b_item in mapped_boq_items:
            r_idx = b_item.get("row_idx")
            ev_raw = b_item.get("evidence", {})
            if isinstance(ev_raw, dict):
                sources_list = [ev_raw]
                if "additional_sources" in b_item:
                    sources_list.extend(b_item["additional_sources"])
                evidence_payload = {
                    "summary": ev_raw,
                    "sources": sources_list
                }
            else:
                evidence_payload = {"summary": {}, "sources": []}
            evidence_str = json.dumps(evidence_payload)

            if r_idx is None:
                cursor.execute(
                    "INSERT INTO boq_items (code, name, unit, rate, quantity, action, comments, category, price_list_id, confidence_score, confidence_level, evidence_json) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        "UNQUOTED",
                        b_item.get("item_name", b_item.get("clean_text", "Unquoted Scope Item")),
                        b_item.get("unit", "each"),
                        float(b_item.get("rate", 0.0)),
                        float(b_item.get("quantity", 0.0)),
                        str(b_item.get("action", "INSTALL")).upper(),
                        str(b_item.get("comment", "Estimator need to fill")),
                        "Non-SOR & Unquoted Items",
                        price_list_id or 1,
                        float(b_item.get("confidence_score", 50.0)),
                        str(b_item.get("confidence_level", "NEEDS_REVIEW")),
                        evidence_str
                    )
                )
                b_item["row_idx"] = cursor.lastrowid

        # Group and write computed quantities, actions, comments, and confidence for standard SOR items
        db_row_updates = {}
        for b_item in mapped_boq_items:
            r_idx = b_item.get("row_idx")
            if r_idx is not None and b_item.get("similarity", 0) > 0:
                r_str = str(r_idx)
                if r_str not in db_row_updates:
                    db_row_updates[r_str] = {
                        "qty": 0.0,
                        "actions": set(),
                        "comments": [],
                        "models": [],
                        "conf_scores": [],
                        "conf_levels": [],
                        "evidences": []
                    }
                db_row_updates[r_str]["qty"] += float(b_item.get("quantity", 0))
                if b_item.get("action"):
                    db_row_updates[r_str]["actions"].add(b_item["action"].upper())
                
                m_name = b_item.get("model") or b_item.get("item_name") or ""
                m_qty = int(b_item.get("quantity", 1))
                if m_name:
                    clean_m = re.sub(r'^(?:ERICSSON|TELSTRA)\s+', '', str(m_name), flags=re.IGNORECASE).strip()
                    db_row_updates[r_str]["models"].append(f"{m_qty}x {clean_m}")

                cmt = str(b_item.get("comment", "")).strip()
                if cmt and cmt not in db_row_updates[r_str]["comments"]:
                    if "Data not matching" in cmt:
                        db_row_updates[r_str]["comments"].insert(0, cmt)
                    else:
                        db_row_updates[r_str]["comments"].append(cmt)
                if b_item.get("confidence_score") is not None:
                    db_row_updates[r_str]["conf_scores"].append(float(b_item["confidence_score"]))
                if b_item.get("confidence_level"):
                    db_row_updates[r_str]["conf_levels"].append(b_item["confidence_level"])
                if b_item.get("evidence_json", {}).get("sources"):
                    db_row_updates[r_str]["evidences"].extend(b_item["evidence_json"]["sources"])
                elif b_item.get("sources"):
                    db_row_updates[r_str]["evidences"].extend(b_item["sources"])
                    if "additional_sources" in b_item:
                        db_row_updates[r_str]["evidences"].extend(b_item["additional_sources"])
                elif b_item.get("evidence"):
                    db_row_updates[r_str]["evidences"].append(b_item["evidence"])
                    if "additional_sources" in b_item:
                        db_row_updates[r_str]["evidences"].extend(b_item["additional_sources"])

        for r_str, data in db_row_updates.items():
            cursor.execute("SELECT code, name, unit, rate, category FROM price_items WHERE id = ?", (int(r_str),))
            sor_row = cursor.fetchone()
            if not sor_row:
                continue

            act_str = ", ".join(sorted(data["actions"]))
            if data["comments"]:
                cmt_str = " | ".join(data["comments"])
            elif len(data.get("models", [])) > 1:
                cmt_str = ", ".join(data["models"])
            else:
                cmt_str = ""
            avg_conf = sum(data["conf_scores"]) / len(data["conf_scores"]) if data["conf_scores"] else 100.0
            worst_level = "NEEDS_REVIEW" if "NEEDS_REVIEW" in data["conf_levels"] else ("MEDIUM" if "MEDIUM" in data["conf_levels"] else "HIGH")
            evidence_payload = {
                "summary": data["evidences"][0] if data["evidences"] else {},
                "sources": data["evidences"]
            }
            ev_str = json.dumps(evidence_payload)
            try:
                cursor.execute(
                    "INSERT INTO boq_items (price_list_id, code, name, unit, rate, quantity, action, comments, category, price_item_id, confidence_score, confidence_level, evidence_json) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (
                        price_list_id or 1,
                        sor_row["code"],
                        sor_row["name"],
                        sor_row["unit"],
                        sor_row["rate"],
                        data["qty"],
                        act_str,
                        cmt_str,
                        sor_row["category"],
                        int(r_str),
                        avg_conf,
                        worst_level,
                        ev_str
                    )
                )
                new_boq_item_id = cursor.lastrowid
                for b_item in mapped_boq_items:
                    if str(b_item.get("row_idx")) == r_str:
                        b_item["row_idx"] = new_boq_item_id
            except Exception as e:
                print(f"[Generate BOQ] Error inserting boq_items row: {e}")
        conn.commit()
        conn.close()

        # Generate the Excel BOQ workbook in-place
        generate_populated_boq_excel(file_path, {}, file_path)
            
        return {
            "status": "success",
            "drawing_name": os.path.basename(pdf_path),
            "raw_count": len(raw_items),
            "consolidated_count": len(consolidated),
            "elements": elements,
            "extracted_tables": extracted_tables,
            "mapped_items": mapped_boq_items,
            "checklist": validation_results
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to generate deduplicated BOQ: {str(e)}")

@app.post("/api/sanitize-pdf")
async def sanitize_pdf(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Accepts a local PDF path, deletes all annotations (Bluebeam markups, clouds,
    highlights, strikethroughs), saves the cleaned PDF to the uploads folder,
    and returns its metadata along with base64 encoded bytes for the frontend.
    """
    pdf_path = payload.get("path")
    if not pdf_path:
        raise HTTPException(status_code=400, detail="Missing 'path' in payload.")
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail=f"PDF not found at: {pdf_path}")
        
    try:
        doc = fitz.open(pdf_path)
        page_count = len(doc)
        
        # Remove all annotations from all pages
        for page in doc:
            annots = list(page.annots())
            for annot in annots:
                page.delete_annot(annot)
                
        # Save to uploads folder
        uploads_dir = os.path.join(os.path.dirname(__file__), "uploads")
        os.makedirs(uploads_dir, exist_ok=True)
        
        base_name = os.path.basename(pdf_path)
        name_part, ext = os.path.splitext(base_name)
        cleaned_filename = f"{name_part}_cleaned{ext}"
        cleaned_pdf_path = os.path.join(uploads_dir, cleaned_filename)
        
        # Cleaned PDF path needs to be absolute
        cleaned_pdf_path = os.path.abspath(cleaned_pdf_path)
        
        # Save changes
        doc.save(cleaned_pdf_path, garbage=3, deflate=True)
        doc.close()
        
        # Read the cleaned PDF to encode in base64
        with open(cleaned_pdf_path, "rb") as f:
            cleaned_bytes = f.read()
            
        base64_str = base64.b64encode(cleaned_bytes).decode("utf-8")
        
        return {
            "name": cleaned_filename,
            "path": cleaned_pdf_path,
            "base64": base64_str,
            "pages": str(page_count)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to sanitize PDF: {str(e)}")

from pydantic import BaseModel

from typing import Optional

class PriceListCreateModel(BaseModel):
    name: str

class PriceItemModel(BaseModel):
    code: str
    name: str
    unit: str
    rate: float
    category: Optional[str] = "General SOR Pricing Items"

class PriceItemUpdateModel(BaseModel):
    code: str
    name: str
    unit: str
    rate: float
    category: Optional[str] = "General SOR Pricing Items"

class DeleteBatchModel(BaseModel):
    row_indices: List[int]

class ClearQuantitiesBatchModel(BaseModel):
    row_indices: List[int]

class CellUpdateModel(BaseModel):
    row_idx: int
    col_idx: int
    value: str

class CategoryRenameModel(BaseModel):
    new_name: str

class CorrectionLogModel(BaseModel):
    pdf_name: str
    original_description: str
    corrected_code: str
    corrected_name: str
    corrected_rate: float
    estimator_username: Optional[str] = "Estimator"


class ExportPayload(BaseModel):
    quantities: Optional[Dict[str, float]] = None

@app.post("/api/clear-cache")
async def clear_cache() -> Dict[str, str]:
    """Clears user mappings cache file."""
    try:
        clear_user_mappings()
        return {"status": "success", "message": "Cache cleared successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to clear cache: {str(e)}")

@app.post("/api/log-correction")
def log_correction(
    payload: CorrectionLogModel,
    background_tasks: BackgroundTasks = None
) -> Dict[str, Any]:
    try:
        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO correction_log (pdf_name, original_description, corrected_code, corrected_name, corrected_rate, estimator_username) VALUES (?, ?, ?, ?, ?, ?)",
            (payload.pdf_name, payload.original_description.upper(), payload.corrected_code.upper(), payload.corrected_name, payload.corrected_rate, payload.estimator_username)
        )
        conn.commit()
        conn.close()
        
        # Also cache the correction in user mappings so it acts as a prior correction
        from services.matcher import save_user_mapping
        save_user_mapping(
            payload.original_description.upper(),
            payload.corrected_code.upper(),
            payload.corrected_name,
            payload.corrected_rate
        )
        return {"status": "success", "message": "Correction logged successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to log correction: {str(e)}")


class PromptModel(BaseModel):
    id: Optional[int] = None
    name: str
    title: str
    prompt: str
    enabled: Optional[int] = 1
    project_type: Optional[str] = "Default"

@app.get("/api/prompts")
def get_prompts():
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM ai_prompts ORDER BY id ASC")
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/prompts")
def save_prompt(prompt: PromptModel):
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    if prompt.id:
        cursor.execute(
            "UPDATE ai_prompts SET name = ?, title = ?, prompt = ?, enabled = ?, project_type = ? WHERE id = ?",
            (prompt.name, prompt.title, prompt.prompt, prompt.enabled, prompt.project_type, prompt.id)
        )
    else:
        cursor.execute(
            "INSERT OR REPLACE INTO ai_prompts (name, title, prompt, enabled, project_type) VALUES (?, ?, ?, ?, ?)",
            (prompt.name, prompt.title, prompt.prompt, prompt.enabled, prompt.project_type)
        )
    conn.commit()
    conn.close()
    return {"status": "success"}

@app.post("/api/project/close")
def close_project_session(payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    Cleans up active project state:
    - Clears SQLite price_items calculated quantities, actions, and comments.
    - Removes extracted_tables.json cache so new projects start completely fresh.
    """
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE price_items SET quantity = 0, action = '', comments = ''")
    cursor.execute("DELETE FROM price_items WHERE code = 'UNQUOTED'")
    cursor.execute("DELETE FROM boq_items")
    conn.commit()
    conn.close()
    
    cache_path = os.path.join(os.path.dirname(__file__), "extracted_tables.json")
    if os.path.exists(cache_path):
        try:
            os.remove(cache_path)
        except Exception:
            pass
            
    return {"status": "success", "message": "Project session closed and state reset cleanly."}

@app.delete("/api/prompts/{prompt_id}")
def delete_prompt(prompt_id: int):
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ai_prompts WHERE id = ?", (prompt_id,))
    conn.commit()
    conn.close()
    return {"status": "success"}

@app.post("/api/prompts/reset")
def reset_prompts():
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DROP TABLE IF EXISTS ai_prompts")
    conn.commit()
    conn.close()
    # Re-trigger init_db to bootstrap defaults
    from services.db import init_db
    init_db()
    return {"status": "success"}

# ==========================================
# DRAWING PARSER CONFIGS & REGEX ENDPOINTS
# ==========================================

class ParserConfigUpdateModel(BaseModel):
    pattern_value: str
    is_active: Optional[int] = 1

class RegexTestRequestModel(BaseModel):
    pattern: str
    sample_text: str

@app.get("/api/parser-configs")
def get_parser_configs():
    """Fetches all drawing parsing regex patterns and keyword configurations from SQLite."""
    from services.db import get_all_parser_configs
    return get_all_parser_configs()

@app.post("/api/parser-configs/test-regex")
def test_regex_endpoint(payload: RegexTestRequestModel):
    """Tests a regex pattern against sample text and returns matches and groups."""
    import re
    try:
        compiled = re.compile(payload.pattern, re.IGNORECASE)
        matches = []
        for m in compiled.finditer(payload.sample_text):
            matches.append({
                "match": m.group(0),
                "groups": list(m.groups()),
                "span": list(m.span())
            })
        return {
            "status": "success",
            "is_valid": True,
            "match_count": len(matches),
            "matches": matches
        }
    except Exception as e:
        return {
            "status": "error",
            "is_valid": False,
            "error": str(e),
            "match_count": 0,
            "matches": []
        }

@app.post("/api/parser-configs/reset-defaults")
def reset_parser_configs_endpoint():
    """Resets all parser configurations to default Telstra standard patterns."""
    from services.db import reset_default_parser_configs
    reset_default_parser_configs()
    return {"status": "success"}

@app.put("/api/parser-configs/{config_id:int}")
@app.post("/api/parser-configs/update/{config_id:int}")
def update_parser_config_endpoint(config_id: int, payload: ParserConfigUpdateModel):
    """Updates a parser configuration pattern and active state."""
    from services.db import update_parser_config
    success = update_parser_config(config_id, payload.pattern_value, payload.is_active or 1)
    if not success:
        raise HTTPException(status_code=404, detail="Parser config not found.")
    return {"status": "success"}

class SettingsModel(BaseModel):
    gemini_rate_limit: int
    target_headers: List[str]
    column_mappings: Dict[str, str]

SETTINGS_FILE = os.path.join(os.path.dirname(__file__), "settings.json")

def load_settings() -> Dict[str, Any]:
    default_settings = {
        "gemini_rate_limit": 0,
        "target_headers": ["ITEM", "EQUIPMENT DETAILS", "EXISTING", "PROPOSED", "TOTAL", "REFERENCE DWG"],
        "column_mappings": {
            "description": "EQUIPMENT DETAILS",
            "quantity": "PROPOSED",
            "reference": "REFERENCE DWG"
        }
    }
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                for k, v in default_settings.items():
                    if k not in data:
                        data[k] = v
                return data
        except Exception:
            pass
    return default_settings

def save_settings(settings: Dict[str, Any]) -> None:
    try:
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(settings, f, indent=2)
    except Exception as e:
        print(f"Error saving settings: {e}")

@app.get("/api/settings")
def get_settings() -> Dict[str, Any]:
    """Gets the current settings."""
    return load_settings()

@app.post("/api/settings")
def update_settings(settings: SettingsModel) -> Dict[str, Any]:
    """Updates the settings."""
    save_settings(settings.dict())
    return {"status": "success", "settings": load_settings()}

def get_price_list_response(price_list_id: Optional[int] = None) -> Dict[str, Any]:
    file_path = get_price_list_path(price_list_id)
    return {
        "items": load_master_price_list(file_path, price_list_id),
        "col_widths": get_sheet_column_widths(file_path)
    }

@app.get("/api/price-lists")
def get_price_lists() -> List[Dict[str, Any]]:
    """Retrieves all price lists in the database."""
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, is_active FROM price_lists ORDER BY id")
    rows = cursor.fetchall()
    conn.close()
    return [{"id": r["id"], "name": r["name"], "is_active": r["is_active"]} for r in rows]

@app.post("/api/price-lists")
def create_price_list(payload: PriceListCreateModel) -> Dict[str, Any]:
    """Creates a new empty price list and generates a blank template Excel file for it."""
    name_clean = payload.name.strip()
    if not name_clean:
        raise HTTPException(status_code=400, detail="Price Book name cannot be empty.")
        
    from services.db import get_db_connection
    import sqlite3
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO price_lists (name, is_active) VALUES (?, 0)", (name_clean,))
        new_id = cursor.lastrowid
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="A Price Book with this name already exists.")
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=f"Failed to create price book: {str(e)}")
    conn.close()
    
    # Generate blank Excel template immediately on disk
    success = sync_db_to_active_excel(new_id)
    if not success:
        # Cleanup DB if file creation fails
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM price_lists WHERE id = ?", (new_id,))
        conn.commit()
        conn.close()
        raise HTTPException(status_code=500, detail="Failed to initialize spreadsheet file for new Price Book.")
        
    return {"status": "success", "id": new_id, "name": name_clean}

@app.post("/api/price-lists/active/{list_id}")
def set_active_price_list(list_id: int) -> Dict[str, Any]:
    """Switches the active price list pointer in the database (global fallback)."""
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Check if exists
    cursor.execute("SELECT id FROM price_lists WHERE id = ?", (list_id,))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Price Book not found.")
        
    cursor.execute("UPDATE price_lists SET is_active = 0")
    cursor.execute("UPDATE price_lists SET is_active = 1 WHERE id = ?", (list_id,))
    conn.commit()
    conn.close()
    
    clear_user_mappings()
    return {"status": "success", "active_id": list_id}

@app.put("/api/price-lists/{list_id}")
def rename_price_list(list_id: int, payload: PriceListCreateModel) -> Dict[str, Any]:
    """Renames an existing Price Book."""
    name_clean = payload.name.strip()
    if not name_clean:
        raise HTTPException(status_code=400, detail="Price Book name cannot be empty.")
        
    from services.db import get_db_connection
    import sqlite3
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Check if another book has the same name
        cursor.execute("SELECT id FROM price_lists WHERE name = ? AND id != ?", (name_clean, list_id))
        if cursor.fetchone():
            conn.close()
            raise HTTPException(status_code=400, detail="A Price Book with this name already exists.")
            
        cursor.execute("UPDATE price_lists SET name = ? WHERE id = ?", (name_clean, list_id))
        conn.commit()
    except HTTPException:
        conn.close()
        raise
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=f"Failed to rename price book: {str(e)}")
    conn.close()
    
    return {"status": "success", "id": list_id, "name": name_clean}

@app.delete("/api/price-lists/{list_id}")
def delete_price_list(list_id: int) -> Dict[str, Any]:
    """Deletes a price list, cascade deletes all items, and deletes the Excel file on disk."""
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Enforce count > 1
    cursor.execute("SELECT COUNT(*) as count FROM price_lists")
    if cursor.fetchone()["count"] <= 1:
        conn.close()
        raise HTTPException(status_code=400, detail="Cannot delete the last remaining Price Book.")
        
    # Check if exists
    cursor.execute("SELECT is_active FROM price_lists WHERE id = ?", (list_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Price Book not found.")
        
    was_active = row["is_active"]
    
    # Acquire file lock to prevent race conditions during deletion
    lock = price_list_locks[list_id]
    with lock:
        # Delete file on disk first
        file_path = get_price_list_path(list_id)
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                conn.close()
                raise HTTPException(status_code=500, detail=f"Failed to delete spreadsheet file: {str(e)}. Aborting deletion.")
                
        # Perform DB transaction to delete price list & its associated items
        try:
            # Delete associated items explicitly (code-level cascade delete)
            cursor.execute("DELETE FROM price_items WHERE price_list_id = ?", (list_id,))
            cursor.execute("DELETE FROM price_lists WHERE id = ?", (list_id,))
            
            # Re-assign active flag to another price list if the deleted one was active
            if was_active:
                cursor.execute("SELECT id FROM price_lists LIMIT 1")
                another_row = cursor.fetchone()
                if another_row:
                    cursor.execute("UPDATE price_lists SET is_active = 1 WHERE id = ?", (another_row["id"],))
                    
            conn.commit()
        except Exception as e:
            conn.rollback()
            conn.close()
            raise HTTPException(status_code=500, detail=f"Database error during deletion: {str(e)}")
            
    conn.close()
    
    # Clean up lock dictionary
    price_list_locks.pop(list_id, None)
    clear_user_mappings()
    
    return {"status": "success"}

@app.get("/api/price-list")
def get_price_list(price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Loads and returns the active price list items and column widths."""
    return get_price_list_response(price_list_id)

@app.get("/api/price-list/file")
def get_price_list_file(price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Returns the active price list Excel file as a base64-encoded string along with layout dimensions for frontend import."""
    file_path = get_price_list_path(price_list_id)
    
    if price_list_id is None:
        price_list_id = get_id_from_path(get_price_list_path()) or get_default_price_list_id()
    
    # Sync DB to Excel before serving to prevent staleness bugs
    sync_db_to_active_excel(price_list_id)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="No price list has been uploaded yet.")
    
    lock = price_list_locks[price_list_id]
    with lock:
        try:
            with open(file_path, "rb") as f:
                raw_bytes = f.read()
            encoded = base64.b64encode(raw_bytes).decode("utf-8")
            layout = get_excel_layout_metadata(file_path)
            return {
                "data": encoded,
                "filename": f"price_list_{price_list_id}.xlsx",
                "col_widths": layout.get("col_widths", {}),
                "row_heights": layout.get("row_heights", {})
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to read price list file: {str(e)}")

@app.post("/api/price-list/upload")
async def upload_price_list(price_list_id: Optional[int] = None, file: UploadFile = File(...)) -> Dict[str, Any]:
    """Uploads a new excel pricebook (legacy alias for import)."""
    return await import_price_list(price_list_id, file)

@app.post("/api/price-list/import")
async def import_price_list(price_list_id: Optional[int] = None, file: UploadFile = File(...)) -> Dict[str, Any]:
    """Imports a new excel pricebook, updating active list."""
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Uploaded file must be an Excel workbook (.xlsx or .xls).")
        
    if price_list_id is None:
        price_list_id = get_id_from_path(get_price_list_path()) or get_default_price_list_id()
        
    try:
        # Enforce size limit (10MB)
        MAX_FILE_SIZE = 10 * 1024 * 1024
        contents = await file.read(MAX_FILE_SIZE + 1)
        if len(contents) > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="File size exceeds the 10MB limit.")
            
        # Parse Excel workbook in-memory to validate
        import io
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid Excel file. The file may be corrupt or encrypted.")
        # Select SOR / Schedule of Rates tab if present, else first worksheet
        target_sheet = None
        for name in wb.sheetnames:
            n_up = name.upper()
            if any(k in n_up for k in ["SOR", "SCHEDULE", "PRICE", "RATE", "PRICING"]):
                target_sheet = wb[name]
                break
        sheet = target_sheet or wb.worksheets[0]
        
        # Identify headers row dynamically
        header_row_idx = 1
        code_col, name_col, unit_col, rate_col, category_col, comments_col = None, None, None, None, None, None
        action_col, quantity_col = None, None
        found_header = False
        
        for r in range(1, min(15, sheet.max_row + 1)):
            row_str = [str(sheet.cell(r, c).value or "").strip().upper() for c in range(1, min(15, sheet.max_column + 1))]
            if any(k in val for val in row_str for k in ["RATE", "EXCLUDING GST", "UNIT OF QTY", "UNIT", "PRICE"]):
                header_row_idx = r
                found_header = True
                for c in range(1, min(15, sheet.max_column + 1)):
                    val = str(sheet.cell(r, c).value or "").strip().upper()
                    if not val:
                        continue
                    if "UNIT" in val:
                        unit_col = c
                    elif "RATE" in val or "PRICE" in val or "EXCLUDING GST" in val or "EXCL GST" in val:
                        rate_col = c
                    elif "CATEGORY" in val or "SECTION" in val or "TRADE" in val:
                        category_col = c
                    elif "COMMENT" in val or "NOTE" in val or "REMARK" in val:
                        comments_col = c
                    elif "ACTION" in val or "SCOPE" in val:
                        action_col = c
                    elif "QTY" in val or "QUANTITY" in val:
                        quantity_col = c
                    elif any(d in val for d in ["DESCRIPTION", "ITEM DESCRIPTION", "DETAILS", "EQUIPMENT"]):
                        name_col = c
                    elif any(cd in val for cd in ["SOR CODE", "ITEM CODE", "ITEM NO", "ITEM #", "CODE", "SOR NO"]):
                        code_col = c
                break

        # If Telstra/standard structure (Unit at col 3 (1-based), Rate at col 4 (1-based))
        if unit_col == 3 and rate_col == 4:
            code_col = 1
            name_col = 2
        elif name_col is None or code_col is None or name_col == code_col:
            code_col = 1
            name_col = 2

        if unit_col is None:
            unit_col = 3
        if rate_col is None:
            rate_col = 4

        # Parse rows
        parsed_items = []
        current_category = "General SOR Pricing Items"
        
        for r in range(header_row_idx + 1, sheet.max_row + 1):
            c0 = sheet.cell(r, 1)
            c1 = sheet.cell(r, 2)
            c_unit = sheet.cell(r, unit_col) if unit_col else None
            c_rate = sheet.cell(r, rate_col) if rate_col else None
            
            val0 = str(c0.value or "").strip()
            val1 = str(c1.value or "").strip()
            unit_str = str(c_unit.value or "").strip().lower() if c_unit else ""
            
            rate_val = None
            if c_rate and c_rate.value is not None:
                try:
                    rate_val = float(str(c_rate.value).replace('$', '').replace(',', '').strip())
                except ValueError:
                    rate_val = None

            bold0 = c0.font.bold if c0.font else False
            bold1 = c1.font.bold if c1.font else False
            
            # Category Header identification rule:
            # 1. Bold text in col A or col B with no rate/unit
            # 2. Or col A has text, col B is completely empty, and rate/unit are empty
            is_category = False
            if rate_val is None and not unit_str:
                if (bold0 or bold1) and (val0 or val1):
                    is_category = True
                elif val0 and not val1:
                    is_category = True
                    
            if is_category:
                current_category = val0 or val1
                continue
                
            if not val0 and not val1:
                continue
                
            code = val0
            name = val1
            if not name and code:
                if len(code) > 15:
                    name = code
                    code = ""
                else:
                    name = code
                    
            unit = unit_str if unit_str else "each"
            rate = rate_val if rate_val is not None else 0.0
            
            comments = str(sheet.cell(r, comments_col).value or "").strip() if (comments_col and sheet.max_column >= comments_col) else ""
            action = str(sheet.cell(r, action_col).value or "").strip() if (action_col and sheet.max_column >= action_col) else ""
            
            qty = 0.0
            if quantity_col and sheet.max_column >= quantity_col:
                qty_cell = sheet.cell(r, quantity_col).value
                if qty_cell is not None:
                    try:
                        qty = float(str(qty_cell).replace(',', '').strip())
                    except ValueError:
                        pass
                        
            category = ""
            if category_col and sheet.max_column >= category_col:
                category = str(sheet.cell(r, category_col).value or "").strip()
            if not category:
                category = current_category or "General SOR Pricing Items"
                
            parsed_items.append((code, name, action, unit, rate, qty, category, comments))
            
        wb.close()
        
        # Concurrency & Lock
        lock = price_list_locks[price_list_id]
        with lock:
            file_path = get_price_list_path(price_list_id)
            
            # High-speed batch DB insertion in transaction
            from services.db import get_db_connection
            conn = get_db_connection()
            cursor = conn.cursor()
            try:
                # Clear items for this price list
                cursor.execute("DELETE FROM price_items WHERE price_list_id = ?", (price_list_id,))
                
                # Insert items in a single executemany call (50x faster)
                cursor.executemany(
                    "INSERT INTO price_items (code, name, action, unit, rate, quantity, category, comments, price_list_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    [(item[0], item[1], item[2], item[3], item[4], item[5], item[6], item[7], price_list_id) for item in parsed_items]
                )
                conn.commit()
            except Exception as db_err:
                conn.rollback()
                conn.close()
                raise HTTPException(status_code=500, detail=f"Database transaction failed: {str(db_err)}")
            conn.close()
            
            # Run spreadsheet file sync asynchronously in background thread
            def _async_post_import_tasks(pid):
                try:
                    sync_db_to_active_excel(pid)
                except Exception as post_err:
                    print(f"[Import Background Task] {post_err}")

            import threading
            threading.Thread(target=_async_post_import_tasks, args=(price_list_id,), daemon=True).start()
            
        clear_user_mappings()
        return {"status": "success", **get_price_list_response(price_list_id)}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

@app.get("/api/price-list/template")
def get_price_list_template():
    """Generates and returns a styled blank Excel template in the correct format."""
    from fastapi.responses import StreamingResponse
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from services.matcher import write_instructions_sheet
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Price List Template"
    sheet.views.sheetView[0].showGridLines = True
    
    # Create second sheet for Instructions
    ins_sheet = wb.create_sheet(title="DOs and DONTs")
    write_instructions_sheet(ins_sheet)
    
    headers = ["SOR Code", "Item Description", "Action", "Unit", "Rate", "Quantity", "Total Cost", "Category", "Comments"]
    sheet.append(headers)
    
    header_fill = PatternFill(start_color="1F376A", end_color="1F376A", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="left", vertical="center")
    
    thin = Side(border_style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for col_idx in range(1, 10):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border
        
    # Sample row
    sheet.append(["W1002", "Sample Installation Service", "", "each", 450.00, 1.0, "=E2*F2", "Installation Services", "Optional comments here"])
    
    sample_font = Font(name="Segoe UI", size=10, color="57606A")
    for col_idx in range(1, 10):
        cell = sheet.cell(row=2, column=col_idx)
        cell.font = sample_font
        cell.border = border
        if col_idx == 5: # Rate
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx == 6: # Quantity
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 7: # Total Cost formula
            cell.number_format = "$#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")
            
    column_widths = {"A": 15, "B": 45, "C": 12, "D": 12, "E": 15, "F": 12, "G": 15, "H": 25, "I": 30}
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    wb.close()
    
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=price_list_template.xlsx"}
    )

@app.post("/api/export-priced-excel")
def export_priced_excel(price_list_id: Optional[int] = None, payload: Optional[ExportPayload] = None, only_priced: bool = False):
    """Generates the populated Excel BOQ workbook on the fly from boq_items and returns it."""
    from fastapi.responses import FileResponse
    from services.db import get_db_connection
    
    file_path = get_price_list_path(price_list_id)
    
    # If custom quantities payload is passed, update boq_items database first
    if payload and payload.quantities is not None:
        conn = get_db_connection()
        cursor = conn.cursor()
        for row_idx_str, qty in payload.quantities.items():
            try:
                row_idx = int(row_idx_str)
                cursor.execute("UPDATE boq_items SET quantity = ? WHERE id = ?", (qty, row_idx))
            except ValueError:
                pass
        conn.commit()
        conn.close()
        # If payload is provided, default to only exporting priced items
        only_priced = True
        
    output_file = os.path.join(os.path.dirname(__file__), "uploads", "Priced_BOQ.xlsx")
    
    success = generate_populated_boq_excel(file_path, {}, output_file, only_priced=only_priced)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to generate BOQ Excel sheet from database.")
        
    return FileResponse(
        path=output_file,
        filename="Priced_BOQ.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/boq-items")
def get_boq_items(price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Loads and returns the active BOQ items (drawn items mapped to SOR + unquoted)."""
    from services.matcher import load_boq_items
    file_path = get_price_list_path(price_list_id)
    return {
        "items": load_boq_items(price_list_id),
        "col_widths": get_sheet_column_widths(file_path)
    }

@app.put("/api/boq-items/cell")
def update_boq_item_cell(payload: CellUpdateModel, price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Writes value directly to specified cell coordinates in SQLite boq_items table."""
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if payload.col_idx in (6, 7):
        cursor.execute("UPDATE boq_items SET comments = ? WHERE id = ?", (payload.value, payload.row_idx))
    elif payload.col_idx == 2:
        cursor.execute("UPDATE boq_items SET action = ? WHERE id = ?", (payload.value.upper(), payload.row_idx))
    else:
        qty = 0.0
        if payload.value.strip() != "":
            try:
                qty = float(payload.value)
            except ValueError:
                pass
        cursor.execute("UPDATE boq_items SET quantity = ? WHERE id = ?", (qty, payload.row_idx))
        
    conn.commit()
    conn.close()
    clear_user_mappings()
    
    from services.matcher import load_boq_items
    file_path = get_price_list_path(price_list_id)
    return {
        "status": "success",
        "items": load_boq_items(price_list_id),
        "col_widths": get_sheet_column_widths(file_path)
    }

@app.post("/api/boq-items/clear-quantities")
def clear_boq_quantities(price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Clears all calculated BOQ items and resets them."""
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM boq_items WHERE price_list_id = ?", (price_list_id or 1,))
    conn.commit()
    conn.close()
    clear_user_mappings()
    
    from services.matcher import load_boq_items
    file_path = get_price_list_path(price_list_id)
    return {
        "status": "success",
        "items": load_boq_items(price_list_id),
        "col_widths": get_sheet_column_widths(file_path)
    }

@app.put("/api/price-list/cell")
def update_price_list_cell(payload: CellUpdateModel, price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Writes value directly to specified cell coordinates in SQLite database (no disk sync for speed)."""
    file_path = get_price_list_path(price_list_id)
    success = write_cell_value_to_excel(file_path, payload.row_idx, payload.col_idx, payload.value)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to write value to database cell.")
    clear_user_mappings()
    return {"status": "success", **get_price_list_response(price_list_id)}

@app.post("/api/price-list/clear-quantities")
def clear_quantities(price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Clears all calculated quantity values."""
    file_path = get_price_list_path(price_list_id)
    success = clear_column_values_in_excel(file_path, "QTY")
    if not success:
        raise HTTPException(status_code=500, detail="Failed to clear quantities in database.")
    
    # Batch action: sync database state to disk
    resolved_id = get_id_from_path(file_path) or get_default_price_list_id()
    sync_db_to_active_excel(resolved_id)
    
    clear_user_mappings()
    return {"status": "success", **get_price_list_response(price_list_id)}

@app.post("/api/price-list/restore-mapped-items")
def restore_mapped_items(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Restores saved mapped_items into SQLite price_items for the active price list."""
    mapped_items = payload.get("mapped_items", [])
    price_list_id = payload.get("price_list_id", 1)
    
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE price_items SET quantity = 0, action = '', comments = '' WHERE price_list_id = ?", (price_list_id,))
    cursor.execute("DELETE FROM price_items WHERE code = 'UNQUOTED' AND price_list_id = ?", (price_list_id,))

    # Re-insert unquoted non-SOR items
    for b_item in mapped_items:
        r_idx = b_item.get("row_idx")
        if r_idx is None or str(b_item.get("sor_code", "")).upper() == "UNQUOTED":
            cursor.execute(
                "INSERT INTO price_items (code, name, unit, rate, quantity, action, comments, category, price_list_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    "UNQUOTED",
                    b_item.get("item_name", b_item.get("clean_text", "Unquoted Scope Item")),
                    b_item.get("unit", "each"),
                    float(b_item.get("rate", 0.0)),
                    float(b_item.get("quantity", 1.0)),
                    str(b_item.get("action", "INSTALL")).upper(),
                    str(b_item.get("comment", "Estimator need to fill")),
                    "Non-SOR & Unquoted Items",
                    price_list_id
                )
            )
            b_item["row_idx"] = cursor.lastrowid
    
    db_row_updates = {}
    for b_item in mapped_items:
        r_idx = b_item.get("row_idx")
        if r_idx is not None and str(b_item.get("sor_code", "")).upper() != "UNQUOTED":
            r_str = str(r_idx)
            if r_str not in db_row_updates:
                db_row_updates[r_str] = {
                    "qty": 0.0,
                    "actions": set(),
                    "comments": []
                }
            db_row_updates[r_str]["qty"] += float(b_item.get("quantity", 0))
            if b_item.get("action"):
                db_row_updates[r_str]["actions"].add(str(b_item["action"]).upper())
            cmt = str(b_item.get("comment", "")).strip()
            if cmt and cmt not in db_row_updates[r_str]["comments"]:
                db_row_updates[r_str]["comments"].append(cmt)
                
    for r_str, data in db_row_updates.items():
        act_str = ", ".join(sorted(data["actions"]))
        cmt_str = " | ".join(data["comments"])
        try:
            cursor.execute(
                "UPDATE price_items SET quantity = ?, action = ?, comments = ? WHERE id = ?",
                (data["qty"], act_str, cmt_str, int(r_str))
            )
        except Exception as e:
            print(f"[Restore Mapped Items] Error updating row {r_str}: {e}")

    # Clear and repopulate boq_items table for the BOQ Viewer
    cursor.execute("DELETE FROM boq_items WHERE price_list_id = ?", (price_list_id,))
    for b_item in mapped_items:
        code = b_item.get("sor_code", "UNQUOTED")
        name = b_item.get("item_name", b_item.get("clean_text", "Unquoted Scope Item"))
        unit = b_item.get("unit", "each")
        rate = float(b_item.get("rate") or 0.0)
        quantity = float(b_item.get("quantity") or 0.0)
        action = str(b_item.get("action", "INSTALL")).upper()
        comments = b_item.get("comment", b_item.get("comments", ""))
        category = b_item.get("category", "General SOR Pricing Items")
        price_item_id = b_item.get("row_idx")
        conf_score = float(b_item.get("confidence_score", 100.0))
        conf_level = str(b_item.get("confidence_level", "HIGH"))

        ev_raw = b_item.get("evidence", b_item.get("evidence_json", {}))
        if isinstance(ev_raw, str):
            evidence_str = ev_raw
        elif isinstance(ev_raw, dict):
            sources_list = [ev_raw]
            if "additional_sources" in b_item:
                sources_list.extend(b_item["additional_sources"])
            evidence_str = json.dumps({
                "summary": ev_raw,
                "sources": sources_list
            })
        else:
            evidence_str = "{}"

        cursor.execute(
            """
            INSERT INTO boq_items (
                price_list_id, code, name, unit, rate, quantity, action, comments, category, 
                price_item_id, confidence_score, confidence_level, evidence_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                price_list_id,
                code,
                name,
                unit,
                rate,
                quantity,
                action,
                comments,
                category,
                price_item_id,
                conf_score,
                conf_level,
                evidence_str
            )
        )

    conn.commit()
    conn.close()
    
    file_path = get_price_list_path(price_list_id)
    generate_populated_boq_excel(file_path, {}, file_path)
    return {"status": "success", **get_price_list_response(price_list_id)}

@app.post("/api/price-list/clear-quantities-batch")
def clear_quantities_batch(payload: ClearQuantitiesBatchModel, price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Clears quantities for selected item IDs in a batch operation."""
    try:
        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.executemany(
            "UPDATE price_items SET quantity = 0, action = '', comments = '' WHERE id = ?",
            [(rid,) for rid in payload.row_indices]
        )
        conn.commit()
        conn.close()
        
        resolved_id = price_list_id
        if resolved_id is None:
            resolved_id = get_id_from_path(get_price_list_path(price_list_id)) or get_default_price_list_id()
        sync_db_to_active_excel(resolved_id)
        
        clear_user_mappings()
        return {"status": "success", **get_price_list_response(price_list_id)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to clear quantities in batch: {str(e)}")

@app.post("/api/price-list")
def add_price_item(item: PriceItemModel, price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Adds a new pricing item to the database."""
    file_path = get_price_list_path(price_list_id)
    success = add_price_item_to_excel(file_path, item.code, item.name, item.unit, item.rate, item.category)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to add pricing item to database.")
    
    # Batch action: sync database state to disk
    resolved_id = get_id_from_path(file_path) or get_default_price_list_id()
    sync_db_to_active_excel(resolved_id)
    
    clear_user_mappings()
    return {"status": "success", **get_price_list_response(price_list_id)}

@app.put("/api/price-list/{row_idx}")
def update_price_item(row_idx: int, item: PriceItemUpdateModel, price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Updates an item in the database."""
    file_path = get_price_list_path(price_list_id)
    success = update_price_item_in_excel(file_path, row_idx, item.code, item.name, item.unit, item.rate, item.category)
    if not success:
        raise HTTPException(status_code=500, detail=f"Failed to update item {row_idx} in database.")
    
    # Batch action: sync database state to disk
    resolved_id = get_id_from_path(file_path) or get_default_price_list_id()
    sync_db_to_active_excel(resolved_id)
    
    clear_user_mappings()
    return {"status": "success", **get_price_list_response(price_list_id)}

@app.put("/api/price-list/category/{category_name}")
def rename_category(category_name: str, payload: CategoryRenameModel, price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Renames a category across all items in SQLite."""
    try:
        from services.db import get_db_connection
        import urllib.parse
        
        # Decode URL-encoded category name
        decoded_name = urllib.parse.unquote(category_name)
        new_name_clean = payload.new_name.strip()
        
        if not new_name_clean:
            raise HTTPException(status_code=400, detail="New category name cannot be empty.")
            
        resolved_id = price_list_id
        if resolved_id is None:
            resolved_id = get_id_from_path(get_price_list_path()) or get_default_price_list_id()
            
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE price_items SET category = ? WHERE category = ? AND price_list_id = ?",
            (new_name_clean, decoded_name, resolved_id)
        )
        conn.commit()
        conn.close()
        
        # Batch action: sync database state to disk
        sync_db_to_active_excel(resolved_id)
        
        clear_user_mappings()
        return {"status": "success", **get_price_list_response(price_list_id)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to rename category: {str(e)}")

@app.delete("/api/price-list/{row_idx}")
def delete_price_item(row_idx: int, price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Deletes an item from the database."""
    file_path = get_price_list_path(price_list_id)
    success = clear_price_item_in_excel(file_path, row_idx)
    if not success:
        raise HTTPException(status_code=500, detail=f"Failed to delete item {row_idx} from database.")
    
    # Batch action: sync database state to disk
    resolved_id = get_id_from_path(file_path) or get_default_price_list_id()
    sync_db_to_active_excel(resolved_id)
    
    clear_user_mappings()
    return {"status": "success", **get_price_list_response(price_list_id)}

@app.post("/api/price-list/clear")
def clear_price_list(price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Clears all records in the SQLite pricing items database."""
    file_path = get_price_list_path(price_list_id)
    success = clear_all_price_items_in_excel(file_path)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to clear database.")
    
    # Batch action: sync database state to disk
    resolved_id = get_id_from_path(file_path) or get_default_price_list_id()
    sync_db_to_active_excel(resolved_id)
    
    clear_user_mappings()
    return {"status": "success", **get_price_list_response(price_list_id)}

@app.post("/api/price-list/delete-batch")
def delete_price_items_batch(payload: DeleteBatchModel, price_list_id: Optional[int] = None) -> Dict[str, Any]:
    """Clears values for multiple row indices in Excel."""
    file_path = get_price_list_path(price_list_id)
    success = clear_price_items_in_excel_batch(file_path, payload.row_indices)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to clear selected row indices.")
    
    # Batch action: sync database state to disk
    resolved_id = get_id_from_path(file_path) or get_default_price_list_id()
    sync_db_to_active_excel(resolved_id)
    
    clear_user_mappings()
    return {"status": "success", **get_price_list_response(price_list_id)}

@app.get("/api/settings/confidence")
def get_confidence_settings() -> Dict[str, Any]:
    """Returns configurable confidence threshold settings."""
    from services.db import get_confidence_thresholds
    return {"status": "success", "thresholds": get_confidence_thresholds()}

@app.post("/api/settings/confidence")
def save_confidence_settings(payload: Dict[str, float]) -> Dict[str, Any]:
    """Updates configurable confidence threshold settings."""
    from services.db import update_confidence_thresholds, get_confidence_thresholds
    success = update_confidence_thresholds(payload)
    if not success:
        raise HTTPException(status_code=500, detail="Failed to update confidence thresholds.")
    return {"status": "success", "thresholds": get_confidence_thresholds()}

@app.get("/api/rules/pending")
def get_pending_rules():
    """Returns all proposed rules that are currently pending human review."""
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mapping_rules WHERE status = 'PENDING_REVIEW' ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/rules/approve/{rule_id}")
def approve_rule(rule_id: int, approved_by: Optional[str] = "Estimator") -> Dict[str, Any]:
    """Approves a pending rule, activating it and deprecating/superseding its parent rule if relevant."""
    from services.db import get_db_connection
    import datetime
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT parent_rule_id, rule_name FROM mapping_rules WHERE id = ?", (rule_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail=f"Rule #{rule_id} not found.")
    
    parent_id = row["parent_rule_id"]
    now_str = datetime.datetime.now().isoformat()
    
    # 1. Update pending rule to ACTIVE
    cursor.execute(
        "UPDATE mapping_rules SET status = 'ACTIVE', approved_by = ?, approved_at = ? WHERE id = ?",
        (approved_by, now_str, rule_id)
    )
    
    # 2. Update parent rule (if any) to SUPERSEDED
    if parent_id:
        cursor.execute(
            "UPDATE mapping_rules SET status = 'SUPERSEDED' WHERE id = ?",
            (parent_id,)
        )
        
    conn.commit()
    conn.close()
    return {"status": "success", "message": f"Rule #{rule_id} approved and activated."}

@app.post("/api/rules/reject/{rule_id}")
def reject_rule(rule_id: int, rejected_by: Optional[str] = "Estimator") -> Dict[str, Any]:
    """Rejects a proposed rule, marking it as REJECTED."""
    from services.db import get_db_connection
    import datetime
    
    now_str = datetime.datetime.now().isoformat()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE mapping_rules SET status = 'REJECTED', rejected_by = ?, rejected_at = ? WHERE id = ?",
        (rejected_by, now_str, rule_id)
    )
    affected = cursor.rowcount > 0
    conn.commit()
    conn.close()
    
    if not affected:
        raise HTTPException(status_code=404, detail=f"Rule #{rule_id} not found.")
    return {"status": "success", "message": f"Rule #{rule_id} marked as REJECTED."}

@app.post("/api/rules/{rule_id}/simulate")
def simulate_rule_endpoint(rule_id: int) -> Dict[str, Any]:
    """Regenerates simulation stats for a rule against the latest historical mappings."""
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT conditions_json, actions_json, target_sor_code, target_sor_name FROM mapping_rules WHERE id = ?", (rule_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail=f"Rule #{rule_id} not found.")
        
    import json
    try:
        c_json = json.loads(row["conditions_json"])
        a_json = json.loads(row["actions_json"])
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=f"Failed to parse rule JSON: {e}")
        
    from services.ai_service import run_proposed_rule_simulation
    sim_stats = run_proposed_rule_simulation(
        c_json,
        a_json,
        row["target_sor_code"],
        row["target_sor_name"]
    )
    
    cursor.execute(
        "UPDATE mapping_rules SET simulation_stats = ? WHERE id = ?",
        (json.dumps(sim_stats), rule_id)
    )
    conn.commit()
    conn.close()
    
    return {"status": "success", "simulation_stats": sim_stats}

@app.get("/api/rules/{rule_id}")
def get_rule_details(rule_id: int):
    """Fetches details of a specific rule by ID."""
    from services.db import get_mapping_rule_by_id
    rule = get_mapping_rule_by_id(rule_id)
    if not rule:
        raise HTTPException(status_code=404, detail=f"Rule #{rule_id} not found.")
    return rule

@app.get("/api/rules/history")
def get_rules_history():
    """Returns rule execution and lifecycle history (ACTIVE, SUPERSEDED, REJECTED, DISABLED, PENDING_REVIEW)."""
    from services.db import get_db_connection
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mapping_rules ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ==========================================
# EQUIPMENT MASTER CATALOG ENDPOINTS
# ==========================================

@app.get("/api/equipment-catalog")
def list_equipment_catalog(
    equipment_class: Optional[str] = None,
    manufacturer: Optional[str] = None,
    search: Optional[str] = None
):
    """Lists all canonical equipment items with optional filtering."""
    from services.equipment_service import get_all_equipment
    return get_all_equipment(equipment_class=equipment_class, manufacturer=manufacturer, search=search)

@app.get("/api/equipment-catalog/{item_id}")
def get_equipment_item(item_id: int):
    """Retrieves a single canonical equipment record by ID."""
    from services.equipment_service import get_equipment_by_id
    item = get_equipment_by_id(item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Equipment item not found.")
    return item

@app.post("/api/equipment-catalog")
def create_equipment_item(payload: Dict[str, Any]):
    """Creates a new canonical equipment item with aliases."""
    from services.equipment_service import create_equipment
    from services.entity_resolution import get_entity_resolver
    item_id = create_equipment(payload)
    get_entity_resolver().reload()
    return {"status": "success", "id": item_id}

@app.put("/api/equipment-catalog/{item_id}")
def update_equipment_item(item_id: int, payload: Dict[str, Any]):
    """Updates an existing canonical equipment record."""
    from services.equipment_service import update_equipment
    from services.entity_resolution import get_entity_resolver
    success = update_equipment(item_id, payload)
    if not success:
        raise HTTPException(status_code=404, detail="Equipment item not found or update failed.")
    get_entity_resolver().reload()
    return {"status": "success", "id": item_id}

@app.delete("/api/equipment-catalog/{item_id}")
def delete_equipment_item(item_id: int):
    """Deletes an equipment record from the catalog."""
    from services.equipment_service import delete_equipment
    from services.entity_resolution import get_entity_resolver
    success = delete_equipment(item_id)
    if not success:
        raise HTTPException(status_code=404, detail="Equipment item not found.")
    get_entity_resolver().reload()
    return {"status": "success", "id": item_id}

@app.post("/api/equipment-catalog/seed")
def seed_equipment_catalog_endpoint(force: bool = False):
    """Seeds default canonical equipment into the database."""
    from services.equipment_service import seed_default_equipment_catalog
    from services.entity_resolution import get_entity_resolver
    count = seed_default_equipment_catalog(force=force)
    get_entity_resolver().reload()
    return {"status": "success", "seeded_count": count}

@app.post("/api/equipment-catalog/add-alias")
def add_equipment_alias_endpoint(payload: Dict[str, Any]):
    """Adds a new drawing alias pattern to an existing equipment record."""
    from services.equipment_service import add_alias_to_equipment
    from services.entity_resolution import get_entity_resolver
    canonical_id = payload.get("canonical_id")
    new_alias = payload.get("alias")
    if not canonical_id or not new_alias:
        raise HTTPException(status_code=400, detail="canonical_id and alias are required.")
    success = add_alias_to_equipment(canonical_id, new_alias)
    if not success:
        raise HTTPException(status_code=404, detail=f"Canonical equipment '{canonical_id}' not found.")
    get_entity_resolver().reload()
    return {"status": "success", "canonical_id": canonical_id, "alias": new_alias}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)


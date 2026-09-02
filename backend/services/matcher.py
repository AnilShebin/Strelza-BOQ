"""
Matcher Service.
Provides fuzzy name search matching for equipment items against the Master Schedule
of Rates (SOR) spreadsheet, utilizing rapidfuzz and a local corrections cache.
"""
import os
import openpyxl
import json
import threading
from typing import List, Dict, Any, Optional
from rapidfuzz import fuzz

import re
from collections import defaultdict

MAPPINGS_CACHE_FILE = os.path.join(os.path.dirname(__file__), "user_mappings.json")
excel_lock = threading.RLock()
price_list_locks = defaultdict(threading.RLock)

def get_id_from_path(file_path: str) -> Optional[int]:
    if not file_path:
        return None
    match = re.search(r"price_list_(\d+)\.xlsx", file_path)
    if match:
        return int(match.group(1))
    return None

def get_price_list_path(price_list_id: Optional[int] = None) -> str:
    """Returns the Excel file path for the given price list ID. Falls back to the globally active pointer if none."""
    if price_list_id is None:
        try:
            from services.db import get_db_connection, get_default_price_list_id
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM price_lists WHERE is_active = 1")
            row = cursor.fetchone()
            conn.close()
            price_list_id = row["id"] if row else get_default_price_list_id()
        except Exception:
            from services.db import get_default_price_list_id
            price_list_id = get_default_price_list_id()
            
    return os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads", f"price_list_{price_list_id}.xlsx")

def locked_excel_write(func):
    def wrapper(*args, **kwargs):
        file_path = None
        if len(args) > 0 and isinstance(args[0], str):
            file_path = args[0]
        elif "file_path" in kwargs:
            file_path = kwargs["file_path"]
        elif "template_path" in kwargs:
            file_path = kwargs["template_path"]
            
        lock_id = get_id_from_path(file_path) if file_path else None
        if lock_id is not None:
            lock = price_list_locks[lock_id]
        else:
            lock = excel_lock
            
        with lock:
            return func(*args, **kwargs)
    return wrapper

def load_user_mappings() -> Dict[str, Any]:
    """Loads user-corrected item mappings from local cache file."""
    if os.path.exists(MAPPINGS_CACHE_FILE):
        try:
            with open(MAPPINGS_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_user_mapping(raw_desc: str, matched_code: str, matched_name: str, rate: float) -> None:
    """Saves a single user-corrected price code alignment to local cache."""
    mappings = load_user_mappings()
    mappings[raw_desc.upper()] = {
        "code": matched_code,
        "name": matched_name,
        "rate": rate
    }
    try:
        with open(MAPPINGS_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(mappings, f, indent=2)
    except Exception as e:
        print(f"[Matcher] Error saving user mapping: {e}")

def clear_user_mappings() -> None:
    """Clears all user-corrected item mappings from the local cache."""
    if os.path.exists(MAPPINGS_CACHE_FILE):
        try:
            os.remove(MAPPINGS_CACHE_FILE)
        except Exception as e:
            print(f"[Matcher] Error removing mappings cache: {e}")
            try:
                with open(MAPPINGS_CACHE_FILE, "w", encoding="utf-8") as f:
                    json.dump({}, f)
            except Exception:
                pass

def evaluate_sheet_formulas(sheet) -> dict:
    """
    Parses and evaluates simple spreadsheet formulas in-memory:
    1. =D{row}*E{row}
    2. =SUM(F{start}:F{end})
    3. Sum chain formulas like =F249+F241+F234+...
    """
    raw_grid = {}
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            raw_grid[(r, c)] = sheet.cell(r, c).value

    def get_cell_float(row, col):
        val = raw_grid.get((row, col))
        if val is None:
            return 0.0
        if isinstance(val, str) and val.startswith("="):
            return 0.0
        try:
            val_str = str(val).replace('$', '').replace(',', '').strip()
            return float(val_str)
        except ValueError:
            return 0.0

    # Pass 1: Evaluate multiplications for all item rows (Column D rate * Column E qty)
    for r in range(1, sheet.max_row + 1):
        val = raw_grid.get((r, 6))
        # Skip if it is a SUM formula or chain total
        if isinstance(val, str) and (val.startswith("=SUM") or "+" in val or "#REF" in val):
            continue
            
        code_val = str(raw_grid.get((r, 1)) or "").strip()
        rate_val = get_cell_float(r, 4)
        if rate_val > 0 or code_val != "":
            qty_val = get_cell_float(r, 5)
            raw_grid[(r, 6)] = rate_val * qty_val

    # Pass 2: Evaluate SUM ranges (=SUM(F{start}:F{end}))
    for r in range(1, sheet.max_row + 1):
        val = raw_grid.get((r, 6))
        if isinstance(val, str) and val.startswith("="):
            val_upper = val.upper().strip()
            if val_upper.startswith("=SUM(F") and ":" in val_upper:
                try:
                    # Strip letters and parentheses to get bounds (e.g. F5:F31 -> 5:31)
                    clean = val_upper.replace("=", "").replace("SUM", "").replace("(", "").replace(")", "").replace("F", "")
                    parts = clean.split(":")
                    start_row = int(parts[0])
                    end_row = int(parts[1])
                    section_sum = sum(get_cell_float(i, 6) for i in range(start_row, end_row + 1))
                    raw_grid[(r, 6)] = section_sum
                except Exception:
                    pass

    # Pass 3: Evaluate section sum chains
    for r in range(1, sheet.max_row + 1):
        val = raw_grid.get((r, 6))
        if isinstance(val, str) and val.startswith("="):
            val_upper = val.upper().strip()
            if val_upper.startswith("=F") and "+" in val_upper:
                try:
                    terms = val_upper.replace("=", "").split("+")
                    total = 0.0
                    for term in terms:
                        term = term.strip()
                        if term.startswith("F") and term[1:].isdigit():
                            target_row = int(term[1:])
                            total += get_cell_float(target_row, 6)
                    raw_grid[(r, 6)] = total
                except Exception:
                    pass

    return raw_grid

def load_master_price_list(file_path: str = "", price_list_id: Optional[int] = None) -> List[Dict[str, Any]]:
    """Loads SOR pricing items directly from the SQLite database."""
    try:
        if price_list_id is None and file_path:
            price_list_id = get_id_from_path(file_path)
        if price_list_id is None:
            price_list_id = get_id_from_path(get_price_list_path())
            if price_list_id is None:
                from services.db import get_default_price_list_id
                price_list_id = get_default_price_list_id()

        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, code, name, unit, rate, quantity, category, action, comments, confidence_score, confidence_level, evidence_json, attributes_json FROM price_items WHERE price_list_id = ? ORDER BY id",
            (price_list_id,)
        )
        rows = cursor.fetchall()
        conn.close()
        
        price_items = []
        current_category = None
        
        for r in rows:
            cat = r["category"] or "General SOR Pricing Items"
            if cat != current_category:
                current_category = cat
                price_items.append({
                    "row_idx": 0,
                    "row_type": "section_header",
                    "code": "",
                    "name": current_category,
                    "unit": "",
                    "rate": 0.0,
                    "cells": [current_category, "", "", "", "", "", "", ""]
                })
                
            qty_val = r["quantity"]
            qty_str = ""
            if qty_val:
                qty_str = str(int(qty_val)) if float(qty_val).is_integer() else str(qty_val)
                
            total_val = r["rate"] * qty_val if r["rate"] and qty_val else 0.0
            total_str = f"${total_val:,.2f}" if total_val > 0 else "$-"
            rate_str = f"${r['rate']:,.2f}" if r["rate"] else ""
            
            price_items.append({
                "row_idx": r["id"],
                "row_type": "data_item",
                "code": r["code"] or "",
                "name": r["name"] or "",
                "unit": r["unit"] or "",
                "rate": r["rate"] or 0.0,
                "category": r["category"] or "General SOR Pricing Items",
                "action": r["action"] or "",
                "comments": r["comments"] or "",
                "confidence_score": r["confidence_score"] if "confidence_score" in r.keys() else 100.0,
                "confidence_level": r["confidence_level"] if "confidence_level" in r.keys() else "HIGH",
                "evidence_json": r["evidence_json"] if "evidence_json" in r.keys() else "",
                "attributes_json": r["attributes_json"] if "attributes_json" in r.keys() else "",
                "cells": [
                    r["code"] or "",
                    r["name"] or "",
                    r["unit"] or "",
                    rate_str,
                    qty_str,
                    total_str,
                    r["comments"] or ""
                ]
            })
            
        return price_items
    except Exception as e:
        print(f"[Matcher] Error loading SQLite price list for ID {price_list_id}: {e}")
        return []

def make_canonical_takeoff_profile(item: Dict[str, Any]) -> Dict[str, Any]:
    text_to_parse = f"{item.get('equipment_type', '')} {item.get('model', '')} {item.get('raw_text', '')}"
    det = deterministic_parse_row(text_to_parse, item.get("unit", "each"))
    
    # Override action if present in item
    if item.get("action"):
        det["commercial_action"] = item["action"].upper()
        det["compatible_actions"] = [item["action"].upper()]
        
    return {
        "semantic_class": det["semantic_class"],
        "commercial_action": det["commercial_action"],
        "compatible_actions": det["compatible_actions"],
        "quantity_basis": det["quantity_basis"],
        "commercial_basis": det["commercial_basis"],
        "attributes": det["attributes"],
        "critical_attributes": det["critical_attributes"],
        "optional_attributes": det["optional_attributes"],
        "quantity": float(item.get("quantity", 1.0)),
        "unit": item.get("unit", "each")
    }

QUANTITY_FAMILIES = {
    "LM": ["LM", "PER_6LM"],
    "PER_6LM": ["LM", "PER_6LM"],
    "M3": ["M3", "PER_M3"],
    "PER_M3": ["M3", "PER_M3"],
    "HOUR": ["HOUR", "WEEK", "DAY"],
    "WEEK": ["HOUR", "WEEK", "DAY"],
    "DAY": ["HOUR", "WEEK", "DAY"],
    "EACH": ["EACH", "PER_SITE", "PER_SECTOR"],
    "PER_SITE": ["EACH", "PER_SITE", "PER_SECTOR"],
    "PER_SECTOR": ["EACH", "PER_SITE", "PER_SECTOR"]
}

def normalize_unit_string(unit: str) -> str:
    u = unit.strip().lower()
    if u in ["each", "ea", "no.", "nr", "item", "no", "nr"]:
        return "EACH"
    if u in ["lm", "linear metre", "lnm", "linear meter", "m"]:
        return "LM"
    if u in ["m3", "m³", "cubic metre", "cubic meter"]:
        return "M3"
    if u in ["m2", "m²", "sqm", "square metre", "square meter"]:
        return "M2"
    if u in ["hour", "hr", "hours", "ph"]:
        return "HOUR"
    if u in ["week", "wk", "weeks"]:
        return "WEEK"
    if u in ["day", "days"]:
        return "DAY"
    if u in ["km", "kilometre", "kilometer"]:
        return "KM"
    if u in ["night", "nights"]:
        return "NIGHT"
    if "site" in u:
        return "PER_SITE"
    if "sector" in u:
        return "PER_SECTOR"
    if "6lm" in u or "6 lm" in u:
        return "PER_6LM"
    if "/m3" in u:
        return "PER_M3"
    return "EACH"

NO_MATCH_THRESHOLDS = {
    "STRUCTURE": 60.0,
    "FOUNDATION": 65.0,
    "FEEDER": 70.0,
    "ANTENNA": 55.0,
    "LABOUR": 45.0,
    "ACTIVITY": 40.0,
    "default": 50.0
}

def match_item_to_price_list(
    item: Dict[str, Any],
    price_list: List[Dict[str, Any]],
    threshold: float = 60.0
) -> Optional[Dict[str, Any]]:
    """Performs fuzzy schema-driven matching on equipment descriptions to locate active SOR codes."""
    equip_type = (item.get("equipment_type") or "").upper()
    raw_desc = (item.get("model") or item.get("raw_text") or "").upper()
    search_query = f"{equip_type} {raw_desc}".strip()
    user_mappings = load_user_mappings()
    
    prior_correction_code = None
    if raw_desc in user_mappings:
        prior_correction_code = user_mappings[raw_desc]["code"]
        
    if not price_list:
        return None
        
    active_price_list = [
        p for p in price_list 
        if p.get("row_type", "data_item") == "data_item" and (p.get("code") or p.get("name"))
    ]
    if not active_price_list:
        return None
        
    # Convert item to canonical profile
    takeoff = make_canonical_takeoff_profile(item)
    
    # 1. Candidate Retrieval
    candidates = []
    for p in active_price_list:
        # Load or compute profile
        prof = {}
        if p.get("evidence_json") and "semantic_class" in p.get("evidence_json"):
            try:
                prof = json.loads(p["evidence_json"])
            except Exception:
                pass
        if not prof and p.get("profile_json"):
            try:
                prof = json.loads(p["profile_json"])
            except Exception:
                pass
        if not prof:
            prof = deterministic_parse_row(p["name"], p["unit"])
            prof["sor_code"] = p["code"]
            prof["description"] = p["name"]
            
        p_profile = prof
        
        # Fast pre-filters:
        class_match = p_profile.get("semantic_class") == takeoff["semantic_class"]
        
        # Word intersection (strip non-alphanumeric to resolve parenthesized tags like (RRU))
        q_tokens = {re.sub(r'\W+', '', w) for w in search_query.split() if re.sub(r'\W+', '', w)}
        p_tokens = {re.sub(r'\W+', '', w) for w in p["name"].upper().split() if re.sub(r'\W+', '', w)}
        shared_tokens = q_tokens.intersection(p_tokens)
        
        # Always retrieve if it matches prior correction code
        is_prior = prior_correction_code and p["code"] == prior_correction_code
        
        if class_match or shared_tokens or is_prior:
            sim = fuzz.token_sort_ratio(search_query, p["name"].upper())
            candidates.append((p, p_profile, sim))
            
    # Sort and take top 50 candidates
    candidates.sort(key=lambda x: x[2], reverse=True)
    shortlist = candidates[:50]
    
    # 2. Hard Constraint Gates
    valid_matches = []
    for p, p_prof, base_sim in shortlist:
        # Action Gate
        t_action = takeoff["commercial_action"]
        p_actions = p_prof.get("compatible_actions", [p_prof.get("commercial_action", "INSTALL")])
        if t_action not in p_actions and not (t_action == "INSTALL" and "INSTALL" in p_actions):
            continue
            
        # Unit Ontology Gate
        t_unit_norm = normalize_unit_string(takeoff["unit"])
        p_unit_norm = normalize_unit_string(p["unit"])
        if t_unit_norm != p_unit_norm:
            # Check family compatibility
            t_family = next((fam for fam, units in QUANTITY_FAMILIES.items() if t_unit_norm in units), None)
            p_family = next((fam for fam, units in QUANTITY_FAMILIES.items() if p_unit_norm in units), None)
            if t_family != p_family:
                continue # Incompatible unit families
                
        # Critical Attributes Gate
        attr_mismatch = False
        for attr_key in p_prof.get("critical_attributes", []):
            if attr_key in p_prof.get("attributes", {}) and attr_key in takeoff["attributes"]:
                p_val = p_prof["attributes"][attr_key]
                t_val = takeoff["attributes"][attr_key]
                
                # Check height range or exact values
                if attr_key == "structure_height_m":
                    if abs(float(p_val) - float(t_val)) > 1.5:
                        attr_mismatch = True
                        break
                elif attr_key == "cable_count":
                    if int(p_val) != int(t_val):
                        attr_mismatch = True
                        break
                elif attr_key == "volume_m3":
                    if abs(float(p_val) - float(t_val)) > 1.0:
                        attr_mismatch = True
                        break
                else:
                    if str(p_val).strip().upper() != str(t_val).strip().upper():
                        attr_mismatch = True
                        break
                        
        if attr_mismatch:
            continue
            
        # 3. Dynamic Feature Matcher
        score = base_sim
        
        # Attribute matching bonuses / penalties
        crit_matched = 0
        crit_total = len(p_prof.get("critical_attributes", []))
        for attr_key in p_prof.get("critical_attributes", []):
            if attr_key in takeoff["attributes"]:
                crit_matched += 1
                
        if crit_total > 0:
            if crit_matched == crit_total:
                score += 20.0
            else:
                score -= 10.0
                
        # Prior correction boost
        is_prior = prior_correction_code and p["code"] == prior_correction_code
        if is_prior:
            score += 35.0
            
        final_score = min(max(score, 0.0), 100.0)
        
        valid_matches.append({
            "code": p["code"],
            "name": p["name"],
            "unit": p["unit"],
            "rate": p["rate"],
            "similarity": round(final_score, 2),
            "row_idx": p.get("row_idx"),
            "profile": p_prof
        })
        
    valid_matches.sort(key=lambda x: x["similarity"], reverse=True)
    
    # 4. Decision Outcomes
    top_match = valid_matches[0] if valid_matches else None
    
    if not top_match:
        return {
            "code": "UNQUOTED",
            "name": "No matching SOR found",
            "unit": item.get("unit", "each"),
            "rate": 0.0,
            "similarity": 0.0,
            "auto_matched": False,
            "row_idx": None,
            "comments": "No matching SOR found — Estimator to provide rate"
        }
        
    top_score = top_match["similarity"]
    sem_class = takeoff["semantic_class"]
    no_match_threshold = min(threshold, NO_MATCH_THRESHOLDS.get(sem_class, NO_MATCH_THRESHOLDS["default"]))
    
    if top_score < no_match_threshold:
        return {
            "code": "UNQUOTED",
            "name": top_match["name"],
            "unit": top_match["unit"],
            "rate": 0.0,
            "similarity": top_score,
            "auto_matched": False,
            "row_idx": None,
            "comments": "No matching SOR found — Estimator to provide rate"
        }
        
    # Check ambiguity margin
    is_ambiguous = False
    if len(valid_matches) > 1:
        # Find the next best option that represents a different commercial item
        different_options = [
            m for m in valid_matches 
            if m["code"] != top_match["code"] and m["code"] != "UNQUOTED"
        ]
        if different_options:
            second_match = different_options[0]
            margin = top_score - second_match["similarity"]
            if margin < 5.0 and top_score < 95.0:
                is_ambiguous = True
            
    top_match_copy = dict(top_match)
    top_match_copy["auto_matched"] = not is_ambiguous
    if is_ambiguous:
        top_match_copy["comments"] = "Ambiguous match (low margin) — review required"
        
    top_match_copy["options"] = [dict(m) for m in valid_matches[:5]]
    return top_match_copy

def save_master_price_list(file_path: str, items: List[Dict[str, Any]]) -> bool:
    """
    Saves the list of items back to the master price list excel file.
    Preserves headers and layout structure.
    """
    try:
        import openpyxl
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            if "SOR" in wb.sheetnames:
                sheet = wb["SOR"]
            elif "Schedule of Rates" in wb.sheetnames:
                sheet = wb["Schedule of Rates"]
            elif "Schedule Of Rates" in wb.sheetnames:
                sheet = wb["Schedule Of Rates"]
            else:
                sheet = wb.active
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Schedule of Rates"
            sheet.append(["Code", "Item Name", "Unit of Qty", "Rate Excluding GST"])
            
        header_row_idx = 1
        code_col, name_col, unit_col, rate_col = 1, 2, 3, 4
        
        # Identify headers row dynamically
        for row in range(1, 15):
            row_values = [str(sheet.cell(row, col).value or "").strip().upper() for col in range(1, 10)]
            if any("ITEM NAME" in val or "RATE" in val or "EXCLUDING GST" in val for val in row_values):
                header_row_idx = row
                for col_idx, val in enumerate(row_values, 1):
                    if any(x in val for x in ["CODE", "WCODE", "SOR"]):
                        code_col = col_idx
                    elif "ITEM NAME" in val or "DESCRIPTION" in val:
                        name_col = col_idx
                    elif "UNIT" in val:
                        unit_col = col_idx
                    elif "RATE" in val or "PRICE" in val:
                        rate_col = col_idx
                break

        # Delete all rows below header row
        max_row = sheet.max_row
        if max_row > header_row_idx:
            sheet.delete_rows(header_row_idx + 1, max_row - header_row_idx)
            
        # Append updated items
        for item in items:
            row_idx = sheet.max_row + 1
            sheet.cell(row=row_idx, column=code_col, value=item.get("code", ""))
            sheet.cell(row=row_idx, column=name_col, value=item.get("name", ""))
            sheet.cell(row=row_idx, column=unit_col, value=item.get("unit", ""))
            
            # Rate column gets the float rate
            sheet.cell(row=row_idx, column=rate_col, value=item.get("rate", 0.0))
            
        wb.save(file_path)
        wb.close()
        return True
    except Exception as e:
        print(f"[Matcher] Error saving master price list: {e}")
        return False

@locked_excel_write
def update_price_item_in_excel(file_path: str, row_idx: int, code: str, name: str, unit: str, rate: float, category: str = "General SOR Pricing Items") -> bool:
    """Updates the details of a specific item in the SQLite database by ID."""
    try:
        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        unit_clean = unit.strip().lower()
        cursor.execute(
            "UPDATE price_items SET code = ?, name = ?, unit = ?, rate = ?, category = ? WHERE id = ?",
            (code, name, unit_clean, rate, category, row_idx)
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"[Matcher] Error updating row {row_idx} in SQLite: {e}")
        return False

@locked_excel_write
def clear_price_item_in_excel(file_path: str, row_idx: int) -> bool:
    """Deletes a specific pricing item from the SQLite database by ID."""
    try:
        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM price_items WHERE id = ?", (row_idx,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"[Matcher] Error deleting row {row_idx} in SQLite: {e}")
        return False

@locked_excel_write
def add_price_item_to_excel(file_path: str, code: str, name: str, unit: str, rate: float, category: str = "General SOR Pricing Items") -> bool:
    """Appends a new pricing item to the SQLite database."""
    try:
        price_list_id = get_id_from_path(file_path) or 1
        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        unit_clean = unit.strip().lower()
        cursor.execute(
            "INSERT INTO price_items (code, name, unit, rate, category, price_list_id) VALUES (?, ?, ?, ?, ?, ?)",
            (code, name, unit_clean, rate, category, price_list_id)
        )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"[Matcher] Error appending item to SQLite: {e}")
        return False

@locked_excel_write
def generate_populated_boq_excel(template_path: str, matched_quantities: dict, output_path: str, only_priced: bool = False) -> bool:
    """Generates a fresh, clean, beautifully styled Excel BOQ workbook directly from SQLite data with proper formulas and filters."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from services.db import get_db_connection

        price_list_id = get_id_from_path(template_path) or 1

        # Initialize fresh workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Schedule of Rates"
        sheet.views.sheetView[0].showGridLines = True

        # Style templates
        font_family = "Segoe UI"
        
        # Branding Color Palette
        color_graphite = "2B3036"
        color_ink = "1F376A"
        color_tropical = "F27E20"
        color_red_bird = "EE4324"
        color_sand = "FFF3EA"
        color_border = "CBD5E1"
        color_active_green = "E2F0D9"    # Soft green fill for priced rows

        font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        font_category = Font(name=font_family, size=11, bold=True, color=color_ink)
        font_data = Font(name=font_family, size=10, color=color_graphite)
        font_data_bold = Font(name=font_family, size=10, bold=True, color=color_graphite)
        
        fill_header = PatternFill(start_color=color_ink, end_color=color_ink, fill_type="solid")
        fill_category = PatternFill(start_color=color_sand, end_color=color_sand, fill_type="solid")
        fill_total = PatternFill(start_color=color_sand, end_color=color_sand, fill_type="solid")
        fill_priced_row = PatternFill(start_color=color_active_green, end_color=color_active_green, fill_type="solid")

        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        border_thin = Border(
            left=Side(style="thin", color=color_border),
            right=Side(style="thin", color=color_border),
            top=Side(style="thin", color=color_border),
            bottom=Side(style="thin", color=color_border)
        )
        border_total = Border(
            top=Side(style="thin", color=color_ink),
            bottom=Side(style="double", color=color_ink) # Double underline
        )

        # Title Block
        sheet.row_dimensions[1].height = 35
        sheet.merge_cells("A1:H1")
        title_cell = sheet.cell(row=1, column=1, value="SCHEDULE OF RATES (SOR) BILL OF QUANTITIES")
        title_cell.font = Font(name=font_family, size=14, bold=True, color=color_ink)
        title_cell.alignment = align_left

        # Table Column Headers
        headers = ["SOR Code", "Item Description", "Action", "Unit", "Rate", "Quantity", "Total Cost", "Comments"]
        header_row_idx = 3
        sheet.row_dimensions[header_row_idx].height = 28

        for col_idx, text in enumerate(headers, 1):
            cell = sheet.cell(row=header_row_idx, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = border_thin
            if text in ["Unit", "Quantity", "Action"]:
                cell.alignment = align_center
            elif text in ["Rate", "Total Cost"]:
                cell.alignment = align_right
            else:
                cell.alignment = align_left

        # Query items
        conn = get_db_connection()
        cursor = conn.cursor()
        if only_priced:
            cursor.execute(
                "SELECT code, name, unit, rate, quantity, category, action, comments FROM boq_items WHERE price_list_id = ? AND quantity > 0 ORDER BY (CASE WHEN code = 'UNQUOTED' THEN 1 ELSE 0 END) ASC, id ASC",
                (price_list_id,)
            )
            db_items = cursor.fetchall()
        else:
            # Query all price items, left joining boq_items to get quantity/action/comments
            cursor.execute(
                """
                SELECT 
                    pi.code, 
                    pi.name, 
                    pi.unit, 
                    pi.rate, 
                    COALESCE(bi.quantity, 0) as quantity, 
                    pi.category, 
                    COALESCE(bi.action, '') as action, 
                    COALESCE(bi.comments, '') as comments
                FROM price_items pi
                LEFT JOIN boq_items bi ON pi.id = bi.price_item_id AND bi.price_list_id = pi.price_list_id
                WHERE pi.price_list_id = ?
                ORDER BY pi.id
                """,
                (price_list_id,)
            )
            db_items = list(cursor.fetchall())
            
            # Query all UNQUOTED items from boq_items and append them at the end
            cursor.execute(
                """
                SELECT code, name, unit, rate, quantity, category, action, comments
                FROM boq_items
                WHERE price_list_id = ? AND code = 'UNQUOTED'
                ORDER BY id
                """,
                (price_list_id,)
            )
            db_items.extend(cursor.fetchall())
        conn.close()

        # Group by category
        from collections import defaultdict
        category_groups = defaultdict(list)
        for item in db_items:
            cat_name = (item["category"] or "General SOR Items").strip()
            category_groups[cat_name].append(item)

        current_row = 4
        import math

        # Write items
        for cat_name, items in category_groups.items():
            # Write Category Header Row
            sheet.row_dimensions[current_row].height = 24
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            
            cat_cell = sheet.cell(row=current_row, column=1, value=cat_name.upper())
            cat_cell.font = font_category
            cat_cell.fill = fill_category
            cat_cell.alignment = align_left
            
            # Apply borders to the merged row cells
            for c in range(1, 9):
                sheet.cell(row=current_row, column=c).border = border_thin
                
            current_row += 1

            # Write Items in Category
            for item in items:
                qty = float(item["quantity"] or 0)
                is_priced = qty > 0
                row_fill = fill_priced_row if is_priced else None

                # Calculate estimated height based on description wrap (B column has width 65, wrap fits ~60 chars)
                desc_text = item["name"] or ""
                lines_count = max(1, math.ceil(len(desc_text) / 60))
                estimated_height = 14 * lines_count + 12
                sheet.row_dimensions[current_row].height = estimated_height
                
                # SOR Code
                c_code = sheet.cell(row=current_row, column=1, value=(item["code"] or ""))
                c_code.font = font_data
                c_code.alignment = align_left
                c_code.border = border_thin
                if row_fill:
                    c_code.fill = row_fill
                
                # Item Description
                c_desc = sheet.cell(row=current_row, column=2, value=desc_text)
                c_desc.font = font_data
                c_desc.alignment = align_wrap_left
                c_desc.border = border_thin
                if row_fill:
                    c_desc.fill = row_fill
                    
                # Action
                c_act = sheet.cell(row=current_row, column=3, value=(item["action"] or ""))
                c_act.font = font_data
                c_act.alignment = align_center
                c_act.border = border_thin
                if row_fill:
                    c_act.fill = row_fill
                
                # Unit
                c_unit = sheet.cell(row=current_row, column=4, value=(item["unit"] or ""))
                c_unit.font = font_data
                c_unit.alignment = align_center
                c_unit.border = border_thin
                if row_fill:
                    c_unit.fill = row_fill
                
                # Rate
                c_rate = sheet.cell(row=current_row, column=5, value=float(item["rate"] or 0))
                c_rate.font = font_data
                c_rate.alignment = align_right
                c_rate.border = border_thin
                c_rate.number_format = "$#,##0.00"
                if row_fill:
                    c_rate.fill = row_fill
                
                # Quantity
                c_qty = sheet.cell(row=current_row, column=6, value=qty)
                c_qty.font = font_data_bold if is_priced else font_data
                c_qty.alignment = align_center
                c_qty.border = border_thin
                c_qty.number_format = "#,##0"
                if row_fill:
                    c_qty.fill = row_fill
                
                # Total Cost Formula (=Rate * Quantity)
                c_total = sheet.cell(row=current_row, column=7, value=f"=E{current_row}*F{current_row}")
                c_total.font = font_data_bold
                c_total.alignment = align_right
                c_total.border = border_thin
                c_total.number_format = "$#,##0.00"
                if row_fill:
                    c_total.fill = row_fill

                # Comments
                c_comments = sheet.cell(row=current_row, column=8, value=(item["comments"] or ""))
                c_comments.font = font_data
                c_comments.alignment = align_left
                c_comments.border = border_thin
                if row_fill:
                    c_comments.fill = row_fill
                
                current_row += 1

        # Summary Row (TOTAL AMOUNT)
        if current_row > 4:
            sheet.row_dimensions[current_row].height = 26
            
            # Label
            label_cell = sheet.cell(row=current_row, column=2, value="TOTAL ESTIMATED COST (EXCLUDING GST)")
            label_cell.font = Font(name=font_family, size=10, bold=True, color=color_ink)
            label_cell.alignment = align_right
            
            # Formula cell
            total_formula_cell = sheet.cell(row=current_row, column=7, value=f"=SUM(G4:G{current_row-1})")
            total_formula_cell.font = Font(name=font_family, size=11, bold=True, color=color_ink)
            total_formula_cell.fill = fill_total
            total_formula_cell.alignment = align_right
            total_formula_cell.number_format = "$#,##0.00"
            total_formula_cell.border = border_total
            
            # Apply light sand background to summary row
            for c in range(1, 9):
                if c != 7:
                    sheet.cell(row=current_row, column=c).fill = fill_total
                    sheet.cell(row=current_row, column=c).border = Border(top=Side(style="thin", color=color_ink))

            # Enable AutoFilter on column headers grid
            sheet.auto_filter.ref = f"A{header_row_idx}:H{current_row-1}"

        # Custom Column Widths
        column_widths = {
            "A": 15,  # SOR Code
            "B": 65,  # Item Description
            "C": 15,  # Action
            "D": 10,  # Unit
            "E": 15,  # Rate
            "F": 12,  # Quantity
            "G": 20,  # Total Cost
            "H": 30   # Comments
        }

        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width

        wb.save(output_path)
        wb.close()
        return True
    except Exception as e:
        print(f"[Matcher] Error generating fresh BOQ Excel: {e}")
        return False

def write_instructions_sheet(sheet) -> None:
    """Populates the Excel worksheet with beautifully styled DOs and DONTs rules."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    sheet.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    color_ink = "1F376A"
    color_border = "D0D7DE"
    
    border_thin = Border(
        left=Side(style="thin", color=color_border),
        right=Side(style="thin", color=color_border),
        top=Side(style="thin", color=color_border),
        bottom=Side(style="thin", color=color_border)
    )
    
    # Title Block
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells("A1:C1")
    title_cell = sheet.cell(row=1, column=1, value="PRICING SPREADSHEET IMPORT RULES & INSTRUCTIONS")
    title_cell.font = Font(name=font_family, size=13, bold=True, color=color_ink)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Headers
    sheet.row_dimensions[3].height = 24
    headers = ["RULE CATEGORY", "RECOMMENDED ACTIONS (DOs)", "THINGS TO AVOID (DONTs)"]
    for col_idx, text in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_idx, value=text)
        cell.font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_ink, end_color=color_ink, fill_type="solid")
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    instructions_data = [
        (
            "Column Structure",
            "Keep the 6 standard columns in the correct layout: A (Code), B (Description), C (Unit), D (Rate), E (Category), F (Comments).",
            "Do NOT delete or rename the standard columns. Empty columns are okay, but their headers and positions must be preserved."
        ),
        (
            "Wording & Headers",
            "Ensure first row contains column header labels. The import engine matches column types automatically using keywords (e.g. 'code', 'rate').",
            "Do NOT add duplicate header labels or nested multi-row header configurations."
        ),
        (
            "Row Spacing & Merges",
            "Place each Schedule of Rates item on its own flat row.",
            "Do NOT merge cells across columns or rows in the data section. Merges break CSV alignment."
        ),
        (
            "Pricing Rates",
            "Type rate values as clean decimal numbers. Comma separators and currency symbols ($) are cleaned automatically on import.",
            "Do NOT write formulas in the Rate column (e.g. =C4*D4). The Rate column must be a static number."
        ),
        (
            "Category Grouping",
            "Classify every item by specifying its Section/Category in Column E. This groups matching items beautifully in the PDF BOQ generator.",
            "Do NOT leave the Category column empty. Uncategorized items default to 'General SOR Pricing Items'."
        ),
        (
            "Active Items",
            "Ensure every pricing row has an Item Description in Column B. Empty descriptions are ignored.",
            "Do NOT delete the column header row or leave descriptions empty for items you want to keep."
        )
    ]
    
    row_font = Font(name=font_family, size=10, color="24292F")
    do_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    dont_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    curr_row = 4
    for cat, do_txt, dont_txt in instructions_data:
        sheet.row_dimensions[curr_row].height = 50
        
        c_cat = sheet.cell(row=curr_row, column=1, value=cat)
        c_cat.font = Font(name=font_family, size=10, bold=True, color=color_ink)
        c_cat.border = border_thin
        c_cat.alignment = Alignment(horizontal="left", vertical="center")
        
        c_do = sheet.cell(row=curr_row, column=2, value=do_txt)
        c_do.font = row_font
        c_do.fill = do_fill
        c_do.border = border_thin
        c_do.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        c_dont = sheet.cell(row=curr_row, column=3, value=dont_txt)
        c_dont.font = row_font
        c_dont.fill = dont_fill
        c_dont.border = border_thin
        c_dont.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        curr_row += 1
        
    column_widths = {"A": 20, "B": 50, "C": 50}
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

@locked_excel_write
def sync_db_to_active_excel(price_list_id: int) -> bool:
    """Regenerates the physical Excel pricebook file on disk from the SQLite database."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from services.db import get_db_connection
        
        file_path = get_price_list_path(price_list_id)
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Schedule of Rates"
        sheet.views.sheetView[0].showGridLines = True
        
        # Create second sheet for Instructions
        ins_sheet = wb.create_sheet(title="DOs and DONTs")
        write_instructions_sheet(ins_sheet)
        
        headers = ["SOR Code", "Item Description", "Action", "Unit", "Rate", "Quantity", "Total Cost", "Category", "Comments"]
        sheet.append(headers)
        
        header_fill = PatternFill(start_color="1F376A", end_color="1F376A", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D0D7DE"),
            right=Side(style="thin", color="D0D7DE"),
            top=Side(style="thin", color="D0D7DE"),
            bottom=Side(style="thin", color="D0D7DE")
        )
        
        for col_idx in range(1, 10):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT code, name, action, unit, rate, quantity, category, comments FROM price_items WHERE price_list_id = ? ORDER BY id",
            (price_list_id,)
        )
        rows = cursor.fetchall()
        conn.close()
        
        data_font = Font(name="Segoe UI", size=10, color="2B3036")
        for r_idx, r in enumerate(rows, 2):
            sheet.append([
                r["code"] or "",
                r["name"] or "",
                r["action"] or "",
                r["unit"] or "",
                r["rate"] or 0.0,
                r["quantity"] or 0.0,
                f"=E{r_idx}*F{r_idx}",
                r["category"] or "General SOR Pricing Items",
                r["comments"] or ""
            ])
            for col_idx in range(1, 10):
                cell = sheet.cell(row=r_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if col_idx == 5: # Rate
                    cell.number_format = "$#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx == 6: # Quantity
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 7: # Total Cost
                    cell.number_format = "$#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    
        column_widths = {"A": 15, "B": 45, "C": 12, "D": 12, "E": 15, "F": 12, "G": 15, "H": 25, "I": 30}
        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width
            
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb.save(file_path)
        wb.close()
        return True
    except Exception as e:
        print(f"[Matcher] Error syncing DB to Excel for list {price_list_id}: {e}")
        return False

@locked_excel_write
def write_cell_value_to_excel(file_path: str, row_idx: int, col_idx: int, value: str) -> bool:
    """Updates a cell value (quantity or comment) directly inside the SQLite database."""
    try:
        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        if col_idx == 6: # Comments (index 6 in cells array)
            cursor.execute("UPDATE price_items SET comments = ? WHERE id = ?", (value, row_idx))
        else: # Default/Quantity (col_idx == 5 or others)
            qty = 0.0
            if value.strip() != "":
                try:
                    qty = float(value)
                except ValueError:
                    pass
            cursor.execute("UPDATE price_items SET quantity = ? WHERE id = ?", (qty, row_idx))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"[Matcher] Error updating cell (col {col_idx}) in SQLite for item {row_idx}: {e}")
        return False

@locked_excel_write
def clear_column_values_in_excel(file_path: str, col_name_keyword: str = "QTY") -> bool:
    """Resets all quantity values to 0 inside the SQLite database for a specific price list."""
    try:
        price_list_id = get_id_from_path(file_path) or 1
        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE price_items SET quantity = 0, action = '', comments = '' WHERE price_list_id = ?", (price_list_id,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"[Matcher] Error resetting quantities in SQLite for list {price_list_id}: {e}")
        return False

@locked_excel_write
def clear_price_items_in_excel_batch(file_path: str, row_indices: list[int]) -> bool:
    """Deletes multiple items from the SQLite database by ID in a batch operation."""
    try:
        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.executemany("DELETE FROM price_items WHERE id = ?", [(rid,) for rid in row_indices])
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"[Matcher] Error batch deleting items from SQLite: {e}")
        return False

@locked_excel_write
def clear_all_price_items_in_excel(file_path: str) -> bool:
    """Wipes all rows from the SQLite database for a specific price list."""
    try:
        price_list_id = get_id_from_path(file_path) or 1
        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM price_items WHERE price_list_id = ?", (price_list_id,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"[Matcher] Error clearing SQLite price list {price_list_id}: {e}")
        return False

def get_sheet_column_widths(file_path: str) -> dict:
    """Returns a dictionary mapping column letters to their widths in Excel."""
    try:
        import openpyxl
        if not os.path.exists(file_path):
            return {}
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if "SOR" in wb.sheetnames:
            sheet = wb["SOR"]
        elif "Schedule of Rates" in wb.sheetnames:
            sheet = wb["Schedule of Rates"]
        elif "Schedule Of Rates" in wb.sheetnames:
            sheet = wb["Schedule Of Rates"]
        else:
            sheet = wb.active
            
        widths = {}
        for col in range(1, sheet.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            width = sheet.column_dimensions[col_letter].width
            widths[col_letter] = width if width is not None else 12
        wb.close()
        return widths
    except Exception as e:
        print(f"[Matcher] Error reading column widths: {e}")
        return {}

@locked_excel_write
def get_excel_layout_metadata(file_path: str) -> dict:
    """Returns a dictionary containing column widths and row heights for all sheets."""
    try:
        import openpyxl
        if not os.path.exists(file_path):
            return {"col_widths": {}, "row_heights": {}}
        wb = openpyxl.load_workbook(file_path, data_only=True)
        col_widths = {}
        row_heights = {}
        for name in wb.sheetnames:
            sheet = wb[name]
            # Column widths
            sheet_widths = {}
            for col in range(1, sheet.max_column + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                width = sheet.column_dimensions[col_letter].width
                if width is not None:
                    sheet_widths[col_letter] = width
            if sheet_widths:
                col_widths[name] = sheet_widths

            # Row heights
            sheet_heights = {}
            for r in range(1, sheet.max_row + 1):
                height = sheet.row_dimensions[r].height
                if height is not None:
                    sheet_heights[str(r)] = height
            if sheet_heights:
                row_heights[name] = sheet_heights
        wb.close()
        return {"col_widths": col_widths, "row_heights": row_heights}
    except Exception as e:
        print(f"[Matcher] Error reading Excel layout metadata: {e}")
        return {"col_widths": {}, "row_heights": {}}

def apply_excel_sorting_and_summary(wb) -> None:
    """
    Finds the active pricing sheet, enables AutoFilters on it, and creates/recreates
    a sorted 'BOQ Summary' sheet at index 0 showing matched quantities.
    """
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        import openpyxl.utils

        # 1. Identify active pricing sheet
        if "SOR" in wb.sheetnames:
            sheet = wb["SOR"]
        elif "Schedule of Rates" in wb.sheetnames:
            sheet = wb["Schedule of Rates"]
        elif "Schedule Of Rates" in wb.sheetnames:
            sheet = wb["Schedule Of Rates"]
        else:
            sheet = wb.active
            
        header_row_idx = 1
        qty_col = 5
        rate_col = 4
        code_col = 1
        name_col = 2
        unit_col = 3
        
        for row in range(1, 15):
            row_values = [str(sheet.cell(row, col).value or "").strip().upper() for col in range(1, 10)]
            if any("ITEM NAME" in val or "RATE" in val or "EXCLUDING GST" in val for val in row_values):
                header_row_idx = row
                for col_idx, val in enumerate(row_values, 1):
                    if ("QTY" in val or "QUANTITY" in val) and "UNIT" not in val:
                        qty_col = col_idx
                    elif "RATE" in val or "PRICE" in val:
                        rate_col = col_idx
                    elif any(x in val for x in ["CODE", "WCODE", "SOR"]):
                        code_col = col_idx
                    elif "ITEM NAME" in val or "DESCRIPTION" in val:
                        name_col = col_idx
                    elif "UNIT" in val:
                        unit_col = col_idx
                break
                
        # 2. Enable AutoFilter on the active pricing sheet
        try:
            max_col_letter = openpyxl.utils.get_column_letter(sheet.max_column)
            sheet.auto_filter.ref = f"A{header_row_idx}:{max_col_letter}{sheet.max_row}"
        except Exception as filter_err:
            print(f"[Matcher] Error applying main sheet auto filter: {filter_err}")
            
        # 3. Read matched items (Qty > 0)
        matched_items = []
        for r in range(header_row_idx + 1, sheet.max_row + 1):
            qty_val = sheet.cell(row=r, column=qty_col).value
            if qty_val is not None:
                try:
                    qty = float(qty_val)
                    if qty > 0:
                        code = str(sheet.cell(row=r, column=code_col).value or "").strip()
                        name = str(sheet.cell(row=r, column=name_col).value or "").strip()
                        unit = str(sheet.cell(row=r, column=unit_col).value or "").strip()
                        rate_val = sheet.cell(row=r, column=rate_col).value
                        rate = 0.0
                        if rate_val is not None:
                            try:
                                rate = float(str(rate_val).replace('$', '').replace(',', '').strip())
                            except ValueError:
                                pass
                        matched_items.append({
                            "code": code,
                            "name": name,
                            "unit": unit,
                            "rate": rate,
                            "qty": qty
                        })
                except ValueError:
                    pass
                    
        # 4. Sort matched items by Quantity descending
        matched_items.sort(key=lambda x: x["qty"], reverse=True)
        
        # 5. Delete existing 'BOQ Summary' sheet if any
        if "BOQ Summary" in wb.sheetnames:
            del wb["BOQ Summary"]
            
        # 6. Create 'BOQ Summary' sheet at index 0
        summary = wb.create_sheet(title="BOQ Summary", index=0)
        summary.views.sheetView[0].showGridLines = True
        
        # Headers styling
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark steel blue
        
        headers = ["SOR Code", "Description", "Unit", "Rate", "Quantity", "Total Cost"]
        for col_idx, h in enumerate(headers, 1):
            cell = summary.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")
            
        summary.row_dimensions[1].height = 25
        
        # 7. Write data rows
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        regular_font = Font(name="Segoe UI", size=10, bold=False)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        if matched_items:
            for i, item in enumerate(matched_items, 2):
                c_code = summary.cell(row=i, column=1, value=item["code"])
                c_name = summary.cell(row=i, column=2, value=item["name"])
                c_unit = summary.cell(row=i, column=3, value=item["unit"])
                c_rate = summary.cell(row=i, column=4, value=item["rate"])
                c_qty = summary.cell(row=i, column=5, value=item["qty"])
                c_tot = summary.cell(row=i, column=6, value=f"=D{i}*E{i}")
                
                # Fonts
                for cell in [c_code, c_name, c_unit, c_rate, c_qty, c_tot]:
                    cell.font = regular_font
                    
                # Formatting & Alignments
                c_code.alignment = center_align
                c_name.alignment = left_align
                c_unit.alignment = center_align
                
                c_rate.alignment = right_align
                c_rate.number_format = '"$"#,##0.00'
                
                c_qty.alignment = center_align
                c_qty.number_format = '#,##0'
                
                c_tot.alignment = right_align
                c_tot.number_format = '"$"#,##0.00'
                
                summary.row_dimensions[i].height = 20
                
            # Grand Total Row
            n = len(matched_items) + 2
            summary.cell(row=n, column=2, value="Grand Total").font = bold_font
            summary.cell(row=n, column=2).alignment = right_align
            
            c_tot_sum = summary.cell(row=n, column=6, value=f"=SUM(F2:F{n-1})")
            c_tot_sum.font = bold_font
            c_tot_sum.alignment = right_align
            c_tot_sum.number_format = '"$"#,##0.00'
            summary.row_dimensions[n].height = 22
            
            # Enable AutoFilter on Summary Sheet
            summary.auto_filter.ref = f"A1:F{n-1}"
        else:
            # Placeholder message
            cell = summary.cell(row=3, column=2, value="No priced quantities matched or input yet.")
            cell.font = Font(name="Segoe UI", size=11, italic=True)
            cell.alignment = left_align
            
        # 8. Set Column Widths
        col_widths = {"A": 15, "B": 50, "C": 10, "D": 15, "E": 12, "F": 18}
        for col_letter, width in col_widths.items():
            summary.column_dimensions[col_letter].width = width
            
    except Exception as e:
        print(f"[Matcher] Error applying sorting and summary: {e}")


import hashlib

def compute_row_hash(code: str, name: str, unit: str, rate: float) -> str:
    raw_str = f"{code or ''}|{name or ''}|{unit or ''}|{rate or 0.0}"
    return hashlib.md5(raw_str.encode('utf-8')).hexdigest()

def deterministic_parse_row(description: str, unit: str) -> Dict[str, Any]:
    desc_raw_upper = description.upper()
    unit_upper = unit.upper()
    
    # Strip mounting adjectives to prevent false-positive structure classification
    desc_upper = desc_raw_upper
    desc_upper = desc_upper.replace("TOWER MOUNTED", " ")
    desc_upper = desc_upper.replace("TOWER-MOUNTED", " ")
    desc_upper = desc_upper.replace("POLE MOUNTED", " ")
    desc_upper = desc_upper.replace("POLE-MOUNTED", " ")
    desc_upper = desc_upper.replace("WALL MOUNTED", " ")
    desc_upper = desc_upper.replace("WALL-MOUNTED", " ")
    desc_upper = desc_upper.replace("RACK MOUNTED", " ")
    desc_upper = desc_upper.replace("RACK-MOUNTED", " ")
    desc_upper = desc_upper.replace("COLLAR MOUNTED", " ")
    desc_upper = desc_upper.replace("COLLAR-MOUNTED", " ")
    
    # 1. semantic_class
    sem_class = "EQUIPMENT"
    if any(x in desc_upper for x in ["RIGGER", "TECH", "LABOUR", "SUPERVISOR", "ELECTRICIAN", "TRADES", "PERSONNEL", "HOURLY"]):
        sem_class = "LABOUR"
    elif any(x in desc_upper for x in ["CRANE", "EWP", "PLANT", "HIRE", "HOIST", "TRUCK", "BOOM", "PLATFORM HIRE"]):
        sem_class = "PLANT_HIRE"
    elif any(x in desc_upper for x in ["PIM", "SWEEP", "TEST", "COMMISSION", "AUDIT", "REPORT", "INSPECT"]):
        sem_class = "TESTING"
    elif any(x in desc_upper for x in ["CONCRETE", "FOUNDATION", "SLAB", "FOOTING", "GROUT", "EXCAVATION"]):
        sem_class = "FOUNDATION"
    elif any(x in desc_upper for x in ["FEEDER", "CABLE", "COAX", "HYBRID", "RUN", "TAIL", "CONDUIT", "TRAY", "FIBER", "OPTICAL"]):
        sem_class = "FEEDER"
    elif any(x in desc_upper for x in ["MONOPOLE", "POLE", "TOWER", "MAST", "STRUCTURE", "HEADFRAME", "LADDER", "MOUNT", "BRACKET"]):
        sem_class = "STRUCTURE"
    elif any(x in desc_upper for x in ["ANTENNA", "AAU", "RRU", "TMA", "GPS", "FILTER", "RRH", "COMBINER", "RADIO", "BASEBAND", "UNIT"]):
        sem_class = "EQUIPMENT"
    else:
        sem_class = "ACTIVITY"
        
    # 2. commercial_action
    action = "INSTALL"
    if any(x in desc_upper for x in ["REMOVE", "RECOVER", "DECOMMISSION", "DECOM", "DISMANTLE", "RECOVERY", "REMOVAL"]):
        action = "REMOVE"
    elif any(x in desc_upper for x in ["RELOCATE", "SHIFT", "REPOSITION", "MOVE"]):
        action = "RELOCATE"
    elif any(x in desc_upper for x in ["TEST", "AUDIT", "MEASURE", "SWEEP", "INSPECT"]):
        action = "TEST"
        
    # 3. quantity_basis / unit normalisation
    qty_basis = "EACH"
    if any(x in unit_upper for x in ["M3", "M³", "CUBIC"]):
        qty_basis = "M3"
    elif any(x in unit_upper for x in ["LM", "LINEAR", "METRE", "METER", " LNM"]) or (unit_upper == "M" and sem_class == "FEEDER"):
        qty_basis = "LM"
    elif any(x in unit_upper for x in ["HOUR", "HR", "PH"]):
        qty_basis = "HOUR"
    elif any(x in unit_upper for x in ["WEEK", "WK"]):
        qty_basis = "WEEK"
    elif any(x in unit_upper for x in ["DAY"]):
        qty_basis = "DAY"
    elif any(x in unit_upper for x in ["KM"]):
        qty_basis = "KM"
    elif any(x in unit_upper for x in ["NIGHT"]):
        qty_basis = "NIGHT"
    elif "SITE" in unit_upper:
        qty_basis = "PER_SITE"
    elif "SECTOR" in unit_upper:
        qty_basis = "PER_SECTOR"
    elif "6LM" in unit_upper or "6 LM" in unit_upper:
        qty_basis = "PER_6LM"
    elif "/M3" in unit_upper:
        qty_basis = "PER_M3"
        
    # 4. commercial_basis
    comm_basis = "BASE"
    if any(x in desc_upper for x in ["FIRST", "1ST", "SINGLE", "1 OFF"]):
        comm_basis = "FIRST"
    elif any(x in desc_upper for x in ["ADDITIONAL", "EXTRA OVER", "EXTRA-OVER", "EXTRA", "ADD"]):
        comm_basis = "ADDITIONAL"
        
    # 5. attributes
    attributes = {}
    critical_attributes = []
    
    # Extract monopole/pole/tower height
    height_match = re.search(r"(\d+(?:\.\d+)?)\s*M\b", desc_upper)
    if height_match and sem_class in ["STRUCTURE", "FEEDER"]:
        height = float(height_match.group(1))
        attributes["structure_height_m"] = height
        critical_attributes.append("structure_height_m")
        
    # Extract cable sizes
    size_match = re.search(r"(7/8|1/2|1-1/4|1/4|3/8|1\s+5/8)", desc_upper)
    if size_match:
        attributes["cable_size"] = size_match.group(1).replace(" ", "")
        critical_attributes.append("cable_size")
        
    # Extract cable count / multiplier
    count_match = re.search(r"X\s*(\d+)\b", desc_upper) or re.search(r"\b(\d+)\s*X\b", desc_upper) or re.search(r"\b(\d+)\s*(?:COAX|FEEDER|CABLE|RUN|OFF)", desc_upper)
    if count_match:
        attributes["cable_count"] = int(count_match.group(1))
        critical_attributes.append("cable_count")
        
    # Extract concrete volume
    vol_match = re.search(r"(\d+(?:\.\d+)?)\s*M3", desc_upper) or re.search(r"(\d+(?:\.\d+)?)\s*M³", desc_upper)
    if vol_match:
        attributes["volume_m3"] = float(vol_match.group(1))
        critical_attributes.append("volume_m3")
        
    return {
        "semantic_class": sem_class,
        "commercial_action": action,
        "compatible_actions": [action],
        "quantity_basis": qty_basis,
        "commercial_basis": comm_basis,
        "attributes": attributes,
        "critical_attributes": critical_attributes,
        "optional_attributes": []
    }

def normalize_price_list_to_knowledge_base(price_list_id: int) -> int:
    """Automatically profiles Excel pricing entries on import or app start, utilizing cache hashing and AI validation."""
    from services.db import get_db_connection
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, code, name, unit, rate, row_hash, profile_json FROM price_items WHERE price_list_id = ?",
        (price_list_id,)
    )
    rows = cursor.fetchall()
    conn.close()
    
    to_profile = []
    
    for r in rows:
        computed_hash = compute_row_hash(r["code"], r["name"], r["unit"], r["rate"])
        if r["row_hash"] == computed_hash and r["profile_json"]:
            continue # Skip, cached profile is valid
        to_profile.append({
            "id": r["id"],
            "code": r["code"] or "",
            "name": r["name"] or "",
            "unit": r["unit"] or "",
            "rate": r["rate"] or 0.0,
            "hash": computed_hash
        })
        
    if not to_profile:
        return 0
        
    # Call Gemini in batches of 50 items
    api_key = os.environ.get("GEMINI_API_KEY")
    ai_profiles = {}
    
    if api_key:
        from services.ai_service import send_gemini_request
        batch_size = 50
        for i in range(0, len(to_profile), batch_size):
            batch = to_profile[i:i+batch_size]
            items_json = json.dumps([
                {"id": item["id"], "description": item["name"], "unit": item["unit"]}
                for item in batch
            ], indent=2)
            
            prompt = f"""You are an expert telecom billing and Schedule of Rates (SOR) engineering systems profiler.
Analyze the following list of SOR items and produce a JSON array of profiled objects.
Each input item has an "id", "description", and "unit".
For each item, output a JSON object containing:
- "id": same as input id
- "semantic_class": one of "STRUCTURE", "FEEDER", "FOUNDATION", "LABOUR", "PLANT_HIRE", "TESTING", "EQUIPMENT", "ACTIVITY", "SITE_SERVICE"
- "commercial_action": one of "INSTALL", "REMOVE", "RELOCATE", "TEST"
- "compatible_actions": list containing commercial_action and alternative actions (e.g. ["INSTALL"])
- "quantity_basis": one of "EACH", "LM", "M3", "M2", "HOUR", "WEEK", "DAY", "KM", "NIGHT", "PER_SITE", "PER_SECTOR", "PER_6LM", "PER_M3"
- "commercial_basis": one of "BASE", "EXTRA_OVER", "FIRST", "ADDITIONAL"
- "attributes": dictionary of key-value attributes (e.g. structure_height_m, volume_m3, cable_size, cable_count, material, structure_type, etc.)
- "critical_attributes": list of keys from attributes that are critical to distinguish this code from other similar codes.
- "optional_attributes": list of optional attribute keys.

Input SOR items:
{items_json}

Return ONLY a valid JSON array matching this exact format:
[
  {{
    "id": 1,
    "semantic_class": "STRUCTURE",
    "commercial_action": "INSTALL",
    "compatible_actions": ["INSTALL"],
    "quantity_basis": "EACH",
    "commercial_basis": "BASE",
    "attributes": {{
       "structure_height_m": 30.0,
       "structure_type": "MONOPOLE",
       "material": "STEEL"
    }},
    "critical_attributes": ["structure_height_m", "structure_type"],
    "optional_attributes": ["material"]
  }}
]
Do not wrap in markdown or add explanations. Return ONLY the raw JSON array.
"""
            model_name = "gemini-3.5-flash-lite"
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={api_key}"
            payload = {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"responseMimeType": "application/json", "temperature": 0.0}
            }
            try:
                res = send_gemini_request(url, payload, timeout=60)
                candidates = res.get("candidates", [])
                if candidates:
                    text = candidates[0].get("content", {}).get("parts", [{}])[0].get("text", "").strip()
                    parsed = json.loads(text)
                    if isinstance(parsed, list):
                        for p in parsed:
                            if isinstance(p, dict) and "id" in p:
                                ai_profiles[int(p["id"])] = p
            except Exception as e:
                print(f"[Matcher] Error during batch profiling: {e}")
                
    # Now merge deterministic & AI profiles, apply validation schema, and save
    conn = get_db_connection()
    cursor = conn.cursor()
    
    updated_count = 0
    for item in to_profile:
        det_prof = deterministic_parse_row(item["name"], item["unit"])
        ai_prof = ai_profiles.get(item["id"], {})
        
        # Merge & Validate
        merged = dict(det_prof)
        if ai_prof:
            # Only trust AI for semantic class if not conflicting with critical deterministic indicator
            if ai_prof.get("semantic_class") in ["STRUCTURE", "FEEDER", "FOUNDATION", "LABOUR", "PLANT_HIRE", "TESTING", "EQUIPMENT", "ACTIVITY", "SITE_SERVICE"]:
                # If deterministic parsed structural attributes, keep class as STRUCTURE
                if det_prof["semantic_class"] == "STRUCTURE" and ai_prof.get("semantic_class") != "STRUCTURE":
                    pass
                else:
                    merged["semantic_class"] = ai_prof["semantic_class"]
            
            # Cross-check attributes: deterministic attributes are highly reliable (extracted via clean regex)
            for k, v in det_prof["attributes"].items():
                merged["attributes"][k] = v
                if k not in merged["critical_attributes"]:
                    merged["critical_attributes"].append(k)
                    
            # Pull in non-conflicting attributes from AI
            for k, v in ai_prof.get("attributes", {}).items():
                if k not in merged["attributes"]:
                    merged["attributes"][k] = v
                    
            # Pull action & unit if valid
            if ai_prof.get("commercial_action"):
                merged["commercial_action"] = ai_prof["commercial_action"]
            if ai_prof.get("quantity_basis"):
                merged["quantity_basis"] = ai_prof["quantity_basis"]
            if ai_prof.get("commercial_basis"):
                merged["commercial_basis"] = ai_prof["commercial_basis"]
                
        # Build final Profile JSON
        profile_data = {
            "sor_code": item["code"],
            "description": item["name"],
            "semantic_class": merged["semantic_class"],
            "attributes": merged["attributes"],
            "critical_attributes": merged["critical_attributes"],
            "optional_attributes": merged.get("optional_attributes", []),
            "commercial_action": merged["commercial_action"],
            "compatible_actions": merged.get("compatible_actions", [merged["commercial_action"]]),
            "quantity_basis": merged["quantity_basis"],
            "commercial_basis": merged["commercial_basis"],
            "dependencies": [],
            "parent_sor": None,
            "exclusions": [],
            "source_metadata": { "rate": item["rate"] },
            "profile_version": 1
        }
        
        profile_str = json.dumps(profile_data)
        cursor.execute(
            "UPDATE price_items SET profile_json = ?, row_hash = ? WHERE id = ?",
            (profile_str, item["hash"], item["id"])
        )
        updated_count += 1
        
    conn.commit()
    conn.close()
    return updated_count

def load_boq_items(price_list_id: Optional[int] = None) -> list[dict[str, Any]]:
    """Loads BOQ pricing items from the boq_items table in SQLite database."""
    try:
        if price_list_id is None:
            from services.db import get_default_price_list_id
            price_list_id = get_default_price_list_id()

        from services.db import get_db_connection
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, code, name, unit, rate, quantity, action, comments, category, price_item_id, confidence_score, confidence_level, evidence_json FROM boq_items WHERE price_list_id = ? ORDER BY (CASE WHEN code = 'UNQUOTED' THEN 1 ELSE 0 END) ASC, id ASC",
            (price_list_id,)
        )
        rows = cursor.fetchall()
        conn.close()
        
        boq_items = []
        current_category = None
        
        for r in rows:
            cat = r["category"] or "General SOR Pricing Items"
            if cat != current_category:
                current_category = cat
                boq_items.append({
                    "row_idx": 0,
                    "row_type": "section_header",
                    "code": "",
                    "name": current_category,
                    "unit": "",
                    "rate": 0.0,
                    "cells": [current_category, "", "", "", "", "", "", ""]
                })
                
            qty_val = r["quantity"]
            qty_str = ""
            if qty_val:
                qty_str = str(int(qty_val)) if float(qty_val).is_integer() else str(qty_val)
                
            total_val = r["rate"] * qty_val if r["rate"] and qty_val else 0.0
            total_str = f"${total_val:,.2f}" if total_val > 0 else "$-"
            rate_str = f"${r['rate']:,.2f}" if r["rate"] else ""
            
            boq_items.append({
                "row_idx": r["id"],
                "row_type": "data_item",
                "code": r["code"] or "",
                "name": r["name"] or "",
                "unit": r["unit"] or "",
                "rate": r["rate"] or 0.0,
                "category": r["category"] or "General SOR Pricing Items",
                "action": r["action"] or "",
                "comments": r["comments"] or "",
                "confidence_score": r["confidence_score"] or 100.0,
                "confidence_level": r["confidence_level"] or "HIGH",
                "evidence_json": r["evidence_json"] or "",
                "price_item_id": r["price_item_id"],
                "quantity": qty_val or 0.0,
                "cells": [
                    r["code"] or "",
                    r["name"] or "",
                    r["unit"] or "",
                    rate_str,
                    qty_str,
                    total_str,
                    r["comments"] or ""
                ]
            })
            
        return boq_items
    except Exception as e:
        print(f"[Matcher] Error loading BOQ items for list {price_list_id}: {e}")
        return []



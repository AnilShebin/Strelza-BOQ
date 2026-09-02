import sqlite3
import os
import json
import openpyxl
from typing import Dict, Any, List, Optional, Tuple

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads", "price_list.db")

def get_db_connection():
    """Establishes and returns a connection to the SQLite database with dict-like row parsing."""
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def get_default_price_list_id() -> int:
    """Dynamically retrieves the current active price list ID, or the first available list ID fallback."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM price_lists WHERE is_active = 1")
        row = cursor.fetchone()
        if not row:
            cursor.execute("SELECT id FROM price_lists LIMIT 1")
            row = cursor.fetchone()
        conn.close()
        return row["id"] if row else 1
    except Exception:
        return 1

def init_db():
    """Initializes the database schema and triggers bootstrap import if empty."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS price_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT,
        name TEXT,
        unit TEXT,
        rate REAL,
        quantity REAL DEFAULT 0,
        category TEXT,
        action TEXT DEFAULT ''
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS price_lists (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        is_active INTEGER DEFAULT 0
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS ai_prompts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        title TEXT,
        prompt TEXT,
        enabled INTEGER DEFAULT 1,
        project_type TEXT DEFAULT 'Default',
        version INTEGER DEFAULT 1
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS mapping_rules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        rule_name TEXT NOT NULL,
        category TEXT NOT NULL DEFAULT 'Antennas',
        equipment_type TEXT NOT NULL DEFAULT 'PANEL ANTENNA',
        match_keywords TEXT NOT NULL,
        exclude_keywords TEXT DEFAULT '',
        condition_expr TEXT DEFAULT '',
        action_filter TEXT DEFAULT 'ALL',
        target_sor_code TEXT NOT NULL,
        target_sor_name TEXT DEFAULT '',
        qty_formula TEXT DEFAULT 'table_qty',
        comment_template TEXT DEFAULT '',
        priority INTEGER DEFAULT 100,
        enabled INTEGER DEFAULT 1,
        internal_id TEXT DEFAULT '',
        conditions_json TEXT DEFAULT '',
        actions_json TEXT DEFAULT '',
        logic_explanation TEXT DEFAULT ''
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS price_registry (
        internal_id TEXT PRIMARY KEY,
        standard_name TEXT NOT NULL,
        category TEXT NOT NULL,
        default_unit TEXT DEFAULT 'each',
        fingerprint TEXT DEFAULT '',
        description TEXT DEFAULT ''
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS carrier_item_links (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        price_list_id INTEGER NOT NULL,
        internal_id TEXT NOT NULL,
        carrier_code TEXT DEFAULT '',
        carrier_name TEXT NOT NULL,
        category TEXT DEFAULT '',
        rate REAL DEFAULT 0.0,
        unit TEXT DEFAULT 'each',
        is_active INTEGER DEFAULT 1
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS correction_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
        pdf_name TEXT,
        original_description TEXT,
        corrected_code TEXT,
        corrected_name TEXT,
        corrected_rate REAL,
        estimator_username TEXT
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS boq_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        price_list_id INTEGER,
        code TEXT,
        name TEXT,
        unit TEXT,
        rate REAL,
        quantity REAL DEFAULT 0,
        action TEXT DEFAULT '',
        comments TEXT DEFAULT '',
        category TEXT,
        price_item_id INTEGER,
        confidence_score REAL DEFAULT 100.0,
        confidence_level TEXT DEFAULT 'HIGH',
        evidence_json TEXT DEFAULT ''
    );
    """)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS equipment_catalog (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        canonical_id TEXT UNIQUE NOT NULL,
        manufacturer TEXT DEFAULT '',
        model_name TEXT NOT NULL,
        equipment_class TEXT NOT NULL,
        category TEXT DEFAULT '',
        aliases_json TEXT DEFAULT '[]',
        attributes_json TEXT DEFAULT '{}',
        default_action TEXT DEFAULT 'INSTALL',
        is_active INTEGER DEFAULT 1,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    """)
    conn.commit()

    # Migrations for mapping_rules Venmo columns
    for col_name in ["target_sor_name", "internal_id", "conditions_json", "actions_json", "logic_explanation", "regex_pattern"]:
        try:
            cursor.execute(f"ALTER TABLE mapping_rules ADD COLUMN {col_name} TEXT DEFAULT ''")
            conn.commit()
        except sqlite3.OperationalError:
            pass

    # Enterprise-grade rule fields
    new_cols = [
        ("status", "TEXT DEFAULT 'ACTIVE'"),
        ("source", "TEXT DEFAULT 'HUMAN'"),
        ("approved_by", "TEXT DEFAULT ''"),
        ("rejected_by", "TEXT DEFAULT ''"),
        ("version", "INTEGER DEFAULT 1"),
        ("parent_rule_id", "INTEGER DEFAULT NULL"),
        ("approved_at", "TEXT DEFAULT NULL"),
        ("rejected_at", "TEXT DEFAULT NULL"),
        ("simulation_stats", "TEXT DEFAULT ''"),
        ("primary_source", "TEXT DEFAULT ''"),
        ("preferred_source_type", "TEXT DEFAULT ''"),
        ("ignore_pages", "TEXT DEFAULT ''"),
        ("duplicate_prone_pages", "TEXT DEFAULT ''"),
        ("matching_conditions", "TEXT DEFAULT ''"),
        ("notes", "TEXT DEFAULT ''"),
        ("rule_text", "TEXT DEFAULT ''")
    ]
    for col_name, col_def in new_cols:
        try:
            cursor.execute(f"ALTER TABLE mapping_rules ADD COLUMN {col_name} {col_def}")
            conn.commit()
        except sqlite3.OperationalError:
            pass

    # Migrations: Add price_list_id and comments columns to price_items if they don't exist
    try:
        cursor.execute("ALTER TABLE price_items ADD COLUMN price_list_id INTEGER")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE price_items ADD COLUMN comments TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE price_items ADD COLUMN confidence_score REAL DEFAULT 100.0")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE price_items ADD COLUMN confidence_level TEXT DEFAULT 'HIGH'")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE price_items ADD COLUMN evidence_json TEXT DEFAULT ''")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE price_items ADD COLUMN profile_json TEXT DEFAULT ''")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE price_items ADD COLUMN row_hash TEXT DEFAULT ''")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE price_items ADD COLUMN attributes_json TEXT DEFAULT ''")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS parser_configs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        config_key TEXT UNIQUE NOT NULL,
        category TEXT NOT NULL,
        name TEXT NOT NULL,
        pattern_value TEXT NOT NULL,
        description TEXT DEFAULT '',
        is_active INTEGER DEFAULT 1
    );
    """)
    conn.commit()

    # Seed default parser configs if table is empty
    cursor.execute("SELECT COUNT(*) as count FROM parser_configs")
    if cursor.fetchone()["count"] == 0:
        seed_default_parser_configs(cursor)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS system_settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL,
        description TEXT
    );
    """)
    conn.commit()

    # Seed default confidence thresholds if not present
    default_settings = [
        ("confidence_auto_approve", "90.0", "Minimum confidence score for automatic approval"),
        ("confidence_review_required", "70.0", "Threshold below which estimator review is mandatory"),
        ("confidence_unquoted_flag", "50.0", "Threshold for flagging unquoted custom scopes")
    ]
    for s_key, s_val, s_desc in default_settings:
        cursor.execute("INSERT OR IGNORE INTO system_settings (key, value, description) VALUES (?, ?, ?)", (s_key, s_val, s_desc))
    conn.commit()

    # Initialize default price list if empty and migrate existing items
    try:
        cursor.execute("SELECT COUNT(*) as count FROM price_lists")
        if cursor.fetchone()["count"] == 0:
            cursor.execute("INSERT INTO price_lists (name, is_active) VALUES ('Default Price List', 1)")
            conn.commit()
            
            cursor.execute("UPDATE price_items SET price_list_id = 1 WHERE price_list_id IS NULL")
            conn.commit()
            
            # Migrate existing active_price_list.xlsx to price_list_1.xlsx
            import shutil
            old_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads", "active_price_list.xlsx")
            new_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads", "price_list_1.xlsx")
            if os.path.exists(old_file) and not os.path.exists(new_file):
                try:
                    os.makedirs(os.path.dirname(new_file), exist_ok=True)
                    shutil.copy2(old_file, new_file)
                    print(f"[DB Migration] Migrated physical Excel file to: {new_file}")
                except Exception as e:
                    print(f"[DB Migration] Error copying file: {e}")
    except Exception as e:
        print(f"[DB Migration] Error during initialization: {e}")

    # Migration: Alter table if version column does not exist
    try:
        cursor.execute("ALTER TABLE ai_prompts ADD COLUMN version INTEGER DEFAULT 1")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    # Bootstrap default AI prompts if table is empty
    cursor.execute("SELECT COUNT(*) as count FROM ai_prompts")
    if cursor.fetchone()["count"] == 0:
        default_prompts = [
            {
                "name": "table_extractor",
                "title": "Vision Table Extractor Prompt",
                "version": 2,
                "prompt": """You are a precise document table detection and extraction vision AI.
Your task is to analyze the provided page image and its native text layer to extract ALL tables present on the page.

CRITICAL INSTRUCTIONS:
1. Scan the page completely. Identify whether any tables exist. A page may contain zero, one, or multiple tables.
2. For each identified table, extract its exact structure and content exactly as it appears. Do not rewrite, normalize, or interpret the values.
3. Preserve all cells, empty cells, special characters, units, column headers, and structural relationships.
4. Support rotated tables (rotated 90 degrees, 270 degrees, or vertically oriented). Extract them in their natural reading flow (left-to-right, top-to-bottom relative to the table's orientation).
5. Preserve merged cell relationships. In the "cells" list, specify row span and column span for merged cells, and map them to their correct 0-based row and column indices. Do not shift or offset columns/rows.
6. Multi-line content in a cell should remain in a single cell, preserving newlines (\\n) within the text.
7. Return ONLY a valid JSON object matching the following structure:

{
  "tables": [
    {
      "title": "Table Title (or null if no title is visible)",
      "bbox": [ymin, xmin, ymax, xmax], // Approximate bounding box of the table on the page using normalized coordinates between 0 and 1000 where [0, 0, 1000, 1000] is the full page.
      "headers": ["Header Column 1", "Header Column 2", ...], // List of header labels. If no clear header row exists, return null.
      "rows": [
        ["Cell 1A", "Cell 1B", ...],
        ["Cell 2A", "Cell 2B", ...]
      ],
      "cells": [
        {
          "row_idx": 0,
          "col_idx": 0,
          "rowspan": 1,
          "colspan": 1,
          "text": "Header Column 1"
        },
        ...
      ],
      "confidence": 0.95 // Float estimate (0.0 to 1.0) of extraction accuracy
    }
  ]
}

If no table is found on the page, return:
{
  "tables": []
}"""
            },
            {
                "name": "action_extractor",
                "title": "Drawing Action Notes Prompt",
                "version": 1,
                "prompt": """You are an expert telecom engineering drawing vision AI.
Inspect the provided full-page telecom layout/elevation drawing sheet carefully and identify all action notes to be performed.

CRITICAL INSTRUCTIONS:
1. Focus ONLY on notes, texts, or paragraphs that are enclosed inside scalloped revision clouds (cloud boxes / revision markups) on the drawing.
2. DO NOT extract any text, levels, labels, or centerlines that are NOT inside a cloud box. Skip plain elevation centerline markings (e.g. notes starting with 'C/L' like 'C/L PROPOSED ...'), relative level marks (e.g. 'E.L.', 'R.L.', 'A.H.D'), reference benchmarks, roof parapets, or overall heights unless they are specifically inside a cloud box.
3. Identify and extract ALL action items and scopes of work inside the cloud boxes. This includes:
   - Equipment installation, removal, recovery, decommissioning, replacement, relocation, modification, raising, or lowering.
   - Required testing (e.g., PIM testing, sweep testing, slab testing, structural integrity tests, or concrete audits).
   - Structural upgrades, mount replacements, feeder/cable re-routing, re-wiring, electrical upgrades, or earthing works.
   - Any contractor obligations or instructions.
4. DO NOT extract notes for items that are purely "EXISTING TO REMAIN" or "REUSED" (where no modification, replacement, or testing is required). Skip them. However, if an existing item requires testing (like PIM testing), extract the test action!
5. Format each extracted action note block or paragraph into a structured table format with the following columns exactly:
   ["ITEM", "EQUIPMENT DETAILS", "EXISTING", "PROPOSED", "TOTAL", "REFERENCE DWG"]
6. Extract each layout note block or paragraph as a single, complete row. Do NOT split a single note or paragraph into multiple rows.
7. In "EQUIPMENT DETAILS", output the EXACT text of the layout note or paragraph as it appears in the PDF. DO NOT omit any words, summarize, paraphrase, or truncate (e.g., retain details like "USING STANDARD MOUNT", "ON EXISTING MOUNT", "BEHIND PANEL ANTENNA (A5)" word-for-word).
8. Follow these numeric mapping rules:
   - For INSTALL / NEW / PROPOSED / TESTING / AUDIT / SURVEY actions: Set EXISTING to "0", PROPOSED to the quantity (default "1" if not specified), and TOTAL to the quantity.
   - For REMOVE / RECOVER / DECOMMISSION / DISMANTLE actions: Set EXISTING to the quantity (default "1"), PROPOSED to negative quantity (e.g. "-1"), and TOTAL to "0".
   - For REPLACE actions (e.g., "Recover A and replace with B"): Do NOT split. Keep as a single row. Set EXISTING to the quantity of the item being replaced (default "1"), PROPOSED to the quantity of the proposed item (default "1"), and TOTAL to the proposed quantity.
   - For RELOCATE / MODIFY / MOVE / RAISE actions: Set EXISTING to quantity (default "1"), PROPOSED to quantity (default "1"), and keep the exact details in the EQUIPMENT DETAILS.
9. In "REFERENCE DWG", extract any referenced drawing sheets (e.g., "T8-7", "S1-1") mentioned in the note.
10. Return ONLY a valid JSON object matching this exact structure:

{
  "tables": [
    {
      "table_title": "LAYOUT ACTION NOTES",
      "table_type": "EQUIPMENT_NOTES",
      "headers": ["ITEM", "EQUIPMENT DETAILS", "EXISTING", "PROPOSED", "TOTAL", "REFERENCE DWG"],
      "rows": [
        ["1", "PROPOSED TELSTRA METAL CHAIN BARRIER WITH SUPPORT POSTS. REFER TO SHEET T8-7 FOR DETAILS", "0", "1", "1", "T8-7"],
        ["2", "EXISTING TELSTRA LTE700 GPS ANTENNA (1 OFF A15) TO BE RECOVERED AND REPLACED WITH PROPOSED TELSTRA KA-7005-1110 GPS ANTENNA (1 OFF A15) TO BE INSTALLED USING STANDARD MOUNT", "1", "1", "1", "N/A"],
        ["3", "BUILD CONTRACTOR TO DO PIM TEST ON EXISTING FEEDER CABLES", "0", "1", "1", "S1-1"]
      ]
    }
  ]
}"""
            },
            {
                "name": "table_validator",
                "title": "Table Rows Alignment Auditor Prompt",
                "version": 1,
                "prompt": """You are a precise table data alignment auditor. Your ONLY task is to correct columns if text from the description columns (ITEM, EQUIPMENT, or EQUIPMENT DETAILS) has accidentally leaked into the quantity columns (EXISTING, PROPOSED, TOTAL) or the REFERENCE DWG column.

CRITICAL INSTRUCTIONS:
1. DO NOT change, modify, clean, simplify, or rewrite any text. Every cell's string content must remain exactly as given, only moved to the correct column if it was in the wrong one.
2. The quantity columns (EXISTING, PROPOSED, TOTAL) can contain numbers, signs (+/-), and expressions like '3 (SPARE)', '(SPARE)', 'SPARE', etc. Do not split or merge quantity descriptors like '(SPARE)' out of these columns.
3. If a cell in the quantity columns contains clear description text (e.g. 'RT11', 'RACK', 'RT/1/1', 'TED', 'SHEET S1-1'), move it back into 'EQUIPMENT DETAILS' or the appropriate description column.
4. The output MUST be a JSON array of lists, matching the exact dimensions of the input table.

Response Format:
Return ONLY the JSON array. Do not wrap in markdown or add explanations."""
            },
            {
                "name": "boq_deduplicator",
                "title": "AI BOQ Consolidator & Deduplicator Prompt",
                "version": 1,
                "prompt": """You are an expert material takeoff auditor. Your task is to analyze, consolidate, and deduplicate a list of raw equipment items extracted from drawing schedules and note pages.

CRITICAL INSTRUCTIONS:
1. Deduplicate duplicates: If the exact same physical equipment installation is listed twice (e.g. repeated lines), consolidate them so we do not double-count.
2. Group similar items: Combine items with identical descriptions, actions, and models, and sum their quantities.
3. Clean descriptions: Simplify descriptions to make them match typical price schedule names where possible.
4. Exclude existing items: Filter out and exclude any items that are purely EXISTING (e.g. equipment listed as existing to remain, or already on-site where no installation, removal, replacement, relocation, or testing work is required). Only output items representing active scopes of work. Do NOT output items with action "EXISTING" or "REUSE".
5. Output a clean, consolidated JSON list of items matching the exact structure below:

[
  {
    "equipment_type": "GPS ANTENNA",
    "model": "KA-7005-1110",
    "action": "INSTALL",
    "quantity": 1,
    "clean_text": "Proposed GPS antenna to be installed using standard mount",
    "source_sheet": "S1-1"
  }
]

Raw items list to deduplicate:
{raw_items}

Response format: Return ONLY the JSON array. Do not wrap in markdown or add explanations."""
            },
            {
                "name": "recheck_generator",
                "title": "AI Price List Match Auditor Prompt",
                "version": 1,
                "prompt": """You are an expert BOQ validation auditor. Inspect the final priced and mapped Bill of Quantities items below.
Identify any potential areas of concern, inconsistencies, or details that the human estimator should recheck.

Specific checks to report:
1. Unmapped items: Any item with "sor_code" set to "UNMAPPED" or rate set to 0.0.
2. Mismatched quantities: Quantities that seem abnormally high or low (e.g., negative proposed quantity, or quantity > 12).
3. Rate validation: High-cost items (e.g. value > $10,000) or items with low similarity score (< 80) mapping.

Return a JSON array of checklist items where each item has this structure:
{
  "check_name": "AI Recommendation: Re-check MW dish rate",
  "status": "WARNING",
  "message": "The proposed 0.6m MW dish was matched to standard mount SOR code. Verify if correct rate is applied."
}

Priced BOQ Items:
{mapped_items}

Response format: Return ONLY the JSON array. Do not wrap in markdown or add explanations. If no warnings or recommendations are found, return an empty array []."""
            },
            {
                "name": "structured_extractor",
                "title": "Vision Structured Tables Extractor",
                "version": 1,
                "prompt": """You are a precise, exhaustive structured document information extraction vision AI.
Your task is to analyze the provided page image (as the primary source of truth) and its native text layer (as supplementary reference) to extract ONLY STRUCTURED elements (tables, grids, schedules, key-value data list, or structured properties).
Do NOT extract unstructured general notes, drawing labels, annotations, or callouts.

CRITICAL INSTRUCTIONS FOR HIGH-FIDELITY TABLE EXTRACTION:
1. Identify and scan all structured tables, schedules, grids, or key-value sections.
2. Extract ALL cells, keeping the exact grid and column layout. Do NOT combine adjacent columns (e.g., if columns are 'EQUIPMENT' and 'EQUIPMENT DETAILS', extract them as separate columns; do not merge them into one).
3. Do NOT skip any rows, even empty-looking spacer rows or rows with empty cells. Keep the original row index matching the sheet.
4. For each structured table element:
   - "type" must be "structured".
   - "content" must be a JSON object with headers and rows: { "headers": ["Col 1", "Col 2", ...], "rows": [["Val 1A", "Val 1B", ...], ...] }
   - Or if it is key-value properties: { "fields": { "Key 1": "Value 1", ... } }
5. Provide the normalized bounding box [ymin, xmin, ymax, xmax] between 0 and 1000 relative to the page/crop.
6. Map elements to the NATIVE PDF TEXT BLOCKS by matching their IDs (e.g. B0, B1) in the "block_ids" array.

Return ONLY a valid JSON object matching the following structure:
{
  "elements": [
    {
      "type": "structured",
      "title": "Equipment Notes Table",
      "bbox": [100, 100, 300, 900],
      "block_ids": ["B0", "B1"],
      "content": {
        "headers": ["ITEM", "EQUIPMENT", "EQUIPMENT DETAILS", "EXISTING", "PROPOSED", "TOTAL", "REFERENCE DWG"],
        "rows": [
          ["1", "ELTEK SP18 RRU HC (PSU)", "600(w) x 600(d) x 2200mm(h)", "1", "0", "1", "SHEET E4-1"]
        ]
      },
      "confidence": 0.99
    }
  ]
}

If no structured content is found, return:
{
  "elements": []
}"""
            },
            {
                "name": "unstructured_extractor",
                "title": "Vision Unstructured Notes Extractor",
                "version": 1,
                "prompt": """You are a precise, exhaustive unstructured document information extraction vision AI.
Your task is to analyze the provided page image (as the primary source of truth) and its native text layer (as supplementary reference) to extract ONLY UNSTRUCTURED elements (general notes, annotations, labels, callouts, paragraph text, and standalone drawing descriptions).
Do NOT extract structured table grids or schedules.

CRITICAL INSTRUCTIONS FOR UNSTRUCTURED SCANNING:
1. Scan the page completely for any drawing notes, layout labels, notes blocks, and annotations.
2. Extract the text exactly as it appears. Do not summarize, format, or convert it into a grid.
3. For each unstructured element:
   - "type" must be "unstructured".
   - "content" must be a plain string containing the exact text of the label or note.
4. Provide the normalized bounding box [ymin, xmin, ymax, xmax] between 0 and 1000 relative to the page/crop.
5. Map elements to the NATIVE PDF TEXT BLOCKS by matching their IDs (e.g. B0, B1) in the "block_ids" array.

Return ONLY a valid JSON object matching the following structure:
{
  "elements": [
    {
      "type": "unstructured",
      "title": "Drawing General Note",
      "bbox": [500, 120, 520, 350],
      "block_ids": ["B2"],
      "content": "EXISTING TELSTRA RBS6102 ODU TO ACCOMMODATE PROPOSED EQUIPMENT.",
      "confidence": 0.95
    }
  ]
}

If no unstructured content is found, return:
{
  "elements": []
}"""
            }
        ]
        for p in default_prompts:
            cursor.execute(
                "INSERT INTO ai_prompts (name, title, prompt, version) VALUES (?, ?, ?, ?)",
                (p["name"], p["title"], p["prompt"], p.get("version", 1))
            )
        conn.commit()

    # Migration 1: Alter table if version column does not exist (already done above)
    # Migration 2: Upgrade existing table_extractor prompt if its version is < 2 (or Null)
    cursor.execute("SELECT version FROM ai_prompts WHERE name = 'table_extractor'")
    row = cursor.fetchone()
    if row:
        current_version = row["version"]
        if current_version is None or current_version < 2:
            # Get prompt string for version 2
            v2_prompt = """You are a precise document table detection and extraction vision AI.
Your task is to analyze the provided page image and its native text layer to extract ALL tables present on the page.

CRITICAL INSTRUCTIONS:
1. Scan the page completely. Identify whether any tables exist. A page may contain zero, one, or multiple tables.
2. For each identified table, extract its exact structure and content exactly as it appears. Do not rewrite, normalize, or interpret the values.
3. Preserve all cells, empty cells, special characters, units, column headers, and structural relationships.
4. Support rotated tables (rotated 90 degrees, 270 degrees, or vertically oriented). Extract them in their natural reading flow (left-to-right, top-to-bottom relative to the table's orientation).
5. Preserve merged cell relationships. In the "cells" list, specify row span and column span for merged cells, and map them to their correct 0-based row and column indices. Do not shift or offset columns/rows.
6. Multi-line content in a cell should remain in a single cell, preserving newlines (\\n) within the text.
7. Return ONLY a valid JSON object matching the following structure:

{
  "tables": [
    {
      "title": "Table Title (or null if no title is visible)",
      "bbox": [ymin, xmin, ymax, xmax], // Approximate bounding box of the table on the page using normalized coordinates between 0 and 1000 where [0, 0, 1000, 1000] is the full page.
      "headers": ["Header Column 1", "Header Column 2", ...], // List of header labels. If no clear header row exists, return null.
      "rows": [
        ["Cell 1A", "Cell 1B", ...],
        ["Cell 2A", "Cell 2B", ...]
      ],
      "cells": [
        {
          "row_idx": 0,
          "col_idx": 0,
          "rowspan": 1,
          "colspan": 1,
          "text": "Header Column 1"
        },
        ...
      ],
      "confidence": 0.95 // Float estimate (0.0 to 1.0) of extraction accuracy
    }
  ]
}

If no table is found on the page, return:
{
  "tables": []
}"""
            cursor.execute(
                "UPDATE ai_prompts SET prompt = ?, version = 2 WHERE name = 'table_extractor'",
                (v2_prompt,)
            )
            conn.commit()
            print("[DB Migration] Upgraded 'table_extractor' prompt to version 2.")

    # Migration 3: Alter table if action column does not exist
    try:
        cursor.execute("ALTER TABLE price_items ADD COLUMN action TEXT DEFAULT ''")
        conn.commit()
    except sqlite3.OperationalError:
        pass

    # Migration 4: Ensure universal_extractor prompt exists (Version 1)
    cursor.execute("SELECT COUNT(*) as count FROM ai_prompts WHERE name = 'universal_extractor'")
    if cursor.fetchone()["count"] == 0:
        universal_prompt = """You are a precise universal document information extraction vision AI.
Your task is to analyze the provided page image (as the primary source of truth) and its native text layer (as supplementary reference) to extract ALL meaningful information present on the page.

CRITICAL INSTRUCTIONS:
1. Scan the page completely. Identify and extract all information exactly as it appears.
2. Group the extracted information into exactly TWO content types:
   - "structured": tables, grids, schedules, key/value lists, and other data that has structured relationships.
   - "unstructured": notes, labels, callouts, paragraphs, annotations, drawing text, and other free-form text.
   Do NOT invent or use other types.
3. For "structured" elements:
   - "content" must be a JSON object preserving the natural structure of the data as found. Do not force every structured element into headers+rows.
   - If it naturally forms a table or grid, use:
     { "headers": ["Col 1", "Col 2", ...], "rows": [["Val 1A", "Val 1B", ...], ...] }
   - If it is key-value or schedule data, use:
     { "fields": { "Key 1": "Value 1", "Key 2": "Value 2", ... } }
4. For "unstructured" elements:
   - "content" must be a plain string containing the exact text of the note, callout, label, annotation, or block. Do not format it as a table.
5. Provide the approximate bounding box of the element on the page using normalized coordinates between 0 and 1000 where [ymin, xmin, ymax, xmax] are relative to [0, 0, 1000, 1000] for the full page.
6. Provide a float estimate (0.0 to 1.0) of extraction confidence for each element, or null if not reliable.
7. Return ONLY a valid JSON object matching the following structure:

{
  "elements": [
    {
      "type": "structured",
      "title": "Equipment Schedule", // optional title string, or null
      "bbox": [ymin, xmin, ymax, xmax], // optional list of 4 floats, or null
      "content": {
        "headers": ["Tag", "Model", "Qty"],
        "rows": [["A1", "ABC", "3"]]
      },
      "confidence": 0.98
    },
    {
      "type": "unstructured",
      "title": null,
      "bbox": [200, 150, 250, 850],
      "content": "EXISTING ANTENNA TO BE REMOVED.",
      "confidence": null
    }
  ]
}

If no extractable content is found, return:
{
  "elements": []
}"""
        cursor.execute(
            "INSERT INTO ai_prompts (name, title, prompt, version) VALUES (?, ?, ?, ?)",
            ("universal_extractor", "Universal Document Extractor Prompt", universal_prompt, 1)
        )
        conn.commit()
        print("[DB Migration] Created default 'universal_extractor' prompt (version 1).")

    # Migration 5: Upgrade universal_extractor prompt to Version 2 (exhaustive scan)
    cursor.execute("SELECT version FROM ai_prompts WHERE name = 'universal_extractor'")
    row = cursor.fetchone()
    if row:
        current_version = row["version"]
        if current_version is None or current_version < 2:
            v2_universal_prompt = """You are a precise, exhaustive universal document information extraction vision AI.
Your task is to analyze the provided page image (as the primary source of truth) and its native text layer (as supplementary reference) to extract ALL readable information present on the page.

CRITICAL INSTRUCTIONS FOR EXHAUSTIVE SCANNING:
1. Scan the page completely and systematically from top to bottom, left to right.
2. Extract ALL readable text pieces, annotations, labels, notes, callouts, diagram descriptions, title-block fields, elevation markers, dimensions, abbreviations, and reference tags (e.g. #2, #4, #13, RL, EL, etc.).
3. Do NOT summarize or generalize any information. Preserve the exact wording, numbers, units, symbols, special characters, and abbreviations as they appear.
4. Do NOT decide what information is important. Do NOT skip small text, repetitively appearing labels, or standalone notes outside tables.
5. Group all extracted elements into exactly TWO types:
   - "structured": tables, grids, schedules, key-value data list, or structured properties.
   - "unstructured": notes, labels, callouts, annotations, drawing text, paragraphs, and standalone text.
   Do NOT invent or use other types.
6. For "structured" elements:
   - "content" must be a JSON object preserving the natural structure. Do not force every structured element into headers+rows.
   - If it naturally forms a table or grid, use: { "headers": ["Col 1", "Col 2", ...], "rows": [["Val 1A", "Val 1B", ...], ...] }
   - If it is key-value or schedule details, use: { "fields": { "Key 1": "Value 1", "Key 2": "Value 2", ... } }
7. For "unstructured" elements:
   - "content" must be a plain string containing the exact text of the annotation/label as found.
8. Provide the bounding box of the element using normalized coordinates [ymin, xmin, ymax, xmax] between 0 and 1000 relative to the full page (or crop area if analyzing a tiled region).
9. Provide a float estimate (0.0 to 1.0) of confidence, or null if not reliable.
10. Return ONLY a valid JSON object matching the following structure:

{
  "elements": [
    {
      "type": "structured",
      "title": "Revision History",
      "bbox": [100, 100, 300, 900],
      "content": {
        "headers": ["Rev", "Date", "Description"],
        "rows": [["A", "12/08/2026", "ISSUED FOR CONSTRUCTION"]]
      },
      "confidence": 0.99
    },
    {
      "type": "unstructured",
      "title": "Equipment Annotation",
      "bbox": [500, 120, 520, 350],
      "content": "EXISTING TELSTRA RBS6102 ODU TO ACCOMMODATE PROPOSED EQUIPMENT.",
      "confidence": 0.95
    }
  ]
}

If no extractable content is found, return:
{
  "elements": []
}"""
            cursor.execute(
                "UPDATE ai_prompts SET prompt = ?, version = 2 WHERE name = 'universal_extractor'",
                (v2_universal_prompt,)
            )
            conn.commit()
            print("[DB Migration] Upgraded 'universal_extractor' prompt to version 2.")

    # Migration 6: Ensure additional_extraction_instructions prompt exists
    cursor.execute("SELECT COUNT(*) as count FROM ai_prompts WHERE name = 'additional_extraction_instructions'")
    if cursor.fetchone()["count"] == 0:
        additional_instructions = """[TELSTRA-SPECIFIC EXTRACTION RULES]:
1. Pay extreme attention to the "Notes", "General Notes", "Equipment Notes", and "Revision History" blocks.
2. Extract all equipment markers and identifiers (e.g. RBS6102, RBS6601, MW, #2, #4, #13).
3. Extract all cable information (e.g. feeder cables, hybrid cables, hybrid tails, lengths, diameters, sizes).
4. Capture existing vs proposed descriptions, elevation markings (e.g. RL, EL, AHD), and GPS antenna references.
5. Ignorable/safe-to-ignore information: You may ignore border drawing labels, general page index lists, or standard engineering diagram line scales unless they contain specific billable material notes."""
        cursor.execute(
            "INSERT INTO ai_prompts (name, title, prompt, version, enabled) VALUES (?, ?, ?, ?, ?)",
            ("additional_extraction_instructions", "Additional Client-Specific Extraction Instructions", additional_instructions, 1, 1)
        )
        conn.commit()
        print("[DB Migration] Created default 'additional_extraction_instructions' prompt (version 1).")

    # Migration 7: Upgrade additional_extraction_instructions prompt to Version 2
    cursor.execute("SELECT version FROM ai_prompts WHERE name = 'additional_extraction_instructions'")
    row = cursor.fetchone()
    if row:
        current_version = row["version"]
        if current_version is None or current_version < 2:
            v2_additional_instructions = """[TELSTRA-SPECIFIC EXTRACTION RULES]:
1. EXHAUSTIVE SCOPE: You MUST scan and extract EVERY single drawing element, label, and annotation. Do not summarize or skip anything.
2. PROPOSED & REMOVAL ITEMS: Capture all "PROPOSED..." (e.g., "PROPOSED TELSTRA HANDRAILS (TYP.)", proposed mounts, proposed feeders) and all "EXISTING... TO BE REMOVED/RELOCATED" items. These represent critical billable project scopes and must never be omitted.
3. PHYSICAL INSTALLATION & SAFETY: Capture all structural items, platform additions, mounts, brackets, handrails, ladders, fall arrest systems, security gates, cable trays, and safety cages.
4. NOTES & LEGENDS: Pay extreme attention to the "Notes", "General Notes", "Equipment Notes", and "Revision History" blocks.
5. CABLES & EQUIPMENT: Extract all cable information (e.g., feeder cables, hybrid cables, hybrid tails, lengths, diameters, sizes) and all equipment markers (e.g., RBS6102, RBS6601, MW dishes, #2, #4, #13).
6. ELEVATION & DETAILS: Capture elevation markings (e.g., RL, EL, AHD), GPS reference locations, and individual equipment details.
7. SAFE-TO-IGNORE: Only general page index lists or standard engineering diagram line scales (with no text labels) can be bypassed. All text labels, dimensions, and callouts must be preserved."""
            cursor.execute(
                "UPDATE ai_prompts SET prompt = ?, version = 2 WHERE name = 'additional_extraction_instructions'",
                (v2_additional_instructions,)
            )
            conn.commit()
            print("[DB Migration] Upgraded 'additional_extraction_instructions' prompt to version 2.")

    # Migration 8: Upgrade additional_extraction_instructions prompt to Version 3
    cursor.execute("SELECT version FROM ai_prompts WHERE name = 'additional_extraction_instructions'")
    row = cursor.fetchone()
    if row:
        current_version = row["version"]
        if current_version is None or current_version < 3:
            v3_additional_instructions = """[TELSTRA-SPECIFIC EXTRACTION RULES]:
1. REVISION CLOUD BOXES / CIRCLED ITEMS ARE HIGHEST PRIORITY: You MUST scan and extract EVERY single text annotation, label, or note that is circled or enclosed inside a scalloped "revision cloud" (cloud box / bubble markup) on the drawing. Do NOT miss any of them. Even if they describe "EXISTING" items (e.g. "EXISTING ROOFTOP ACCESS DOOR", "EXISTING CABLE LADDER", "EXISTING TELSTRA EQUIPMENT INSTALLED INSIDE CABINET"), they are high priority and must be extracted because they are highlighted for verification/work.
2. ANNOTATIONS INSIDE DRAWING VIEWS: Extract all text annotations, descriptions, and notes embedded directly inside the drawing/elevation views (e.g. "EXISTING SIGN SECURED #13 TO TELSTRA ODU DOOR BY USING TEK SCREWS", "EXISTING SIGN SECURED #4 ROOFTOP ACCESS DOOR BY USING SELF ADHESIVES", "EXISTING CABLE LADDER", etc.). Human errors often happen inside these images, so capturing these annotations is critical to prevent mistakes.
3. PROPOSED & REMOVAL ITEMS: Capture all "PROPOSED..." (e.g., "PROPOSED TELSTRA HANDRAILS (TYP.)", proposed mounts, proposed feeders) and all "EXISTING... TO BE REMOVED/RELOCATED/UPGRADED" items. These represent critical billable project scopes and must never be omitted.
4. PHYSICAL INSTALLATION & SAFETY: Capture all structural items, platform additions, mounts, brackets, handrails, ladders, fall arrest systems, security gates, cable trays, and safety cages.
5. NOTES & LEGENDS: Pay extreme attention to the "Notes", "General Notes", "Equipment Notes", and "Revision History" blocks.
6. CABLES & EQUIPMENT: Extract all cable information (e.g., feeder cables, hybrid cables, hybrid tails, lengths, diameters, sizes) and all equipment markers (e.g., RBS6102, RBS6601, MW dishes, #2, #4, #13).
7. ELEVATION & DETAILS: Capture elevation markings (e.g., RL, EL, AHD), GPS reference locations, and individual equipment details.
8. SAFE-TO-IGNORE: Only general page index lists or standard engineering diagram line scales (with no text labels) can be bypassed. All text labels, dimensions, and callouts must be preserved."""
            cursor.execute(
                "UPDATE ai_prompts SET prompt = ?, version = 3 WHERE name = 'additional_extraction_instructions'",
                (v3_additional_instructions,)
            )
            conn.commit()
            print("[DB Migration] Upgraded 'additional_extraction_instructions' prompt to version 3.")

    # Migration 9: Upgrade additional_extraction_instructions prompt to Version 4 (Zero-Miss Mandate)
    cursor.execute("SELECT version FROM ai_prompts WHERE name = 'additional_extraction_instructions'")
    row = cursor.fetchone()
    if row:
        current_version = row["version"]
        if current_version is None or current_version < 4:
            v4_additional_instructions = """[TELSTRA-SPECIFIC EXTRACTION RULES - ZERO-MISS MANDATE]:
CRITICAL DIRECTIVE: You MUST operate under a strict "Zero-Miss Policy". Extract EVERY single text element, annotation, callout, symbol label, and dimension visible on the sheet. Do NOT summarize, merge, generalize, or skip ANY text.

1. REVISION CLOUDS & BUBBLES ARE ABSOLUTE PRIORITY:
   - Identify every scalloped "revision cloud" box or circle markup on the drawing sheet.
   - You MUST scan and extract EVERY word, number, and note inside or directly pointing to these clouds (e.g. "EXISTING ROOFTOP ACCESS DOOR", "EXISTING CABLE LADDER", "EXISTING TELSTRA EQUIPMENT INSTALLED INSIDE CABINET IN THE BUILDING PLANTROOM AREA").
   - These are verified modifications/notes and must NEVER be missed.

2. EMBEDDED DIAGRAM VIEWS & PICTURE ANNOTATIONS:
   - Carefully scan inside all elevation diagrams, drawings, symbols, and equipment profiles.
   - Extract text annotations pointing to cables, mounts, platforms, and structures (e.g. "EXISTING SIGN SECURED #13 TO TELSTRA ODU DOOR BY USING TEK SCREWS", "EXISTING SIGN SECURED #4 ROOFTOP ACCESS DOOR BY USING SELF ADHESIVES").
   - Capturing notes inside these graphics is mandatory to prevent project estimation human errors.

3. ELEVATION, REFERENCE levels, & SYSTEM REF LABELS:
   - Extract every reference label and elevation height (e.g., "E.L. 0.00m (±100mm), R.L. 25.6m A.H.D GROUND LEVEL", "E.L. 23.0m (±100mm) R.L. 54.9m A.H.D").
   - Exposing heights and reference levels prevents critical engineering and safety issues.

4. CABLES, FEEDERS, & HARDWARE NOTES:
   - Capture all cable counts, diameters, technology details (e.g., "EXISTING COAX FEEDERS", "HYBRID CABLES (4 OFF)", "LTE700/NR850").
   - Extract safety cages, handrails, fall arrest lines, cable trays, support pipes, brackets, and grounding details.

5. GENERAL MANDATE:
   - If a piece of text is readable, it MUST be extracted.
   - Do NOT ignore "EXISTING" labels. Existing items that undergo verification are billing-relevant and safety-critical.
   - Only standard visual sheet margins, drawing sheet frames, or line scales without any text can be bypassed."""
            cursor.execute(
                "UPDATE ai_prompts SET prompt = ?, version = 4 WHERE name = 'additional_extraction_instructions'",
                (v4_additional_instructions,)
            )
            conn.commit()
            print("[DB Migration] Upgraded 'additional_extraction_instructions' prompt to version 4.")

    # Migration 10: Upgrade additional_extraction_instructions prompt to Version 5 (Complete Multi-Line Blocks)
    cursor.execute("SELECT version FROM ai_prompts WHERE name = 'additional_extraction_instructions'")
    row = cursor.fetchone()
    if row:
        current_version = row["version"]
        if current_version is None or current_version < 5:
            v5_additional_instructions = """[TELSTRA-SPECIFIC EXTRACTION RULES - ZERO-MISS MANDATE v5]:
CRITICAL DIRECTIVE: You MUST operate under a strict "Zero-Miss Policy". Extract EVERY single text element, annotation, callout, symbol label, and dimension visible on the sheet. Do NOT summarize, merge, generalize, or skip ANY text.

1. MULTI-LINE INSTALLATION DESCRIPTIONS MUST BE COMPLETE:
   - Cable ladder, cable routing, and equipment installation notes often span multiple lines of text.
   - You MUST extract the FULL verbatim text of such blocks — do NOT truncate, shorten, or paraphrase them.
   - Include the complete text exactly: cable model numbers (e.g. LCF78-50JA, LCF12-50J, RFS LCF12-50J), technology band labels (e.g. LTE1800/LTE2100/NR/LTE2600/NR3600), quantities (e.g. 4 OFF, 1 OFF, 2 OFF), and action keywords (e.g. TO BE REUSED, TO BE RECOVERED, TO BE REMOVED, TO BE RELOCATED).
   - Example of a block that MUST be captured as a single complete unstructured element:
     "EXISTING TELSTRA CABLE LADDER MOUNTED ON OUTSIDE OF PARAPET ABOVE CONCRETE WALL FEATURE WITH CUSTOM MADE LIDS TO ACCOMMADATE PROPOSED TELSTRA RFS LCF78-50JA FEEDER CABLES (4 OFF). EXISTING TELSTRA W&B HYBRID CABLE (1 OFF) TO BE REUSED FOR LTE1800/LTE2100/NR/LTE2600/NR3600. EXISTING RFS LCF12-50J FEEDER CABLES (2 OFF) TO BE RECOVERED"

2. REVISION CLOUDS & BUBBLES ARE ABSOLUTE PRIORITY:
   - Identify every scalloped "revision cloud" box or circle markup on the drawing sheet.
   - You MUST scan and extract EVERY word, number, and note inside or directly pointing to these clouds.
   - These are verified modifications/notes and must NEVER be missed.

3. EMBEDDED DIAGRAM VIEWS & PICTURE ANNOTATIONS:
   - Carefully scan inside all elevation diagrams, drawings, symbols, and equipment profiles.
   - Extract text annotations pointing to cables, mounts, platforms, and structures.
   - Capturing notes inside these graphics is mandatory to prevent project estimation human errors.

4. ELEVATION, REFERENCE LEVELS & SYSTEM REF LABELS:
   - Extract every reference label and elevation height (e.g., "E.L. 0.00m (±100mm), R.L. 25.6m A.H.D GROUND LEVEL").
   - Exposing heights and reference levels prevents critical engineering and safety issues.

5. CABLES, FEEDERS & HARDWARE NOTES:
   - Capture all cable counts, diameters, technology details (e.g., "EXISTING COAX FEEDERS", "HYBRID CABLES (4 OFF)", "LTE700/NR850").
   - Extract safety cages, handrails, fall arrest lines, cable trays, support pipes, brackets, and grounding details.

6. GENERAL MANDATE:
   - If a piece of text is readable, it MUST be extracted in full.
   - Do NOT ignore "EXISTING" labels. Existing items that undergo verification are billing-relevant and safety-critical.
   - Only standard visual sheet margins, drawing sheet frames, or line scales without any text can be bypassed."""
            cursor.execute(
                "UPDATE ai_prompts SET prompt = ?, version = 5 WHERE name = 'additional_extraction_instructions'",
                (v5_additional_instructions,)
            )
            conn.commit()
            print("[DB Migration] Upgraded 'additional_extraction_instructions' prompt to version 5.")

    # Migration 11: Upgrade additional_extraction_instructions prompt to Version 6
    # Phase-based extraction + bbox accuracy + individual elevation markers
    cursor.execute("SELECT version FROM ai_prompts WHERE name = 'additional_extraction_instructions'")
    row = cursor.fetchone()
    if row:
        current_version = row["version"]
        if current_version is None or current_version < 6:
            v6_additional_instructions = """[TELSTRA-SPECIFIC EXTRACTION RULES - ZERO-MISS MANDATE v6]:
CRITICAL DIRECTIVE: Extract EVERY piece of readable text. Do NOT summarize, merge into groups, or skip ANY text element.

══════════════════════════════════════════
PHASE 1 — REVISION CLOUDS (EXTRACT FIRST):
══════════════════════════════════════════
Before extracting anything else, scan the ENTIRE page for revision clouds (scalloped cloud outlines / bubble markups).
For EACH revision cloud found:
  - Extract the cloud as its own SEPARATE unstructured element.
  - Include the COMPLETE verbatim multi-line text inside the cloud — every word, every model number, every quantity, every action keyword.
  - Do NOT merge multiple clouds into one element.
  - Example of a complete cloud element that must NOT be truncated:
    "EXISTING TELSTRA CABLE LADDER MOUNTED ON OUTSIDE OF PARAPET ABOVE CONCRETE WALL FEATURE WITH CUSTOM MADE LIDS TO ACCOMMADATE PROPOSED TELSTRA RFS LCF78-50JA FEEDER CABLES (4 OFF). EXISTING TELSTRA W&B HYBRID CABLE (1 OFF) TO BE REUSED FOR LTE1800/LTE2100/NR/LTE2600/NR3600. EXISTING RFS LCF12-50J FEEDER CABLES (2 OFF) TO BE RECOVERED"

══════════════════════════════════════════
PHASE 2 — ELEVATION / REFERENCE MARKERS (EACH AS INDIVIDUAL ITEM):
══════════════════════════════════════════
Extract every E.L. / R.L. / C/L elevation marker as its OWN separate unstructured element.
  - Do NOT group all elevation markers into a single structured block.
  - Each line such as "E.L. 29.0m (±100mm) R.L. 54.6m A.H.D C/L PROPOSED TELSTRA NR3600 PANEL ANTENNAS (2 OFF A6 & A10)" is ONE separate element.
  - This is critical because each marker corresponds to a different elevation level and a different piece of equipment.

══════════════════════════════════════════
PHASE 3 — ALL OTHER LABELS, CALLOUTS & ANNOTATIONS:
══════════════════════════════════════════
After completing Phases 1 and 2, extract all remaining text:

1. MULTI-LINE INSTALLATION DESCRIPTIONS MUST BE COMPLETE:
   - Cable ladder, cable routing, and equipment installation notes often span multiple lines.
   - Extract the FULL verbatim text — do NOT truncate, shorten, or paraphrase.
   - Include: cable model numbers (LCF78-50JA, LCF12-50J), technology bands (LTE1800/LTE2100/NR/LTE2600/NR3600), quantities (4 OFF, 1 OFF, 2 OFF), action keywords (TO BE REUSED, TO BE RECOVERED, TO BE REMOVED, TO BE RELOCATED).

2. EACH LEADER LINE = SEPARATE ELEMENT:
   - If multiple callout boxes each have their own separate leader/pointer line, extract each callout as a SEPARATE unstructured element.
   - Do NOT combine callouts that point to different equipment or locations into one element.

3. EMBEDDED DIAGRAM ANNOTATIONS:
   - Extract all text annotations inside elevation diagrams, drawings, equipment profiles.
   - Capturing inside-drawing notes is mandatory to prevent project estimation errors.

4. CABLES, FEEDERS & HARDWARE NOTES:
   - Capture all cable counts, diameters, technology details, safety equipment.

5. BOUNDING BOX ACCURACY RULE:
   - The bbox must surround the TEXT LABEL itself, NOT the physical equipment or structure the text points to.
   - Example: For the label "EXISTING CABLE LADDER" with a leader line pointing to a cable ladder structure, the bbox must cover the text words "EXISTING CABLE LADDER", not the cable ladder drawing.

6. GENERAL MANDATE:
   - If a piece of text is readable, it MUST be extracted in full.
   - Do NOT ignore "EXISTING" labels — they are billing-relevant and safety-critical.
   - Only standard visual sheet margins, drawing sheet frames, or line scales without text can be bypassed."""
            cursor.execute(
                "UPDATE ai_prompts SET prompt = ?, version = 6 WHERE name = 'additional_extraction_instructions'",
                (v6_additional_instructions,)
            )
            conn.commit()
            print("[DB Migration] Upgraded 'additional_extraction_instructions' prompt to version 6.")

    # Migration 12: Ensure unified_extractor prompt exists (Version 1)
    cursor.execute("SELECT COUNT(*) as count FROM ai_prompts WHERE name = 'unified_extractor'")
    if cursor.fetchone()["count"] == 0:
        unified_prompt = """You are a precise, exhaustive document extraction vision AI.
Your task is to analyze the provided page image (as the primary source of truth) and its native text layer (as supplementary reference) to extract BOTH:
1. STRUCTURED elements (tables, grids, schedules, key-value data list, or structured properties).
2. UNSTRUCTURED elements (general notes, annotations, labels, callouts, paragraph text, or standalone drawing descriptions).

CRITICAL INSTRUCTIONS FOR HIGH-FIDELITY EXTRACTION:
1. Identify and scan all structured tables, schedules, grids, or key-value sections.
2. Extract ALL cells, keeping the exact grid and column layout. Do NOT combine adjacent columns (e.g., if columns are 'EQUIPMENT' and 'EQUIPMENT DETAILS', extract them as separate columns; do not merge them into one).
3. Do NOT skip any rows, even empty-looking spacer rows or rows with empty cells. Keep the original row index matching the sheet.
4. If there are multiple side-by-side or separate tables on the page (e.g., two antenna configuration tables side-by-side), extract them as SEPARATE structured elements with their own bounding boxes and titles. Do NOT join or merge separate/side-by-side tables into a single table.
5. Scan the page completely for any drawing notes, layout labels, notes blocks, and annotations.
6. For unstructured elements:
   - "type" must be "unstructured".
   - "content" must be a plain string containing the exact text of the label or note. Do not summarize or format it.
7. Provide the normalized bounding box [ymin, xmin, ymax, xmax] between 0 and 1000 relative to the page.
8. Map elements to the NATIVE PDF TEXT BLOCKS by matching their IDs (e.g. B0, B1) in the "block_ids" array.

Return ONLY a valid JSON object matching the following structure:
{
  "elements": [
    {
      "type": "structured",
      "title": "Equipment Notes Table",
      "bbox": [100, 100, 300, 900],
      "block_ids": ["B0", "B1"],
      "content": {
        "headers": ["ITEM", "EQUIPMENT", "EQUIPMENT DETAILS", "EXISTING", "PROPOSED", "TOTAL", "REFERENCE DWG"],
        "rows": [
          ["1", "ELTEK SP18 RRU HC (PSU)", "600(w) x 600(d) x 2200mm(h)", "1", "0", "1", "SHEET E4-1"]
        ]
      },
      "confidence": 0.99
    },
    {
      "type": "unstructured",
      "title": "Drawing General Note",
      "bbox": [500, 120, 520, 350],
      "block_ids": ["B2"],
      "content": "EXISTING TELSTRA RBS6102 ODU TO ACCOMMODATE PROPOSED EQUIPMENT.",
      "confidence": 0.95
    }
  ]
}

If no extractable content is found, return:
{
  "elements": []
}"""
        cursor.execute(
            "INSERT INTO ai_prompts (name, title, prompt, version, enabled) VALUES (?, ?, ?, ?, ?)",
            ("unified_extractor", "Vision Unified Structured/Unstructured Extractor Prompt", unified_prompt, 1, 1)
        )
        conn.commit()
        print("[DB Migration] Created default 'unified_extractor' prompt (version 1).")

    # Migration 13: Ensure boq_mapping_engine and client_mapping_rules prompts exist (Version 1)
    cursor.execute("SELECT COUNT(*) as count FROM ai_prompts WHERE name = 'boq_mapping_engine'")
    if cursor.fetchone()["count"] == 0:
        mapping_prompt = """You are an expert telecom BOQ auditor and price schedule mapping AI.
Your task is to analyze extracted drawing tables, layout notes/callouts, and map the required active work scopes to the active Price Book Schedule of Rates (SOR) items.

CRITICAL INSTRUCTIONS & HIERARCHY:

1. TABLE-FIRST PRIMARY AUTHORITY (80-90% of BOQ):
   - Structured Tables are the primary source of truth for equipment quantities and specifications.
   - Antenna Configuration Tables take precedence for all Antenna items (Panel Antennas, AAU, GPS antennas) and antenna removals.
   - Equipment Notes tables take precedence for internal equipment, racks, RRUs, TMAs, Surge Protection Devices (SPD), and shelter hardware.

2. CROSS-VERIFICATION WITH LAYOUT CLOUDS & REFERENCED DRAWINGS:
   - For every table item, cross-check against the unstructured layout callouts, scalloped revision clouds, elevation markers, and referenced drawing sheets (e.g. from the 'REFERENCE DWG' column).
   - If the table quantity and layout annotations MATCH: Output the table quantity with comment "".
   - If the table quantity and layout annotations DIFFER: Use the TABLE quantity as authoritative count, and set the comment to: "Data not matching with antenna layout".

3. ACTION CLASSIFICATION & FILTERING:
   - Active Actions: INSTALL, PROPOSED, NEW, TO BE INSTALLED, REMOVE, RECOVER, TO BE REMOVED, TO BE RECOVERED, TO BE REPLACED, TO BE RELOCATED, TO BE MODIFIED, TO BE MOVED.
   - Existing / Non-Action: If an item is marked as EXISTING, REUSE, or designated as SPARE / MADE SPARE with NO active work scope or proposed count (PROPOSED: 0), SKIP and DO NOT output it.
   - Spare with active action: If an item marked as SPARE explicitly has an active action (e.g. REMOVE SPARE ANTENNA), process the action.

4. PRICE BOOK MAPPING & UNQUOTED ITEMS:
   - Map each active item to the most appropriate Price Book item matching its action, equipment type, and description.
   - Removal actions: Concentrate strictly on item count/quantity (do not require technology breakdown for removals).
   - If an active scope of work is required for an item, but NO matching SOR code / item exists in the Price Book:
     - Set "row_idx" to null, "sor_code" to "UNQUOTED", "rate" to 0.0, "total_cost" to 0.0.
     - Set "comment" to: "Estimator need to fill".

5. OUTPUT STRUCTURE:
Return ONLY a valid JSON array of mapped BOQ objects matching this structure:
[
  {
    "equipment_type": "PANEL ANTENNA",
    "model": "KAELUS F6RHEU01",
    "action": "INSTALL",
    "quantity": 3,
    "source_sheet": "Sheet 9",
    "clean_text": "Install proposed Telstra Kaelus F6RHEU01 panel antenna",
    "row_idx": 45,
    "sor_code": "ANT-001",
    "item_name": "Install 4G Panel Antenna (>1.5m)",
    "unit": "each",
    "rate": 450.0,
    "total_cost": 1350.0,
    "comment": ""
  }
]"""
        cursor.execute(
            "INSERT INTO ai_prompts (name, title, prompt, version, enabled) VALUES (?, ?, ?, ?, ?)",
            ("boq_mapping_engine", "AI BOQ Mapping & Table-First Deduplication Engine Prompt", mapping_prompt, 1, 1)
        )
        conn.commit()
        print("[DB Migration] Created default 'boq_mapping_engine' prompt (version 1).")

    cursor.execute("SELECT COUNT(*) as count FROM ai_prompts WHERE name = 'client_mapping_rules'")
    if cursor.fetchone()["count"] == 0:
        client_rules = """[TELSTRA WIRELESS CLIENT-SPECIFIC DOMAIN RULES]:
1. ANTENNA TECHNOLOGY CLASSIFICATION (INSTALLATIONS):
   - 4G Panel Antennas: Antenna length/height > 1.5m (1500mm), e.g., Kaelus F6RHEU01 (2705mm), Argus RVVPX series.
   - 5G AAU (Active Antenna Unit): Antenna length/height < 1.0m (1000mm), or Ericsson models starting with 'AIR' (e.g., AIR3258, AIR6488, AIRXXXX).
   - Removals: Do NOT differentiate technology; match to general antenna removal/recovery SOR based strictly on quantity.

2. RADIO & RRU (Remote Radio Unit):
   - Provides power and RF signals to antennas. Associated with RF tails and hybrid trunk cables.

3. TMA (Tower Mounted Amplifier) / TMD (Tower Mounted Device):
   - Transfers low-band signals to antennas. If TMA is proposed/installed, feeder cables are required.

4. FILTERS, COMBINERS & MHA:
   - Combine and split technology bands.

5. JUNCTION BOX (W&B JB / Samsung JB):
   - Connects hybrid trunk cables from shelter to JB, split into separate tails for RRU and AAU.

6. FEEDERS & TRUNK CABLES:
   - Feeders triggered if TMA proposed/existing.
   - Hybrid Trunk cables triggered if RRU proposed/existing."""
        cursor.execute(
            "INSERT INTO ai_prompts (name, title, prompt, version, enabled) VALUES (?, ?, ?, ?, ?)",
            ("client_mapping_rules", "Client-Specific Domain Mapping Instructions (Telstra)", client_rules, 1, 1)
        )
        conn.commit()
        print("[DB Migration] Created default 'client_mapping_rules' prompt (version 1).")

    # Migration 14: Upgrade boq_mapping_engine and client_mapping_rules to Version 2 (Universal Zero-Loss Telecom Architecture)
    cursor.execute("SELECT version FROM ai_prompts WHERE name = 'boq_mapping_engine'")
    row = cursor.fetchone()
    if row:
        current_version = row["version"]
        if current_version is None or current_version < 2:
            v2_mapping_prompt = """You are a senior telecom Bill of Quantities (BOQ) estimator and universal pricing AI engine.
Your task is to analyze all extracted drawing data (structured schedules, layout notes, revision clouds, elevation details, and equipment notes) and map every single active scope of work to the active Price Book Schedule of Rates (SOR).

CRITICAL ARCHITECTURAL DIRECTIVES:

1. ZERO-LOSS SCOPE GUARANTEE (FINANCIAL CRITICAL):
   - Every active work scope, proposed equipment item, removal action, structural modification, civil fixing, testing requirement, or preliminary task extracted from the drawing MUST appear in the final BOQ output.
   - If an active item/work scope MATCHES an item in the Price Book:
     - Map to that Price Book item with its exact "row_idx", "sor_code", "rate", "unit", and compute "total_cost".
   - If an active item/work scope DOES NOT exist in the Price Book (or is a custom civil/structural/non-SOR scope):
     - DO NOT OMIT OR DISCARD IT!
     - Set "row_idx": null, "sor_code": "UNQUOTED", "rate": 0.0, "total_cost": 0.0.
     - Set "comment": "Estimator need to fill: <Detailed extracted scope, dimensions, hardware specs, and drawing sheet reference>".

2. TABLE-FIRST PRIMARY AUTHORITY & CROSS-VERIFICATION:
   - Structured Tables are the primary source of truth for equipment quantities and specifications.
   - Antenna Configuration Tables take precedence for Antenna items (Panel Antennas, AAU, GPS antennas) and antenna removals.
   - Equipment Notes tables take precedence for internal shelter hardware, racks, RRUs, TMAs, filters, and DC power equipment.
   - For every table item, cross-check with layout callouts and revision clouds:
     - If table quantity and layout annotations MATCH: output table quantity with comment "".
     - If table quantity and layout annotations DIFFER: output the TABLE quantity as authoritative count, and set comment: "Data not matching with antenna layout".

3. 5-TIER SCOPE TAXONOMY:
   Tier 1: RF & Antennas (4G Panels >1.5m, 5G AAUs <1.0m / active beamforming, 1st antenna per sector vs extra-over, general antenna removals).
   Tier 2: Active Radios (RRU) & Tower Mounted Devices (TMA, Filters, Combiners, Diplexers, MHAs) for both install and removal.
   Tier 3: Internal Shelter, Baseband & DC Power (Baseband units/RP6672, Cell Site Routers, Digital Units, internal filter recoveries, battery strings, rectifiers).
   Tier 4: Feeders, Tails, Cabling & Commissioning Testing:
     - Automatically derive Blackbird testing line items: 1st Carrier testing per sector (count = active sectors) + Subsequent Carrier testing (count = total active carriers across sectors minus 1st carriers).
     - Feeder PIM / Sweep testing for reused or proposed feeder lines.
   Tier 5: Structural Mounts, Plinths, Civil & Preliminaries:
     - New mounts, mount relocations, plinth removals/replacements, Hilti chemical anchors with embedment depth, EME chain barriers, roof handrails, tower inspections, FIM waste management, crane hire, traffic control.

4. ACTION FILTERING:
   - Active Actions to include: INSTALL, PROPOSED, NEW, TO BE INSTALLED, REMOVE, RECOVER, TO BE REMOVED, TO BE RECOVERED, TO BE REPLACED, TO BE RELOCATED, TO BE MODIFIED, TO BE MOVED.
   - Non-Action: Items marked purely as EXISTING, REUSE, or SPARE / MADE SPARE with NO active work scope (PROPOSED: 0) must be skipped.
   - If an item marked as SPARE explicitly has an active action (e.g. REMOVE SPARE ANTENNA), include and process it.

5. OUTPUT STRUCTURE:
Return ONLY a valid JSON array of mapped BOQ objects matching this schema:
[
  {
    "equipment_type": "PANEL ANTENNA",
    "model": "KAELUS F6RHEU01",
    "action": "INSTALL",
    "quantity": 3,
    "source_sheet": "Sheet S3-3",
    "clean_text": "Install proposed Telstra Kaelus F6RHEU01 panel antenna",
    "row_idx": 45,
    "sor_code": "W7520",
    "item_name": "One panel Antenna",
    "unit": "each",
    "rate": 675.0,
    "total_cost": 2025.0,
    "comment": ""
  }
]"""
            cursor.execute(
                "UPDATE ai_prompts SET prompt = ?, version = 2 WHERE name = 'boq_mapping_engine'",
                (v2_mapping_prompt,)
            )
            conn.commit()
            print("[DB Migration] Upgraded 'boq_mapping_engine' prompt to version 2.")

    cursor.execute("SELECT version FROM ai_prompts WHERE name = 'client_mapping_rules'")
    row = cursor.fetchone()
    if row:
        current_version = row["version"]
        if current_version is None or current_version < 2:
            v2_client_rules = """[UNIVERSAL TELECOM CLIENT-SPECIFIC DOMAIN RULES]:
1. ANTENNA TECHNOLOGY CLASSIFICATION:
   - Primary 4G Panel Antenna: Length/height > 1.5m (1500mm), e.g. Kaelus F6RHEU01, Argus RVVPX series. First panel antenna per sector maps to primary antenna SOR (e.g. W7520).
   - Extra-Over 4G Panel Antenna: Second or additional panel antenna on the same sector maps to Extra-Over SOR (e.g. W13360).
   - 5G AAU (Active Antenna Unit) / Massive MIMO: Compact height < 1.0m (1000mm) or active beamforming (e.g. AIR3258, AIR6488, AAU series) maps to 5G AAU SOR (e.g. W13358).
   - Antenna Removals: Map strictly by total quantity count to general antenna removal/recovery SOR (e.g. R12513), without differentiating technology.

2. RADIOS (RRU) & TOWER MOUNTED DEVICES (TMD):
   - RRU Installation: All remote radio units mounted on tower/mounts map to Remote Radio Unit SOR (e.g. W12252).
   - RRU Removal: Map to RRU Removal SOR (e.g. R12513).
   - Tower Filters / TMA / Combiners: Map to Tower Mounted Device SOR (e.g. W7893) for install, and removal SOR (e.g. R12513).

3. INTERNAL SHELTER, BASEBAND & POWER:
   - Baseband / Radio Processors: Proposed baseband units (e.g. RP6672, Baseband 6630/6648) map to Baseband Unit Installation SOR (e.g. W13393).
   - Baseband Recovery: Recovered DUS, R503, or baseband units map to Baseband Recovery SOR (e.g. R13701).
   - Cell Site Routers: Relocations or installs map to Router SOR (e.g. W13700).
   - Internal Filters: Recovered internal filters/combiners map to internal filter recovery SOR (e.g. R13169).

4. COMMISSIONING & TESTING:
   - 4G/5G Testing (Blackbird / Call & Data Tests):
     - First carrier per sector: Qty = total active sectors (e.g. W13374).
     - Subsequent carriers per sector: Qty = sum of (carriers per sector - 1) across all sectors (e.g. W13400).
   - PIM / Sweep Testing: Map to PIM testing SOR (e.g. W13375) when reusing existing feeder lines or installing new RF tails.

5. STRUCTURAL, CIVIL & PRELIMINARIES:
   - Tier 2 Tower Inspections: Auto-include standard tower inspection SOR (e.g. W13398) for macro build completion.
   - Antenna Mounts, Plinths & Hilti Anchors: If new mounts, plinth replacements, or Hilti chemical anchors are specified in notes/clouds, map to matching SOR or emit as UNQUOTED with exact specs for estimator pricing.
   - Site Safety & Preliminaries: EME chain barrier, roof handrail, crane hire, traffic control, and FIM waste management should be captured with full estimator notes."""
            cursor.execute(
                "UPDATE ai_prompts SET prompt = ?, version = 2 WHERE name = 'client_mapping_rules'",
                (v2_client_rules,)
            )
            conn.commit()
            print("[DB Migration] Upgraded 'client_mapping_rules' prompt to version 2.")

    # Migration 15: Seed DEFAULT_MAPPING_RULES if mapping_rules table is empty
    cursor.execute("SELECT COUNT(*) as count FROM mapping_rules")
    if cursor.fetchone()["count"] == 0:
        seed_default_mapping_rules(cursor)
        conn.commit()
        print("[DB Migration] Seeded default dynamic mapping rules into SQLite.")

    # Check if empty, bootstrap from Excel if so

    cursor.execute("SELECT COUNT(*) as count FROM price_items")
    count = cursor.fetchone()["count"]
    if count == 0:
        bootstrap_from_excel(conn, 1)
    conn.close()

def bootstrap_from_excel(conn, price_list_id=1):
    """Parses original master Excel workbook and populates the SQLite database."""
    template_paths = [
        r"C:\Users\AnilShebinSJ\Downloads\Telstra Wireless pricebook V5.7-Copy.xlsx",
        r"C:\Users\AnilShebinSJ\Downloads\Telstra Wireless pricebook V5.7 - Copy.xlsx",
        r"C:\Users\AnilShebinSJ\Downloads\Telstra Wireless pricebook V5.6 - Copy.xlsx",
        os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads", "active_price_list.xlsx")
    ]
    excel_path = None
    for p in template_paths:
        if os.path.exists(p):
            excel_path = p
            break

    if not excel_path:
        print("[DB] No Excel template found to bootstrap SQLite.")
        return

    print(f"[DB] Bootstrapping SQLite from Excel template: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "SOR" in wb.sheetnames:
        sheet = wb["SOR"]
    elif "Schedule of Rates" in wb.sheetnames:
        sheet = wb["Schedule of Rates"]
    elif "Schedule Of Rates" in wb.sheetnames:
        sheet = wb["Schedule Of Rates"]
    else:
        sheet = wb.active

    header_row_idx = 1
    code_col, name_col, unit_col, rate_col = 1, 2, 3, 4

    for row in range(1, 15):
        row_values = [str(sheet.cell(row, col).value or "").strip().upper() for col in range(1, 10)]
        if any("ITEM NAME" in val or "RATE" in val or "EXCLUDING GST" in val for val in row_values):
            header_row_idx = row
            for col_idx, val in enumerate(row_values, 1):
                if any(x in val for x in ["CODE", "WCODE", "SOR"]):
                    code_col = col_idx
                elif "ITEM NAME" in val or "DESCRIPTION" in val:
                    name_col = col_idx
                elif "UNIT" in val:
                    unit_col = col_idx
                elif "RATE" in val or "PRICE" in val:
                    rate_col = col_idx
            break

    col1_is_code = False
    for r in range(header_row_idx + 1, min(header_row_idx + 15, sheet.max_row + 1)):
        val_a = str(sheet.cell(r, 1).value or "").strip()
        if val_a and any(val_a.upper().startswith(x) for x in ["W", "R"]) and val_a[1:].isdigit():
            col1_is_code = True
            break

    if col1_is_code:
        code_col = 1
        if name_col == 1:
            name_col = 2
    elif code_col == name_col:
        code_col = 1
        name_col = 2

    current_category = "General SOR Pricing Items"
    cursor = conn.cursor()

    for r in range(1, sheet.max_row + 1):
        code = str(sheet.cell(r, code_col).value or "").strip()
        name = str(sheet.cell(r, name_col).value or "").strip()
        unit = str(sheet.cell(r, unit_col).value or "").strip()
        rate_val = sheet.cell(r, rate_col).value

        is_empty = not code and not name and not unit and rate_val is None
        if is_empty:
            continue

        is_valid_code = False
        if code:
            is_valid_code = (" " not in code) and (len(code) <= 12)

        is_action_item = False
        if name:
            name_lower = name.lower().strip()
            if name_lower.startswith(("recover", "remove", "uninstall", "relocate", "install")):
                is_action_item = True

        if r > header_row_idx and (rate_val is not None or unit or is_valid_code or is_action_item):
            rate = 0.0
            if rate_val is not None:
                try:
                    rate = float(str(rate_val).replace('$', '').replace(',', '').strip())
                except ValueError:
                    pass
            qty = 0.0
            qty_val = sheet.cell(r, 5).value
            if qty_val is not None:
                try:
                    qty = float(str(qty_val).strip())
                except ValueError:
                    pass
            cursor.execute(
                "INSERT INTO price_items (code, name, unit, rate, quantity, category, price_list_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (code, name, unit, rate, qty, current_category, price_list_id)
            )
        else:
            cat_name = name or code
            if cat_name and "ITEM NAME" not in cat_name.upper() and "RATE" not in cat_name.upper():
                current_category = cat_name

    conn.commit()
    wb.close()
    print("[DB] Bootstrapping completed successfully.")


# =========================================================================
# DYNAMIC USER-CONFIGURABLE MAPPING RULES ENGINE (SQLite + Excel)
# =========================================================================
# TELECOM PRICE REGISTRY (CANONICAL INTERNAL ITEMS)
# =========================================================================

DEFAULT_PRICE_REGISTRY_ITEMS = [
    {
        "internal_id": "TEL_ANT_PANEL_PRIMARY_4G",
        "standard_name": "Standard 4G Panel Antenna (>= 1500mm) - Primary",
        "category": "Antennas",
        "default_unit": "each",
        "fingerprint": "PANEL, 4G, PRIMARY, ANTENNA, >1.5M",
        "description": "First panel antenna installed per sector (> 1500mm)"
    },
    {
        "internal_id": "TEL_ANT_PANEL_EXTRA_OVER",
        "standard_name": "Standard 4G Panel Antenna (>= 1500mm) - Extra-Over",
        "category": "Antennas",
        "default_unit": "each",
        "fingerprint": "PANEL, 4G, EXTRA OVER, SUBSEQUENT, ANTENNA",
        "description": "Subsequent panel antenna installed on same sector"
    },
    {
        "internal_id": "TEL_ANT_5G_AAU",
        "standard_name": "5G Active Antenna Unit (AAU / Massive MIMO)",
        "category": "Antennas",
        "default_unit": "each",
        "fingerprint": "5G, ACTIVE, AAU, MASSIVE MIMO, AIR",
        "description": "Active beamforming 5G antenna unit"
    },
    {
        "internal_id": "TEL_ANT_REMOVAL_ALL",
        "standard_name": "Remove Panel / Macro Antenna",
        "category": "Antennas",
        "default_unit": "each",
        "fingerprint": "REMOVE, RECOVER, ANTENNA, PANEL",
        "description": "Removal and recovery of existing macro antenna"
    },
    {
        "internal_id": "TEL_GPS_REPLACE_TOWER",
        "standard_name": "GPS Antenna & Receiver Replacement (Tower)",
        "category": "Antennas",
        "default_unit": "each",
        "fingerprint": "GPS, RECEIVER, REPLACEMENT, KA-7005",
        "description": "Tower top GPS antenna and receiver replacement"
    },
    {
        "internal_id": "TEL_RADIO_TOWER_RRU_INSTALL",
        "standard_name": "Remote Radio Unit (RRU) Install - Tower Top",
        "category": "Radios & TMD",
        "default_unit": "each",
        "fingerprint": "RRU, RADIO, TOWER, INSTALL, REMOTE RADIO",
        "description": "Installation of Remote Radio Unit on tower/mount"
    },
    {
        "internal_id": "TEL_RADIO_TOWER_RRU_REMOVE",
        "standard_name": "Remote Radio Unit (RRU) Removal - Tower Top",
        "category": "Radios & TMD",
        "default_unit": "each",
        "fingerprint": "RRU, RADIO, TOWER, REMOVE, RECOVERY",
        "description": "Removal and recovery of tower-mounted RRU"
    },
    {
        "internal_id": "TEL_TMD_TMA_INSTALL",
        "standard_name": "Tower Mounted Device (TMA / Filter) Installation",
        "category": "Radios & TMD",
        "default_unit": "each",
        "fingerprint": "TMA, TMD, FILTER, COMBINER, TOWER",
        "description": "Installation of TMA or filter on tower top"
    },
    {
        "internal_id": "TEL_TMD_FILTER_REMOVE",
        "standard_name": "Tower Mounted Device (TMA / Filter) Removal",
        "category": "Radios & TMD",
        "default_unit": "each",
        "fingerprint": "TMA, TMD, FILTER, REMOVE, TOWER",
        "description": "Removal of tower-mounted TMA or filter"
    },
    {
        "internal_id": "TEL_RADIO_SHELTER_RRU_INSTALL",
        "standard_name": "RRU / Radio Installation in Shelter",
        "category": "Shelter & Baseband",
        "default_unit": "each",
        "fingerprint": "RRU, RADIO, SHELTER, INTERNAL, RACK",
        "description": "Internal shelter rack radio unit installation"
    },
    {
        "internal_id": "TEL_RADIO_SHELTER_RRU_RECOVER",
        "standard_name": "RRU / Radio Recovery from Shelter",
        "category": "Shelter & Baseband",
        "default_unit": "each",
        "fingerprint": "RRU, RADIO, SHELTER, RECOVERY",
        "description": "Recovery of radio unit from internal shelter"
    },
    {
        "internal_id": "TEL_BASEBAND_INSTALL",
        "standard_name": "Baseband Unit / RP6672 Installation",
        "category": "Shelter & Baseband",
        "default_unit": "each",
        "fingerprint": "BASEBAND, RP6672, BB6630, DUS, SHELTER",
        "description": "Baseband or digital unit installation in shelter"
    },
    {
        "internal_id": "TEL_BASEBAND_RECOVER",
        "standard_name": "Recovery of Baseband Unit, DUS or R503",
        "category": "Shelter & Baseband",
        "default_unit": "each",
        "fingerprint": "BASEBAND, DUS, R503, RECOVERY, SHELTER",
        "description": "Recovery of legacy baseband or DUS unit"
    },
    {
        "internal_id": "TEL_TMD_SHELTER_FILTER_RECOVER",
        "standard_name": "Internal Shelter Filter / Combiner Recovery",
        "category": "Shelter & Baseband",
        "default_unit": "each",
        "fingerprint": "FILTER, COMBINER, BANDSTOP, SHELTER, RECOVER",
        "description": "Removal of internal shelter bandstop filter or combiner"
    },
    {
        "internal_id": "TEL_TEST_4G_FIRST_CARRIER",
        "standard_name": "4G/5G Testing - First Carrier Per Sector",
        "category": "Testing",
        "default_unit": "each",
        "fingerprint": "TESTING, BLACKBIRD, FIRST CARRIER, SECTOR",
        "description": "Commissioning test for 1st carrier per sector"
    },
    {
        "internal_id": "TEL_TEST_5G_CARRIER",
        "standard_name": "4G/5G Testing - Subsequent Carrier Per Sector",
        "category": "Testing",
        "default_unit": "each",
        "fingerprint": "TESTING, BLACKBIRD, EXTRA CARRIER",
        "description": "Commissioning test for subsequent carriers per sector"
    },
    {
        "internal_id": "TEL_TEST_PIM_SWEEP",
        "standard_name": "PIM & Sweep Testing on Reused Feeders",
        "category": "Testing",
        "default_unit": "each",
        "fingerprint": "PIM, SWEEP, FEEDER, TEST",
        "description": "PIM and return loss sweep testing on feeders"
    },
    {
        "internal_id": "TEL_PRELIM_TOWER_INSPECTION",
        "standard_name": "Tier 2 Tower Inspections BAU",
        "category": "Civil & Preliminaries",
        "default_unit": "each",
        "fingerprint": "TIER 2, TOWER, INSPECTION, BAU",
        "description": "Tier 2 tower inspection on build completion"
    },
    {
        "internal_id": "TEL_SHELTER_DCDU_INSTALL",
        "standard_name": "DCDU Installation in Shelter",
        "category": "Equipment installed in the shelter",
        "default_unit": "each",
        "fingerprint": "DCDU, DC POWER, DISTRIBUTION UNIT, SHELTER",
        "description": "Installation of DC Distribution Unit (DCDU) in equipment rack"
    }
]


# =========================================================================
# VENMO-STANDARD DECLARATIVE MAPPING RULES (SQLite + JSON)
# =========================================================================

DEFAULT_MAPPING_RULES = [
    {
        "rule_name": "One panel Antenna",
        "internal_id": "R001",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "PANEL_ANTENNA",
        "match_keywords": "KAELUS, RVVPX, ARGUS, PANEL, 4G, T-MOBILE, COMMSCOPE, DELTEC, RFS",
        "exclude_keywords": "FILTER, COMBINER, TMA, TMD, RRU, RADIO, GPS, OMNI, AIR, AAU, 5G, ACTIVE",
        "condition_expr": "",
        "action_filter": "INSTALL",
        "target_sor_code": "W7520",
        "target_sor_name": "One panel Antenna",
        "qty_formula": "first_per_sector",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Site Layout; Antenna Layout; Elevation",
        "matching_conditions": "New/proposed panel antenna",
        "notes": "Install 4G antenna",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "ANTENNA"},
                {"name": "action", "operator": "equal_to", "value": "INSTALL"},
                {"name": "sector_index", "operator": "equal_to", "value": 1},
                {"name": "is_active", "operator": "is_false", "value": True}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R001",
                    "sor_code": "W7520",
                    "target_name": "One panel Antenna"
                }
            }
        ])
    },
    {
        "rule_name": "One panel Antenna - extra over",
        "internal_id": "R002",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "PANEL_ANTENNA",
        "match_keywords": "KAELUS, RVVPX, ARGUS, PANEL, 4G, T-MOBILE, COMMSCOPE, DELTEC, RFS",
        "exclude_keywords": "FILTER, COMBINER, TMA, TMD, RRU, RADIO, GPS, OMNI, AIR, AAU, 5G, ACTIVE",
        "condition_expr": "",
        "action_filter": "INSTALL",
        "target_sor_code": "W13360",
        "target_sor_name": "One panel Antenna - extra over",
        "qty_formula": "extra_per_sector",
        "comment_template": "",
        "priority": 90,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Site Layout; Antenna Layout; Elevation",
        "matching_conditions": "Additional panel antenna after first",
        "notes": "Applies to additional antenna",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "ANTENNA"},
                {"name": "action", "operator": "equal_to", "value": "INSTALL"},
                {"name": "sector_index", "operator": "greater_than", "value": 1},
                {"name": "is_active", "operator": "is_false", "value": True}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R002",
                    "sor_code": "W13360",
                    "target_name": "One panel Antenna - extra over"
                }
            }
        ])
    },
    {
        "rule_name": "One 5G AAU - first",
        "internal_id": "R003",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "5G_AAU",
        "match_keywords": "AIR, AAU, 5G, MASSIVE MIMO, AIR3258, AIR6488, AIR3268, AIR1641",
        "exclude_keywords": "FILTER, COMBINER, TMA, TMD, RRU, RADIO, GPS",
        "condition_expr": "",
        "action_filter": "INSTALL",
        "target_sor_code": "W13358",
        "target_sor_name": "One 5G AAU - first",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Site Layout; Antenna Layout; Elevation",
        "matching_conditions": "New/proposed 5G AAU",
        "notes": "First 5G AAU",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "ANTENNA"},
                {"name": "action", "operator": "equal_to", "value": "INSTALL"},
                {"name": "is_active", "operator": "is_true", "value": True},
                {"name": "sector_index", "operator": "equal_to", "value": 1}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R003",
                    "sor_code": "W13358",
                    "target_name": "One 5G AAU - first"
                }
            }
        ])
    },
    {
        "rule_name": "One 5G AAU - extra over",
        "internal_id": "R004",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "5G_AAU",
        "match_keywords": "AIR, AAU, 5G, MASSIVE MIMO, AIR3258, AIR6488, AIR3268, AIR1641",
        "exclude_keywords": "FILTER, COMBINER, TMA, TMD, RRU, RADIO, GPS",
        "condition_expr": "",
        "action_filter": "INSTALL",
        "target_sor_code": "W13359",
        "target_sor_name": "One 5G AAU - extra over",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 90,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Site Layout; Antenna Layout; Elevation",
        "matching_conditions": "Additional 5G AAU",
        "notes": "Additional AAU",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "ANTENNA"},
                {"name": "action", "operator": "equal_to", "value": "INSTALL"},
                {"name": "is_active", "operator": "is_true", "value": True},
                {"name": "sector_index", "operator": "greater_than", "value": 1}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R004",
                    "sor_code": "W13359",
                    "target_name": "One 5G AAU - extra over"
                }
            }
        ])
    },
    {
        "rule_name": "Remove Panel Antenna or tower mounted device",
        "internal_id": "R005",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "PANEL_ANTENNA",
        "match_keywords": "ANTENNA, PANEL, ARGUS, DELTEC, AIR, KAELUS, RVVPX, RFS, COMMSCOPE",
        "exclude_keywords": "GPS, OMNI, FILTER, COMBINER, TMA, TMD, RRU, RADIO",
        "condition_expr": "",
        "action_filter": "REMOVE",
        "target_sor_code": "R12513",
        "target_sor_name": "Remove Panel Antenna or tower mounted device",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Site Layout; Antenna Layout; Elevation",
        "matching_conditions": "Existing antenna marked REMOVE/RECOVER",
        "notes": "Remove only",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "ANTENNA"},
                {"name": "action", "operator": "equal_to", "value": "REMOVE"}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R005",
                    "sor_code": "R12513",
                    "target_name": "Remove Panel Antenna or tower mounted device"
                }
            }
        ])
    },
    {
        "rule_name": "One Panel antenna relocation",
        "internal_id": "R006",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "PANEL_ANTENNA",
        "match_keywords": "ANTENNA, PANEL, ARGUS, DELTEC, AIR, KAELUS, RVVPX, RFS, COMMSCOPE",
        "exclude_keywords": "GPS, OMNI, FILTER, COMBINER, TMA, TMD, RRU, RADIO",
        "condition_expr": "",
        "action_filter": "RELOCATE",
        "target_sor_code": "W13178",
        "target_sor_name": "One Panel antenna relocation",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Site Layout; Antenna Layout; Elevation",
        "matching_conditions": "Existing antenna moved from one location to another",
        "notes": "Same antenna, new location",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "ANTENNA"},
                {"name": "action", "operator": "equal_to", "value": "RELOCATE"}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R006",
                    "sor_code": "W13178",
                    "target_name": "One Panel antenna relocation"
                }
            }
        ])
    },
    {
        "rule_name": "GPS system",
        "internal_id": "R007",
        "category": "Recover GPS system",
        "equipment_type": "GPS",
        "match_keywords": "GPS, GPS ANTENNA, KA-7005, LTE700 GPS",
        "exclude_keywords": "",
        "condition_expr": "",
        "action_filter": "INSTALL",
        "target_sor_code": "W12804",
        "target_sor_name": "GPS system",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Site Layout; Antenna Layout; Elevation",
        "matching_conditions": "New GPS antenna/system",
        "notes": "Install new GPS antenna",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "GPS"},
                {"name": "action", "operator": "equal_to", "value": "INSTALL"}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R007",
                    "sor_code": "W12804",
                    "target_name": "GPS system"
                }
            }
        ])
    },
    {
        "rule_name": "Recover GPS system",
        "internal_id": "R008",
        "category": "Recover GPS system",
        "equipment_type": "GPS",
        "match_keywords": "GPS, GPS ANTENNA, KA-7005",
        "exclude_keywords": "",
        "condition_expr": "",
        "action_filter": "REMOVE",
        "target_sor_code": "",
        "target_sor_name": "Recover GPS system",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Site Layout; Antenna Layout; Elevation",
        "matching_conditions": "GPS marked REMOVE/RECOVER",
        "notes": "Remove only",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "GPS"},
                {"name": "action", "operator": "equal_to", "value": "REMOVE"}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R008",
                    "sor_code": "",
                    "target_name": "Recover GPS system"
                }
            }
        ])
    },
    {
        "rule_name": "Tower Mounted Device (TMA, COM, FILTER)",
        "internal_id": "R009",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "TMA_FILTER",
        "match_keywords": "TMA, TMD, FILTER, COMBINER, DIPLEXER, TRIPLEXER",
        "exclude_keywords": "",
        "condition_expr": "",
        "action_filter": "INSTALL",
        "target_sor_code": "W7893",
        "target_sor_name": "Tower Mounted Device (TMA, COM, FILTER)",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Antenna Layout; Elevation; Site Layout",
        "matching_conditions": "TMA/COM/FILTER marked INSTALL",
        "notes": "Separate scope from antenna",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "TMA_FILTER"},
                {"name": "action", "operator": "equal_to", "value": "INSTALL"}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R009",
                    "sor_code": "W7893",
                    "target_name": "Tower Mounted Device (TMA, COM, FILTER)"
                }
            }
        ])
    },
    {
        "rule_name": "Tower Mounted Device Removal",
        "internal_id": "R010",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "TMA_FILTER",
        "match_keywords": "TMA, TMD, FILTER, COMBINER, DIPLEXER, TRIPLEXER",
        "exclude_keywords": "",
        "condition_expr": "",
        "action_filter": "REMOVE",
        "target_sor_code": "R12513",
        "target_sor_name": "Tower Mounted Device Removal",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Antenna Layout; Elevation; Site Layout",
        "matching_conditions": "TMA/COM/FILTER marked REMOVE",
        "notes": "Do not merge with antenna",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "TMA_FILTER"},
                {"name": "action", "operator": "equal_to", "value": "REMOVE"}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R010",
                    "sor_code": "R12513",
                    "target_name": "Tower Mounted Device Removal"
                }
            }
        ])
    },
    {
        "rule_name": "Tower Mounted Device Relocation",
        "internal_id": "R011",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "TMA_FILTER",
        "match_keywords": "TMA, TMD, FILTER, COMBINER, DIPLEXER, TRIPLEXER",
        "exclude_keywords": "",
        "condition_expr": "",
        "action_filter": "RELOCATE",
        "target_sor_code": "W13178",
        "target_sor_name": "Tower Mounted Device Relocation",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Antenna Layout; Elevation; Site Layout",
        "matching_conditions": "TMA/COM/FILTER relocated",
        "notes": "Separate scope",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "TMA_FILTER"},
                {"name": "action", "operator": "equal_to", "value": "RELOCATE"}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R011",
                    "sor_code": "W13178",
                    "target_name": "Tower Mounted Device Relocation"
                }
            }
        ])
    },
    {
        "rule_name": "Remote Radio Unit (RRU)",
        "internal_id": "R012",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "RRU",
        "match_keywords": "RADIO, RRU, RRUS, RUS",
        "exclude_keywords": "",
        "condition_expr": "",
        "action_filter": "INSTALL",
        "target_sor_code": "W12252",
        "target_sor_name": "Remote Radio Unit (RRU)",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Antenna Layout; Elevation",
        "matching_conditions": "RRU marked INSTALL",
        "notes": "First RRU",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "RRU"},
                {"name": "action", "operator": "equal_to", "value": "INSTALL"},
                {"name": "sector_index", "operator": "equal_to", "value": 1}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R012",
                    "sor_code": "W12252",
                    "target_name": "Remote Radio Unit (RRU)"
                }
            }
        ])
    },
    {
        "rule_name": "RRU x 1 - Extra over W12252",
        "internal_id": "R013",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "RRU",
        "match_keywords": "RADIO, RRU, RRUS, RUS",
        "exclude_keywords": "",
        "condition_expr": "",
        "action_filter": "INSTALL",
        "target_sor_code": "W12253",
        "target_sor_name": "RRU x 1 - Extra over W12252",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 90,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Antenna Layout; Elevation",
        "matching_conditions": "Additional RRU",
        "notes": "Additional RRU",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "RRU"},
                {"name": "action", "operator": "equal_to", "value": "INSTALL"},
                {"name": "sector_index", "operator": "greater_than", "value": 1}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R013",
                    "sor_code": "W12253",
                    "target_name": "RRU x 1 - Extra over W12252"
                }
            }
        ])
    },
    {
        "rule_name": "Remote Radio Unit (RRU) Removal",
        "internal_id": "R014",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "RRU",
        "match_keywords": "RADIO, RRU, RRUS, RUS",
        "exclude_keywords": "",
        "condition_expr": "",
        "action_filter": "REMOVE",
        "target_sor_code": "R12513",
        "target_sor_name": "Remote Radio Unit (RRU) Removal",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Antenna Layout; Elevation",
        "matching_conditions": "RRU marked REMOVE",
        "notes": "Separate from antenna",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "RRU"},
                {"name": "action", "operator": "equal_to", "value": "REMOVE"}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R014",
                    "sor_code": "R12513",
                    "target_name": "Remote Radio Unit (RRU) Removal"
                }
            }
        ])
    },
    {
        "rule_name": "Remote Radio Unit (RRU) Relocation",
        "internal_id": "R015",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "RRU",
        "match_keywords": "RADIO, RRU, RRUS, RUS",
        "exclude_keywords": "",
        "condition_expr": "",
        "action_filter": "RELOCATE",
        "target_sor_code": "W13178",
        "target_sor_name": "Remote Radio Unit (RRU) Relocation",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Antenna Layout; Elevation",
        "matching_conditions": "RRU moved to another location",
        "notes": "Same RRU, new location",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "RRU"},
                {"name": "action", "operator": "equal_to", "value": "RELOCATE"}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R015",
                    "sor_code": "W13178",
                    "target_name": "Remote Radio Unit (RRU) Relocation"
                }
            }
        ])
    },
    {
        "rule_name": "External Combiner",
        "internal_id": "R016",
        "category": "Antennas, RRUs, TMDs (Ex plant, Inc testing, FIM)",
        "equipment_type": "COMBINER",
        "match_keywords": "COMBINER, EXTERNAL COMBINER",
        "exclude_keywords": "",
        "condition_expr": "",
        "action_filter": "INSTALL",
        "target_sor_code": "",
        "target_sor_name": "External Combiner",
        "qty_formula": "table_qty",
        "comment_template": "",
        "priority": 100,
        "enabled": 1,
        "primary_source": "",
        "preferred_source_type": "TABLE",
        "ignore_pages": "Drawing Index; Document Control",
        "duplicate_prone_pages": "Site Layout; Elevation",
        "matching_conditions": "Combiner marked INSTALL + technology specified",
        "notes": "700/800 technology",
        "conditions_json": json.dumps({
            "all": [
                {"name": "category", "operator": "equal_to", "value": "COMBINER"},
                {"name": "action", "operator": "equal_to", "value": "INSTALL"}
            ]
        }),
        "actions_json": json.dumps([
            {
                "name": "assign_price_item",
                "params": {
                    "internal_id": "R016",
                    "sor_code": "",
                    "target_name": "External Combiner"
                }
            }
        ])
    }
]


def seed_default_price_registry(cursor):
    """Inserts default canonical items into price_registry table."""
    for item in DEFAULT_PRICE_REGISTRY_ITEMS:
        cursor.execute("""
            INSERT OR REPLACE INTO price_registry (
                internal_id, standard_name, category, default_unit, fingerprint, description
            ) VALUES (?, ?, ?, ?, ?, ?)
        """, (
            item["internal_id"],
            item["standard_name"],
            item.get("category", "General"),
            item.get("default_unit", "each"),
            item.get("fingerprint", ""),
            item.get("description", "")
        ))


def seed_default_mapping_rules(cursor):
    """Inserts DEFAULT_MAPPING_RULES into SQLite mapping_rules table."""
    from services.venmo_engine import generate_plain_english_logic
    seed_default_price_registry(cursor)
    for rule in DEFAULT_MAPPING_RULES:
        c_json = rule.get("conditions_json", "")
        a_json = rule.get("actions_json", "")
        logic_exp = generate_plain_english_logic(c_json, a_json)
        cursor.execute("""
            INSERT INTO mapping_rules (
                rule_name, category, equipment_type, match_keywords, exclude_keywords,
                condition_expr, action_filter, target_sor_code, target_sor_name, qty_formula,
                comment_template, priority, enabled, internal_id, conditions_json, actions_json, logic_explanation,
                primary_source, preferred_source_type, ignore_pages, duplicate_prone_pages, matching_conditions, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            rule["rule_name"],
            rule.get("category", "General"),
            rule.get("equipment_type", "EQUIPMENT"),
            rule.get("match_keywords", ""),
            rule.get("exclude_keywords", ""),
            rule.get("condition_expr", ""),
            rule.get("action_filter", "ALL"),
            rule.get("target_sor_code", "UNQUOTED"),
            rule.get("target_sor_name", ""),
            rule.get("qty_formula", "table_qty"),
            rule.get("comment_template", ""),
            rule.get("priority", 100),
            rule.get("enabled", 1),
            rule.get("internal_id", ""),
            c_json,
            a_json,
            logic_exp,
            rule.get("primary_source", ""),
            rule.get("preferred_source_type", "TABLE"),
            rule.get("ignore_pages", ""),
            rule.get("duplicate_prone_pages", ""),
            rule.get("matching_conditions", ""),
            rule.get("notes", "")
        ))


def get_all_mapping_rules() -> List[Dict[str, Any]]:
    """Fetches all mapping rules ordered by priority descending, then id ascending."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mapping_rules ORDER BY priority DESC, id ASC")
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_all_price_registry_items() -> List[Dict[str, Any]]:
    """Fetches all internal canonical price registry items."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM price_registry ORDER BY category ASC, standard_name ASC")
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]


def build_default_venmo_conditions_and_actions(rule_data: Dict[str, Any]) -> Tuple[str, str]:
    """Generates standard Venmo conditions_json and actions_json from UI rule fields if not present."""
    cond_all = []
    
    eq_type = str(rule_data.get("equipment_type") or rule_data.get("category") or "").upper().strip()
    if any(k in eq_type for k in ["PANEL", "AAU", "ANTENNA"]):
        cond_all.append({"name": "category", "operator": "equal_to", "value": "ANTENNA"})
        if "AAU" in eq_type or "5G" in eq_type:
            cond_all.append({"name": "is_active", "operator": "is_true", "value": True})
        elif "PANEL" in eq_type or "4G" in eq_type:
            cond_all.append({"name": "is_active", "operator": "is_false", "value": True})
    elif "RRU" in eq_type or "RADIO" in eq_type:
        cond_all.append({"name": "category", "operator": "equal_to", "value": "RRU"})
    elif any(k in eq_type for k in ["TMA", "FILTER", "TMD", "COMBINER"]):
        cond_all.append({"name": "category", "operator": "equal_to", "value": "TMA_FILTER"})
    elif "GPS" in eq_type:
        cond_all.append({"name": "category", "operator": "equal_to", "value": "GPS"})
    elif any(k in eq_type for k in ["BASEBAND", "ROUTER", "SWITCH"]):
        cond_all.append({"name": "category", "operator": "equal_to", "value": "BASEBAND"})
    elif "DCDU" in eq_type:
        cond_all.append({"name": "category", "operator": "equal_to", "value": "DCDU"})
        
    act = str(rule_data.get("action_filter") or "ALL").upper().strip()
    if act != "ALL":
        cond_all.append({"name": "action", "operator": "equal_to", "value": act})

    formula = str(rule_data.get("qty_formula") or "").lower().strip()
    if formula == "first_per_sector":
        cond_all.append({"name": "sector_index", "operator": "equal_to", "value": 1})
    elif formula == "extra_per_sector":
        cond_all.append({"name": "sector_index", "operator": "greater_than", "value": 1})

    conditions_json = json.dumps({"all": cond_all})
    
    actions_json = json.dumps([
        {
            "name": "assign_price_item",
            "params": {
                "internal_id": rule_data.get("internal_id", ""),
                "sor_code": rule_data.get("target_sor_code", ""),
                "target_name": rule_data.get("target_sor_name", rule_data.get("rule_name", "")),
                "comment": rule_data.get("comment_template", "")
            }
        }
    ])
    
    return conditions_json, actions_json


def create_mapping_rule(rule_data: Dict[str, Any]) -> int:
    """Inserts a new mapping rule and returns its inserted ID."""
    from services.venmo_engine import generate_plain_english_logic
    c_json = rule_data.get("conditions_json")
    a_json = rule_data.get("actions_json")
    matching_conditions = rule_data.get("matching_conditions", "")

    # Auto-generate internal R-number ID if not provided
    internal_id = str(rule_data.get("internal_id") or "").strip()
    if not internal_id:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT internal_id FROM mapping_rules")
        rows = cursor.fetchall()
        conn.close()
        max_num = 0
        import re
        for r in rows:
            val = r["internal_id"] or ""
            match = re.match(r"^R(\d+)$", val, re.IGNORECASE)
            if match:
                num = int(match.group(1))
                if num > max_num:
                    max_num = num
        internal_id = f"R{max_num + 1:03d}"

    # Call dynamic rule compiler if human conditions are provided and JSON is absent
    if matching_conditions and (not c_json or not a_json or str(c_json).strip() == "" or str(a_json).strip() == ""):
        from services.ai_service import compile_natural_language_conditions_with_ai
        c_json, a_json = compile_natural_language_conditions_with_ai(
            rule_name=rule_data.get("rule_name", ""),
            action=rule_data.get("action_filter", "INSTALL"),
            item_type=rule_data.get("equipment_type", "EQUIPMENT"),
            matching_conditions=matching_conditions,
            target_sor_code=rule_data.get("target_sor_code", ""),
            target_sor_name=rule_data.get("rule_name", "")
        )

    if not c_json or not a_json or str(c_json).strip() == "" or str(a_json).strip() == "":
        auto_c, auto_a = build_default_venmo_conditions_and_actions(rule_data)
        c_json = c_json or auto_c
        a_json = a_json or auto_a

    logic_exp = rule_data.get("logic_explanation") or generate_plain_english_logic(c_json, a_json)

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO mapping_rules (
            rule_name, category, equipment_type, match_keywords, exclude_keywords,
            condition_expr, action_filter, target_sor_code, target_sor_name, qty_formula,
            comment_template, priority, enabled, internal_id, conditions_json, actions_json, logic_explanation, regex_pattern,
            status, source, approved_by, rejected_by, version, parent_rule_id, approved_at, rejected_at, simulation_stats,
            primary_source, preferred_source_type, ignore_pages, duplicate_prone_pages, matching_conditions, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        rule_data.get("rule_name", "New Rule"),
        rule_data.get("category", "General"),
        rule_data.get("equipment_type", "EQUIPMENT"),
        rule_data.get("match_keywords", ""),
        rule_data.get("exclude_keywords", ""),
        rule_data.get("condition_expr", ""),
        rule_data.get("action_filter", "ALL"),
        rule_data.get("target_sor_code", "UNQUOTED"),
        rule_data.get("target_sor_name", rule_data.get("rule_name", "")),
        rule_data.get("qty_formula", "table_qty"),
        rule_data.get("comment_template", ""),
        int(rule_data.get("priority", 100)),
        int(rule_data.get("enabled", 1)),
        internal_id,
        c_json,
        a_json,
        logic_exp,
        rule_data.get("regex_pattern", ""),
        rule_data.get("status", "ACTIVE"),
        rule_data.get("source", "HUMAN"),
        rule_data.get("approved_by", ""),
        rule_data.get("rejected_by", ""),
        int(rule_data.get("version", 1)),
        rule_data.get("parent_rule_id"),
        rule_data.get("approved_at"),
        rule_data.get("rejected_at"),
        rule_data.get("simulation_stats", ""),
        rule_data.get("primary_source", ""),
        rule_data.get("preferred_source_type", ""),
        rule_data.get("ignore_pages", ""),
        rule_data.get("duplicate_prone_pages", ""),
        matching_conditions,
        rule_data.get("notes", "")
    ))
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return new_id


def update_mapping_rule(rule_id: int, rule_data: Dict[str, Any]) -> bool:
    """Updates an existing mapping rule."""
    from services.venmo_engine import generate_plain_english_logic
    c_json = rule_data.get("conditions_json")
    a_json = rule_data.get("actions_json")
    matching_conditions = rule_data.get("matching_conditions", "")

    # Call dynamic rule compiler if human conditions are provided and JSON is absent or modified
    if matching_conditions and (not c_json or not a_json or str(c_json).strip() == "" or str(a_json).strip() == ""):
        from services.ai_service import compile_natural_language_conditions_with_ai
        c_json, a_json = compile_natural_language_conditions_with_ai(
            rule_name=rule_data.get("rule_name", ""),
            action=rule_data.get("action_filter", "INSTALL"),
            item_type=rule_data.get("equipment_type", "EQUIPMENT"),
            matching_conditions=matching_conditions,
            target_sor_code=rule_data.get("target_sor_code", ""),
            target_sor_name=rule_data.get("target_sor_name", rule_data.get("rule_name", ""))
        )

    if not c_json or not a_json or str(c_json).strip() == "" or str(a_json).strip() == "":
        auto_c, auto_a = build_default_venmo_conditions_and_actions(rule_data)
        c_json = c_json or auto_c
        a_json = a_json or auto_a

    logic_exp = rule_data.get("logic_explanation") or generate_plain_english_logic(c_json, a_json)

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE mapping_rules SET
            rule_name = ?,
            category = ?,
            equipment_type = ?,
            match_keywords = ?,
            exclude_keywords = ?,
            condition_expr = ?,
            action_filter = ?,
            target_sor_code = ?,
            target_sor_name = ?,
            qty_formula = ?,
            comment_template = ?,
            priority = ?,
            enabled = ?,
            internal_id = ?,
            conditions_json = ?,
            actions_json = ?,
            logic_explanation = ?,
            regex_pattern = ?,
            status = ?,
            source = ?,
            approved_by = ?,
            rejected_by = ?,
            version = ?,
            parent_rule_id = ?,
            approved_at = ?,
            rejected_at = ?,
            simulation_stats = ?,
            primary_source = ?,
            preferred_source_type = ?,
            ignore_pages = ?,
            duplicate_prone_pages = ?,
            matching_conditions = ?,
            notes = ?
        WHERE id = ?
    """, (
        rule_data.get("rule_name", "Updated Rule"),
        rule_data.get("category", "General"),
        rule_data.get("equipment_type", "EQUIPMENT"),
        rule_data.get("match_keywords", ""),
        rule_data.get("exclude_keywords", ""),
        rule_data.get("condition_expr", ""),
        rule_data.get("action_filter", "ALL"),
        rule_data.get("target_sor_code", "UNQUOTED"),
        rule_data.get("target_sor_name", rule_data.get("rule_name", "")),
        rule_data.get("qty_formula", "table_qty"),
        rule_data.get("comment_template", ""),
        int(rule_data.get("priority", 100)),
        int(rule_data.get("enabled", 1)),
        rule_data.get("internal_id", ""),
        c_json,
        a_json,
        logic_exp,
        rule_data.get("regex_pattern", ""),
        rule_data.get("status", "ACTIVE"),
        rule_data.get("source", "HUMAN"),
        rule_data.get("approved_by", ""),
        rule_data.get("rejected_by", ""),
        int(rule_data.get("version", 1)),
        rule_data.get("parent_rule_id"),
        rule_data.get("approved_at"),
        rule_data.get("rejected_at"),
        rule_data.get("simulation_stats", ""),
        rule_data.get("primary_source", ""),
        rule_data.get("preferred_source_type", ""),
        rule_data.get("ignore_pages", ""),
        rule_data.get("duplicate_prone_pages", ""),
        matching_conditions,
        rule_data.get("notes", ""),
        rule_id
    ))
    affected = cursor.rowcount > 0
    conn.commit()
    conn.close()
    return affected


def get_mapping_rule_by_id(rule_id: int) -> Optional[Dict[str, Any]]:
    """Fetches single mapping rule by ID."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mapping_rules WHERE id = ?", (rule_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None


def delete_mapping_rule(rule_id: int) -> bool:
    """Deletes a mapping rule by ID."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM mapping_rules WHERE id = ?", (rule_id,))
    affected = cursor.rowcount > 0
    conn.commit()
    conn.close()
    return affected


def toggle_mapping_rule(rule_id: int) -> Optional[int]:
    """Toggles enabled state (0 <-> 1) for a rule and returns new state."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT enabled FROM mapping_rules WHERE id = ?", (rule_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return None
    new_val = 0 if row["enabled"] == 1 else 1
    cursor.execute("UPDATE mapping_rules SET enabled = ? WHERE id = ?", (new_val, rule_id))
    conn.commit()
    conn.close()
    return new_val


def reset_default_mapping_rules() -> bool:
    """Wipes mapping_rules table and reseeds from DEFAULT_MAPPING_RULES."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM mapping_rules")
    seed_default_mapping_rules(cursor)
    conn.commit()
    conn.close()
    return True


def export_mapping_rules_to_excel(output_path: str) -> bool:
    """Exports all current mapping rules to a formatted Excel workbook matching your custom structure."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        rules = get_all_mapping_rules()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mapping Rules"

        headers = [
            "Rule ID", "Rule Name", "Action", "Item Type", "Category",
            "Preferred Source Type", "Ignore Pages", "Duplicate-Prone Pages",
            "Matching Conditions", "Match Keywords", "Exclude Keywords", "SOR Code", "Notes"
        ]
        
        # Header style
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        fill_dark = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        border_thin = Side(style="thin", color="CBD5E1")
        border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
 
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = fill_dark
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
 
        ws.row_dimensions[1].height = 28
 
        for row_idx, r in enumerate(rules, 2):
            vals = [
                r.get("internal_id") or f"R{r.get('id'):03d}",
                r.get("rule_name", ""),
                r.get("action_filter", "INSTALL"),
                r.get("equipment_type", "EQUIPMENT"),
                r.get("category", "General"),
                r.get("preferred_source_type", "TABLE"),
                r.get("ignore_pages", ""),
                r.get("duplicate_prone_pages", ""),
                r.get("matching_conditions", ""),
                r.get("match_keywords", ""),
                r.get("exclude_keywords", ""),
                r.get("target_sor_code", ""),
                r.get("notes", "")
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Segoe UI", size=9)
                cell.border = border
                if col_idx in [1, 3, 6, 12]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_idx].height = 20

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, len(rules) + 2))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
 
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        wb.close()
        return True
    except Exception as e:
        print(f"[DB] Error exporting mapping rules to Excel: {e}")
        return False
 
 
def import_mapping_rules_from_excel(file_path: str) -> Tuple[int, int]:
    """
    Parses mapping rules from uploaded Excel workbook and updates SQLite mapping_rules table.
    Wipes existing mapping rules before importing to prevent duplicates.
    Returns (imported_count, updated_count).
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Determine column indexes with exact matches to prevent naming collisions
        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(1, col).value or "").strip().upper()
            if "PREFERRED SOURCE TYPE" in val:
                header_map["preferred_source_type"] = col
            elif "CATEGORY" in val:
                header_map["category"] = col
            elif "RULE ID" in val or "INTERNAL ID" in val:
                header_map["internal_id"] = col
            elif "RULE NAME" in val or "NAME" in val:
                header_map["rule_name"] = col
            elif "ITEM TYPE" in val or "EQUIPMENT TYPE" in val:
                header_map["equipment_type"] = col
            elif "IGNORE PAGES" in val:
                header_map["ignore_pages"] = col
            elif "DUPLICATE" in val:
                header_map["duplicate_prone_pages"] = col
            elif "MATCHING CONDITIONS" in val:
                header_map["matching_conditions"] = col
            elif "SOR CODE" in val or "CODE" in val:
                header_map["target_sor_code"] = col
            elif "ACTION" in val:
                header_map["action_filter"] = col
            elif "MATCH KEYWORDS" in val:
                header_map["match_keywords"] = col
            elif "EXCLUDE KEYWORDS" in val:
                header_map["exclude_keywords"] = col
            elif "PRIORITY" in val:
                header_map["priority"] = col
            elif "NOTES" in val:
                header_map["notes"] = col
 
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Clear existing rules before importing new list
        cursor.execute("DELETE FROM mapping_rules")
        
        imported = 0
        for r in range(2, ws.max_row + 1):
            rule_name = str(ws.cell(r, header_map.get("rule_name", 2)).value or "").strip()
            if not rule_name:
                continue
            
            internal_id = str(ws.cell(r, header_map.get("internal_id", 1)).value or "").strip()
            act = str(ws.cell(r, header_map.get("action_filter", 3)).value or "INSTALL").strip().upper()
            eq_type = str(ws.cell(r, header_map.get("equipment_type", 4)).value or "EQUIPMENT").strip()
            category_val = str(ws.cell(r, header_map.get("category", 5)).value or "General").strip()
            pref_src_type = str(ws.cell(r, header_map.get("preferred_source_type", 6)).value or "TABLE").strip()
            ignore_p = str(ws.cell(r, header_map.get("ignore_pages", 7)).value or "").strip()
            dup_prone_p = str(ws.cell(r, header_map.get("duplicate_prone_pages", 8)).value or "").strip()
            matching_cond = str(ws.cell(r, header_map.get("matching_conditions", 9)).value or "").strip()
            match_kws = str(ws.cell(r, header_map.get("match_keywords", 0)).value or "").strip() if header_map.get("match_keywords") else ""
            exclude_kws = str(ws.cell(r, header_map.get("exclude_keywords", 0)).value or "").strip() if header_map.get("exclude_keywords") else ""
            sor = str(ws.cell(r, header_map.get("target_sor_code", 10)).value or "UNQUOTED").strip()
            target_name = rule_name
            notes_val = str(ws.cell(r, header_map.get("notes", 11)).value or "").strip()
            
            prio_col = header_map.get("priority")
            if prio_col is not None:
                prio_val = ws.cell(r, prio_col).value
                try:
                    priority = int(prio_val) if prio_val is not None else 100
                except ValueError:
                    priority = 100
            else:
                priority = 100
                
            enabled = 1  # Default imported rules to enabled
            
            # Call AI rule compiler if Matching Conditions exist
            c_json, a_json = "", ""
            if matching_cond:
                from services.ai_service import compile_natural_language_conditions_with_ai
                c_json, a_json = compile_natural_language_conditions_with_ai(
                    rule_name=rule_name,
                    action=act,
                    item_type=eq_type,
                    matching_conditions=matching_cond,
                    target_sor_code=sor,
                    target_sor_name=target_name
                )
            
            if not c_json or not a_json or str(c_json).strip() == "" or str(a_json).strip() == "":
                # Fallback to defaults
                rule_data_dummy = {
                    "equipment_type": eq_type,
                    "action_filter": act,
                    "qty_formula": "table_qty",
                    "internal_id": internal_id,
                    "target_sor_code": sor,
                    "target_sor_name": target_name
                }
                c_json, a_json = build_default_venmo_conditions_and_actions(rule_data_dummy)
                
            from services.venmo_engine import generate_plain_english_logic
            logic_exp = generate_plain_english_logic(c_json, a_json)
 
            cursor.execute("""
                INSERT INTO mapping_rules (
                    rule_name, category, equipment_type, match_keywords, exclude_keywords,
                    condition_expr, action_filter, target_sor_code, target_sor_name, qty_formula,
                    comment_template, priority, enabled, internal_id, conditions_json, actions_json, logic_explanation,
                    primary_source, preferred_source_type, ignore_pages, duplicate_prone_pages, matching_conditions, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                rule_name, category_val, eq_type, match_kws, exclude_kws,
                "", act, sor, target_name, "table_qty",
                "", priority, enabled, internal_id, c_json, a_json, logic_exp,
                "", pref_src_type, ignore_p, dup_prone_p, matching_cond, notes_val
            ))
            imported += 1

        conn.commit()
        conn.close()
        wb.close()
        return (imported, 0)
    except Exception as e:
        print(f"[DB] Error importing mapping rules from Excel: {e}")
        return (0, 0)


def get_confidence_thresholds() -> Dict[str, float]:
    """Retrieves the configurable confidence thresholds from system settings."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT key, value FROM system_settings WHERE key LIKE 'confidence_%'")
    rows = cursor.fetchall()
    conn.close()

    thresholds = {
        "confidence_auto_approve": 90.0,
        "confidence_review_required": 70.0,
        "confidence_unquoted_flag": 50.0
    }
    for row in rows:
        k = row["key"]
        try:
            thresholds[k] = float(row["value"])
        except (ValueError, TypeError):
            pass
    return thresholds


def update_confidence_thresholds(thresholds: Dict[str, float]) -> bool:
    """Updates the configurable confidence thresholds in system settings."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        for k, v in thresholds.items():
            if k in ["confidence_auto_approve", "confidence_review_required", "confidence_unquoted_flag"]:
                cursor.execute(
                    "INSERT INTO system_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                    (k, str(float(v)))
                )
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"[DB] Error updating confidence thresholds: {e}")
        return False


DEFAULT_PARSER_CONFIGS = [
    {
        "config_key": "antenna_id_regex",
        "category": "REGEX",
        "name": "Antenna Position ID Pattern",
        "pattern_value": r"\b(A[0-9]{1,2}(?:\s*\(OLD\))?)\b",
        "description": "Regex pattern used to extract antenna position IDs (e.g. A1, A5, A12) from drawing tables and notes."
    },
    {
        "config_key": "note_callout_qty_regex",
        "category": "REGEX",
        "name": "Drawing Note Stated Quantity Pattern",
        "pattern_value": r"\((\d+)\s*OFF(?:\s+([A-Z0-9,\s&]+))?\)",
        "description": "Regex pattern used to extract stated quantities and referenced antenna IDs from drawing notes (e.g. (3 OFF A1, A5 & A9))."
    },
    {
        "config_key": "dimension_3d_regex",
        "category": "REGEX",
        "name": "3D Physical Dimension Pattern",
        "pattern_value": r"(\d+)\s*(?:MM)?\s*[Xx*]\s*(\d+)\s*(?:MM)?\s*[Xx*]\s*(\d+)",
        "description": "Regex pattern used to parse physical dimensions (Height x Width x Depth in mm)."
    },
    {
        "config_key": "dimension_2d_regex",
        "category": "REGEX",
        "name": "2D Physical Dimension Pattern",
        "pattern_value": r"(\d+)\s*(?:MM)?\s*[Xx*]\s*(\d+)",
        "description": "Regex pattern used to parse physical dimensions (Height x Width in mm)."
    },
    {
        "config_key": "sector_pattern_regex",
        "category": "REGEX",
        "name": "Sector Designation Pattern",
        "pattern_value": r"\b(S[1-4]|SEC(?:TOR)?\s*[1-4]|ALPHA|BETA|GAMMA)\b",
        "description": "Regex pattern used to identify sector designations from sheet and table strings."
    },
    {
        "config_key": "hardware_exclusion_keywords",
        "category": "KEYWORDS",
        "name": "Structural Hardware Exclusion Keywords",
        "pattern_value": "MOUNT, BRACKET, U-BOLT, STAND-OFF, PIPE, TUBE, CANTILEVER, CLAMP, BOOT, CABLE TIE, ADAPTOR, WASHER, NUT, HEADFRAME",
        "description": "Comma-separated list of hardware/steelwork keywords to ignore during equipment takeoff."
    },
    {
        "config_key": "patch_panel_keywords",
        "category": "KEYWORDS",
        "name": "Patch Panel & Distribution Keywords",
        "pattern_value": "PATCH PANEL, FIBRE PANEL, FIBER PANEL, DISTRIBUTION PANEL, ODF, BREAKOUT",
        "description": "Comma-separated list of keywords identifying shelter distribution and termination panels."
    },
    {
        "config_key": "active_5g_keywords",
        "category": "KEYWORDS",
        "name": "Active 5G AAU Detection Keywords",
        "pattern_value": "AIR, AAU, MASSIVE MIMO, ACTIVE ANTENNA, AIR6488, AIR3268, AIR1641",
        "description": "Comma-separated list of keywords identifying active 5G beamforming antenna units."
    },
    {
        "config_key": "antenna_table_titles",
        "category": "TABLE_TITLES",
        "name": "Antenna Configuration Table Titles",
        "pattern_value": "TELSTRA MOBILES ANTENNA CONFIGURATION TABLE, ANTENNA CONFIGURATION TABLE, ANTENNA SCHEDULE, ANTENNA SUMMARY",
        "description": "Comma-separated list of table titles identifying master antenna configuration schedules."
    },
    {
        "config_key": "equipment_schedule_titles",
        "category": "TABLE_TITLES",
        "name": "Equipment Schedule Table Titles",
        "pattern_value": "EQUIPMENT SCHEDULE, EQUIPMENT SUMMARY, RADIO EQUIPMENT SCHEDULE, RADIO SCHEDULE",
        "description": "Comma-separated list of table titles identifying radio and baseband equipment schedules."
    }
]


def seed_default_parser_configs(cursor):
    """Seeds default parser configurations into parser_configs SQLite table."""
    for cfg in DEFAULT_PARSER_CONFIGS:
        cursor.execute("""
            INSERT OR IGNORE INTO parser_configs (config_key, category, name, pattern_value, description, is_active)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (
            cfg["config_key"],
            cfg["category"],
            cfg["name"],
            cfg["pattern_value"],
            cfg.get("description", "")
        ))


def get_all_parser_configs() -> List[Dict[str, Any]]:
    """Fetches all drawing parsing configs from SQLite."""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM parser_configs ORDER BY category ASC, id ASC")
    rows = cursor.fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_parser_config_map() -> Dict[str, str]:
    """
    Returns a dictionary mapping config_key -> pattern_value for active configs.
    Falls back to DEFAULT_PARSER_CONFIGS defaults if any key is missing.
    """
    config_map = {cfg["config_key"]: cfg["pattern_value"] for cfg in DEFAULT_PARSER_CONFIGS}
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT config_key, pattern_value FROM parser_configs WHERE is_active = 1")
        for row in cursor.fetchall():
            config_map[row["config_key"]] = row["pattern_value"]
        conn.close()
    except Exception as e:
        print(f"[DB] Warning: Failed to load parser_configs from DB: {e}")
    return config_map


def update_parser_config(config_id: int, pattern_value: str, is_active: int = 1) -> bool:
    """Updates a parser configuration pattern and active state."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE parser_configs SET pattern_value = ?, is_active = ? WHERE id = ?
        """, (pattern_value, is_active, config_id))
        affected = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return affected
    except Exception as e:
        print(f"[DB] Error updating parser config #{config_id}: {e}")
        return False


def reset_default_parser_configs() -> bool:
    """Resets all parser configurations to default Telstra patterns."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM parser_configs")
        seed_default_parser_configs(cursor)
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"[DB] Error resetting parser configs: {e}")
        return False


